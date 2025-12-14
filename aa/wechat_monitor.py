import os
import re
import time
import random
import threading
import sqlite3
import sys
from pathlib import Path
from ctypes import windll, c_int, byref

# PySide2 imports
from PySide2.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                               QLabel, QPushButton, QTextEdit, QProgressBar, QFileDialog,
                               QMessageBox, QGroupBox, QGridLayout, QLineEdit, QFrame)
from PySide2.QtCore import Qt, QThread, Signal, QTimer, QSize
from PySide2.QtGui import QFont, QIcon, QPalette, QColor, QTextCursor

# 其他库导入
import pyautogui
import pyperclip
import mss
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# ===================================== 全局配置（可修改）=====================================
DEFAULT_SAVE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "头像")  # 截图保存路径
EXCEL_FILE_NAME = "微信数据.xlsx"  # Excel表格文件名
DB_FILE_NAME = "wechat_monitor.db"  # 数据库文件名
PREPARE_TIME = 5  # 准备时间（秒）
OPERATE_DELAY_MIN = 0.3  # 操作间隔最小值（秒）
OPERATE_DELAY_MAX = 0.8  # 操作间隔最大值（秒）
CELL_PADDING = 5  # 图片与单元格边框的内边距（像素）

# 拟人化配置
MOUSE_MOVE_DURATION_MIN = 0.2  # 鼠标移动时间最小值（秒）
MOUSE_MOVE_DURATION_MAX = 0.5  # 鼠标移动时间最大值（秒）
INPUT_INTERVAL_MIN = 0.03  # 输入字符间隔最小值（秒）
INPUT_INTERVAL_MAX = 0.1  # 输入字符间隔最大值（秒）
BATCH_REST_COUNT = 5  # 每处理N个手机号后休息
BATCH_REST_MIN = 3  # 批量休息最小值（秒）
BATCH_REST_MAX = 6  # 批量休息最大值（秒）
CLICK_OFFSET_RANGE = 5  # 点击坐标随机偏移范围（像素）

# 表格样式配置（专业风）
HEADER_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")  # 表头字体：白色14号加粗
CONTENT_FONT = Font(name="微软雅黑", size=11, bold=True, color="333333")  # 正文字体：深灰11号加粗
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 表头背景：深蓝色
BORDER = Border(  # 表格边框：细黑边
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# 坐标配置（改为「目标区域」而非固定点，容错率更高）
COORDS = {
    "input_activate": (1250, 625, 1260, 635),  # 左上x,左上y,右下x,右下y（区域）
    "after_enter": (1150, 670, 1155, 675),  # 回车后点击区域
    "copy_double_click": (1170, 765, 1180, 770),  # 双击复制区域
    "screenshot_activate": (1420, 720, 1428, 728),  # 截图前激活区域
    "fail_click": (1385, 745, 1393, 750),  # 复制失败时点击区域
    "capture_left_top": (1275, 783),  # 截图左上角（固定）
    "capture_right_bottom": (1561, 1067)  # 截图右下角（固定）
}

# 隐藏工具特征：禁用pyautogui的失败安全（避免触发弹窗）
pyautogui.FAILSAFE = False


# ===================================== 数据库工具类=====================================
class Database:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        self.init_db()  # 初始化数据库（创建表）

    def connect(self):
        """连接数据库"""
        try:
            self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
            self.cursor = self.conn.cursor()
            return True
        except Exception as e:
            print(f"数据库连接失败：{str(e)}")
            return False

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            self.conn = None
            self.cursor = None

    def init_db(self):
        """初始化数据库表结构"""
        if not self.connect():
            return

        # 1. 业务数据表：存储手机号、微信名、头像路径
        self.cursor.execute('''
                            CREATE TABLE IF NOT EXISTS phone_data
                            (
                                phone
                                TEXT
                                PRIMARY
                                KEY
                                NOT
                                NULL, -- 手机号（主键，唯一）
                                wechat_name
                                TEXT
                                NOT
                                NULL, -- 微信名
                                avatar_path
                                TEXT
                                NOT
                                NULL, -- 头像路径
                                create_time
                                TEXT
                                NOT
                                NULL  -- 处理时间
                            )
                            ''')

        # 2. 进度表：记录每个TXT文件的最后处理行数（断点续读用）
        self.cursor.execute('''
                            CREATE TABLE IF NOT EXISTS process_progress
                            (
                                txt_path
                                TEXT
                                PRIMARY
                                KEY
                                NOT
                                NULL, -- TXT文件路径（主键，唯一标识文件）
                                last_line
                                INT
                                NOT
                                NULL
                                DEFAULT
                                0,    -- 最后处理的行号（从0开始）
                                update_time
                                TEXT
                                NOT
                                NULL  -- 最后更新时间
                            )
                            ''')

        self.conn.commit()
        self.close()

    def get_last_process_line(self, txt_path: str) -> int:
        """获取指定TXT文件的最后处理行号（断点续读）"""
        if not self.connect():
            return 0

        try:
            self.cursor.execute('''
                                SELECT last_line
                                FROM process_progress
                                WHERE txt_path = ?
                                ''', (txt_path,))
            result = self.cursor.fetchone()
            return result[0] if result else 0
        except Exception as e:
            print(f"获取进度失败：{str(e)}")
            return 0
        finally:
            self.close()

    def update_process_line(self, txt_path: str, last_line: int):
        """更新指定TXT文件的最后处理行号"""
        if not self.connect():
            return

        try:
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            # 存在则更新，不存在则插入
            self.cursor.execute('''
                INSERT OR REPLACE INTO process_progress 
                (txt_path, last_line, update_time) 
                VALUES (?, ?, ?)
            ''', (txt_path, last_line, current_time))
            self.conn.commit()
        except Exception as e:
            print(f"更新进度失败：{str(e)}")
        finally:
            self.close()

    def insert_phone_data(self, phone: str, wechat_name: str, avatar_path: str):
        """插入业务数据到数据库"""
        if not self.connect():
            return

        try:
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            # 存在则更新（避免重复插入），不存在则插入
            self.cursor.execute('''
                INSERT OR REPLACE INTO phone_data 
                (phone, wechat_name, avatar_path, create_time) 
                VALUES (?, ?, ?, ?)
            ''', (phone, wechat_name, avatar_path, current_time))
            self.conn.commit()
        except Exception as e:
            print(f"插入数据失败：{str(e)}")
        finally:
            self.close()

    def is_phone_processed(self, phone: str) -> bool:
        """判断手机号是否已处理过（避免重复）"""
        if not self.connect():
            return False

        try:
            self.cursor.execute('''
                                SELECT 1
                                FROM phone_data
                                WHERE phone = ? LIMIT 1
                                ''', (phone,))
            return self.cursor.fetchone() is not None
        except Exception as e:
            print(f"查询手机号状态失败：{str(e)}")
            return False
        finally:
            self.close()


# ===================================== 工具函数（拟人化优化版）=====================================
def is_11_digit(text: str) -> bool:
    """验证是否为11位纯数字"""
    return re.fullmatch(r"\d{11}", text.strip()) is not None


def replace_invalid_filename_chars(text: str) -> str:
    """替换文件名中的非法字符"""
    invalid_chars = r'[\\/:*?"<>|]'
    return re.sub(invalid_chars, "_", text)


def create_folder_if_not_exist(folder_path: str):
    """创建文件夹（不存在则创建）"""
    Path(folder_path).mkdir(parents=True, exist_ok=True)
    return folder_path


def capture_screen(save_path: str) -> bool:
    """截取指定区域并保存图片（降低截图频率，优化参数）"""
    try:
        x1, y1 = COORDS["capture_left_top"]
        x2, y2 = COORDS["capture_right_bottom"]
        with mss.mss() as sct:
            monitor = {
                "top": y1, "left": x1,
                "width": x2 - x1, "height": y2 - y1
            }
            sct_img = sct.grab(monitor)
            # 优化：降低图片质量，减少特征
            mss.tools.to_png(sct_img.rgb, sct_img.size, output=save_path, compress_level=6)
        # 截图后随机停顿
        time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))
        return True
    except Exception as e:
        print(f"截图失败：{str(e)}")
        return False


def clear_input_box() -> None:
    """清空输入框（模拟人类可能的重复操作）"""
    # 随机选择清空方式：Ctrl+A+Delete 或 多次Backspace
    if random.random() > 0.3:
        pyautogui.keyDown("ctrl")
        pyautogui.press("a")
        pyautogui.keyUp("ctrl")
        time.sleep(random.uniform(0.1, 0.2))
        pyautogui.press("delete")
    else:
        pyautogui.keyDown("ctrl")
        pyautogui.press("a")
        pyautogui.keyUp("ctrl")
        time.sleep(random.uniform(0.1, 0.2))
        pyautogui.press("backspace")
        # 偶尔多按一次Backspace（人类失误）
        if random.random() > 0.7:
            time.sleep(0.1)
            pyautogui.press("backspace")
    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))


def get_random_point_in_area(area: tuple) -> tuple:
    """在目标区域内获取随机坐标（避免固定点点击）"""
    x1, y1, x2, y2 = area
    x = random.randint(x1, x2)
    y = random.randint(y1, y2)
    return (x, y)


def human_like_move_click(area: tuple) -> None:
    """拟人化移动并点击（随机路径、随机偏移、模拟犹豫）"""
    # 1. 获取目标区域内随机点
    target_x, target_y = get_random_point_in_area(area)
    # 2. 随机偏移目标点（模拟人类点击不准）
    offset_x = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    offset_y = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    target_x += offset_x
    target_y += offset_y
    # 3. 模拟人类鼠标移动曲线（非匀速）
    move_duration = random.uniform(MOUSE_MOVE_DURATION_MIN, MOUSE_MOVE_DURATION_MAX)
    pyautogui.moveTo(target_x, target_y, duration=move_duration, tween=pyautogui.easeInOutQuad)
    # 4. 点击前随机停顿（犹豫）
    time.sleep(random.uniform(0.05, 0.2))
    # 5. 偶尔双击（人类操作误差）
    if random.random() > 0.85:
        pyautogui.doubleClick()
    else:
        pyautogui.click()
    # 6. 点击后随机停顿
    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))


def human_like_double_click(area: tuple) -> None:
    """拟人化双击（避免机械双击）"""
    target_x, target_y = get_random_point_in_area(area)
    offset_x = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    offset_y = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    target_x += offset_x
    target_y += offset_y
    move_duration = random.uniform(MOUSE_MOVE_DURATION_MIN, MOUSE_MOVE_DURATION_MAX)
    pyautogui.moveTo(target_x, target_y, duration=move_duration, tween=pyautogui.easeInOutQuad)
    time.sleep(random.uniform(0.05, 0.15))
    # 双击间隔随机
    pyautogui.click()
    time.sleep(random.uniform(0.03, 0.08))
    pyautogui.click()
    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))


def copy_selected_content() -> str:
    """复制选中内容（模拟人类可能的重复复制）"""
    pyperclip.copy("")
    time.sleep(random.uniform(0.05, 0.15))

    # 随机选择复制方式：Ctrl+C 或 右键复制（降低快捷键频率）
    if random.random() > 0.4:
        pyautogui.keyDown("ctrl")
        # 偶尔多按一次C（人类失误）
        if random.random() > 0.8:
            pyautogui.press("c")
            time.sleep(0.05)
        pyautogui.press("c")
        pyautogui.keyUp("ctrl")
    else:
        pyautogui.rightClick()
        time.sleep(random.uniform(0.1, 0.2))
        pyautogui.press("c")  # 右键菜单中选择复制（需确保菜单位置正确）

    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))
    content = pyperclip.paste().strip()
    return content if content else ""


def human_like_typewrite(text: str) -> None:
    """拟人化输入（随机间隔、偶尔回退重输）"""
    for char in text:
        pyautogui.press(char)
        # 输入间隔随机
        time.sleep(random.uniform(INPUT_INTERVAL_MIN, INPUT_INTERVAL_MAX))
        # 10%概率回退重输（人类输入错误修正）
        if random.random() < 0.1 and len(text) > 3:
            time.sleep(random.uniform(0.1, 0.2))
            pyautogui.press("backspace")
            time.sleep(random.uniform(0.05, 0.1))
            pyautogui.press(char)
            time.sleep(random.uniform(INPUT_INTERVAL_MIN, INPUT_INTERVAL_MAX))


def random_rest() -> None:
    """随机休息（模拟人类走神）"""
    if random.random() > 0.7:
        rest_time = random.uniform(0.5, 1.5)
        time.sleep(rest_time)


def batch_rest(count: int, log_func) -> None:
    """批量处理后休息（避免高强度操作）"""
    if count % BATCH_REST_COUNT == 0 and count != 0:
        rest_time = random.uniform(BATCH_REST_MIN, BATCH_REST_MAX)
        log_func(f"批量处理{count}个，休息{rest_time:.1f}秒...")
        time.sleep(rest_time)


# ===================================== Excel表格生成函数（图片嵌入版）=====================================
class ExcelExporter:
    def __init__(self, save_path: str):
        self.save_path = os.path.join(save_path, EXCEL_FILE_NAME)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "微信数据"
        self.init_table_style()

    def init_table_style(self):
        """初始化表格样式（专业美化+嵌入适配）"""
        # 列名：手机号、微信名、头像
        headers = ["手机号", "微信名", "头像"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            # 表头样式：加大加粗、白色文字、蓝色背景、居中对齐、边框
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

            # 调整列宽（适配内容和嵌入图片）
            if col == 1:  # 手机号列（11位数字+边距）
                self.ws.column_dimensions[get_column_letter(col)].width = 18
            elif col == 2:  # 微信名列（适配较长昵称）
                self.ws.column_dimensions[get_column_letter(col)].width = 28
            elif col == 3:  # 头像列（适配图片+内边距，1列≈8px）
                self.ws.column_dimensions[get_column_letter(col)].width = 12  # 12*8=96px

        # 表头行高（适配14号字体）
        self.ws.row_dimensions[1].height = 35

        # 设置工作表默认字体
        self.ws.font = CONTENT_FONT

    def add_row(self, phone: str, wechat_name: str, avatar_path: str):
        """添加一行数据（图片嵌入单元格）"""
        next_row = self.ws.max_row + 1

        # 头像列单元格尺寸（用于图片适配）
        cell_width_px = self.ws.column_dimensions["C"].width * 8  # 1列≈8px
        cell_height_px = 70  # 固定行高，适配图片嵌入
        self.ws.row_dimensions[next_row].height = cell_height_px

        # 1. 手机号列：加粗、居中、边框
        phone_cell = self.ws.cell(row=next_row, column=1, value=phone)
        phone_cell.alignment = Alignment(horizontal="center", vertical="center")
        phone_cell.border = BORDER

        # 2. 微信名列：加粗、居中、边框、无昵称显示
        wechat_name = wechat_name if wechat_name.strip() else "无昵称"
        wechat_cell = self.ws.cell(row=next_row, column=2, value=wechat_name)
        wechat_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wechat_cell.border = BORDER

        # 3. 头像列：嵌入图片+边框
        avatar_cell = self.ws.cell(row=next_row, column=3, value="")
        avatar_cell.alignment = Alignment(horizontal="center", vertical="center")
        avatar_cell.border = BORDER

        # 有头像时嵌入单元格（图片尺寸适配单元格，带内边距）
        if avatar_path and avatar_path.strip() and os.path.exists(avatar_path):
            try:
                # 计算图片最大可用尺寸（单元格尺寸 - 2*内边距）
                max_img_width = cell_width_px - 2 * CELL_PADDING
                max_img_height = cell_height_px - 2 * CELL_PADDING

                # 插入图片并按单元格尺寸等比例缩放
                img = Image(avatar_path)
                # 等比例缩放，确保图片不超出单元格
                img_ratio = img.width / img.height
                max_ratio = max_img_width / max_img_height

                if img_ratio > max_ratio:
                    # 按宽度适配
                    img.width = max_img_width
                    img.height = max_img_width / img_ratio
                else:
                    # 按高度适配
                    img.height = max_img_height
                    img.width = max_img_height * img_ratio

                # 绑定到单元格，实现嵌入效果（随单元格移动）
                img.anchor = f"C{next_row}"

                # 设置图片位置，实现带内边距的居中
                img.left = int(CELL_PADDING * 9525)  # 内边距，EMU单位（1px≈9525）
                img.top = int(CELL_PADDING * 9525)

                self.ws.add_image(img)
            except Exception as e:
                print(f"插入图片失败：{str(e)}")
                # 失败时仍保持单元格空状态，不显示错误文字

    def save(self):
        """保存Excel文件（处理权限问题）"""
        try:
            # 先检查文件是否被占用
            if os.path.exists(self.save_path):
                with open(self.save_path, "rb") as f:
                    pass  # 能打开说明未被占用

            self.wb.save(self.save_path)
            return self.save_path
        except PermissionError:
            return None
        except Exception as e:
            print(f"Excel保存失败：{str(e)}")
            return None


# ===================================== 工作线程 =====================================
class MonitorThread(QThread):
    """监控工作线程"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)
    update_button_signal = Signal(bool)  # True: 开始执行，False: 停止执行

    def __init__(self, txt_file_path, save_folder, db, prepare_time,
                 operate_delay_min, operate_delay_max, start_line=0):
        super().__init__()
        self.txt_file_path = txt_file_path
        self.save_folder = save_folder
        self.db = db
        self.prepare_time = prepare_time
        self.operate_delay_min = operate_delay_min
        self.operate_delay_max = operate_delay_max
        self.start_line = start_line

        self.is_running = True
        self.excel_exporter = None

    def run(self):
        try:
            self.update_button_signal.emit(True)

            # 读取TXT文件
            with open(self.txt_file_path, "r", encoding="utf-8") as f:
                all_lines = f.readlines()

            total_lines = len(all_lines)
            self.log_signal.emit(f"成功读取TXT文件，总行数：{total_lines}")

            if self.start_line > 0:
                self.log_signal.emit(f"断点续读：从第{self.start_line}行开始处理（前{self.start_line}行已跳过）")

            if self.start_line >= total_lines:
                self.finished_signal.emit(True, "该文件已全部处理完成，无需重复执行！")
                return

            # 筛选有效手机号
            valid_tasks = []
            for line_num in range(self.start_line, total_lines):
                if not self.is_running:
                    break
                line_content = all_lines[line_num].strip()
                if is_11_digit(line_content) and not self.db.is_phone_processed(line_content):
                    valid_tasks.append((line_num, line_content))

            total_valid = len(valid_tasks)
            if total_valid == 0:
                self.finished_signal.emit(True, "当前文件无未处理的有效手机号！")
                return

            self.log_signal.emit(f"筛选出未处理的有效手机号：{total_valid}个")

            # 初始化Excel
            excel_path = os.path.join(self.save_folder, EXCEL_FILE_NAME)
            if os.path.exists(excel_path):
                self.excel_exporter = ExcelExporter(self.save_folder)
                self.excel_exporter.wb = load_workbook(excel_path)
                self.excel_exporter.ws = self.excel_exporter.wb.active
                self.log_signal.emit("Excel文件已存在，将追加数据（图片嵌入格式）")
            else:
                self.excel_exporter = ExcelExporter(self.save_folder)
                self.log_signal.emit("Excel表格初始化完成（专业美化+图片嵌入）")

            # 倒计时
            self.log_signal.emit(f"准备倒计时：{self.prepare_time}秒，请打开目标窗口...")
            for i in range(self.prepare_time, 0, -1):
                if not self.is_running:
                    break
                self.log_signal.emit(f"倒计时：{i}秒")
                time.sleep(1)

            if not self.is_running:
                self.finished_signal.emit(False, "执行已停止")
                return

            self.log_signal.emit("倒计时结束，开始执行操作！")

            # 执行任务
            processed_count = 0
            for line_num, phone in valid_tasks:
                if not self.is_running:
                    break

                processed_count += 1
                self.log_signal.emit(
                    f"\n===== 正在处理第{processed_count}/{total_valid}个手机号（行号：{line_num}）：{phone} ====="
                )

                try:
                    self._process_phone(line_num, phone, processed_count)
                except Exception as e:
                    error_msg = f"处理失败：{str(e)}"
                    self.log_signal.emit(error_msg)
                    self.excel_exporter.add_row(phone, "", "")
                    self.db.insert_phone_data(phone, error_msg, "处理失败")
                    self.db.update_process_line(self.txt_file_path, line_num)
                    time.sleep(random.uniform(1, 2))

                self.progress_signal.emit(processed_count, total_valid)

                # 操作间隔
                if processed_count < total_valid:
                    time.sleep(random.uniform(
                        self.operate_delay_min * 1.5,
                        self.operate_delay_max * 1.5
                    ))

            # 保存Excel
            if self.excel_exporter:
                excel_path = self.excel_exporter.save()
                if excel_path:
                    self.log_signal.emit(f"Excel表格保存成功：{excel_path}（图片嵌入+专业美化）")
                else:
                    self.log_signal.emit("Excel表格保存失败！")

            if self.is_running:
                self.finished_signal.emit(
                    True,
                    f"任务处理完成！\n共处理{processed_count}个手机号\n存储路径：{self.save_folder}"
                )
            else:
                self.finished_signal.emit(False, "执行已停止，当前进度已保存到数据库")

        except Exception as e:
            self.log_signal.emit(f"执行过程发生错误：{str(e)}")
            self.finished_signal.emit(False, f"执行失败：{str(e)}")
        finally:
            self.update_button_signal.emit(False)

    def _process_phone(self, line_num, phone, processed_count):
        """处理单个手机号"""
        avatar_path = ""
        wechat_name = ""

        # 模拟人类随机休息
        random_rest()

        # 1. 激活输入框 + 清空内容
        human_like_move_click(COORDS["input_activate"])
        self.log_signal.emit("激活输入框并清空内容")
        clear_input_box()

        # 2. 输入手机号 + 回车
        human_like_typewrite(phone)
        time.sleep(random.uniform(self.operate_delay_min, self.operate_delay_max))
        time.sleep(random.uniform(0.1, 0.3))
        pyautogui.press("enter")
        self.log_signal.emit("输入手机号并回车")
        time.sleep(random.uniform(self.operate_delay_min * 2, self.operate_delay_max * 2))

        # 3. 点击坐标2
        human_like_move_click(COORDS["after_enter"])
        self.log_signal.emit("点击坐标2")

        # 4. 双击复制微信名
        human_like_double_click(COORDS["copy_double_click"])
        wechat_name = copy_selected_content()
        wechat_name = replace_invalid_filename_chars(wechat_name) if wechat_name else ""
        if wechat_name:
            self.log_signal.emit(f"获取微信名：{wechat_name}")
        else:
            self.log_signal.emit("未获取到微信名（无昵称）")

        # 5. 判断是否复制成功
        if wechat_name:
            if random.random() > 0.6:
                random_x = random.randint(100, 300)
                random_y = random.randint(200, 400)
                pyautogui.moveTo(random_x, random_y, duration=random.uniform(0.2, 0.4))
                time.sleep(random.uniform(0.1, 0.2))
                human_like_move_click(COORDS["screenshot_activate"])
            else:
                human_like_move_click(COORDS["screenshot_activate"])

            self.log_signal.emit("点击坐标4，准备截图")
            avatar_filename = f"{phone}_{wechat_name}.png"
            avatar_path = os.path.join(self.save_folder, avatar_filename)
            if capture_screen(avatar_path):
                self.log_signal.emit(f"截图保存成功：{avatar_filename}")
            else:
                self.log_signal.emit("截图失败！")
                avatar_path = ""
        else:
            human_like_move_click(COORDS["fail_click"])
            self.log_signal.emit("未获取到微信名，点击备用坐标，跳过截图")
            avatar_path = ""

        # 6. 保存数据
        db_wechat_name = wechat_name if wechat_name else "无昵称"
        db_avatar_path = avatar_path if avatar_path else "无截图"
        self.excel_exporter.add_row(phone, wechat_name, avatar_path)
        self.db.insert_phone_data(phone, db_wechat_name, db_avatar_path)
        self.log_signal.emit("数据已保存到Excel（图片嵌入）和数据库")

        # 7. 更新进度
        self.db.update_process_line(self.txt_file_path, line_num)
        self.log_signal.emit(f"进度更新：已处理到第{line_num}行")

        # 8. 批量休息
        batch_rest(processed_count, self.log_signal.emit)

    def stop(self):
        """停止线程"""
        self.is_running = False


# ===================================== 主窗口类 =====================================
class WeChatMonitorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("微信头像工具 v3.0 - 专业商务版")
        self.setGeometry(100, 100, 1200, 800)

        # 初始化变量
        self.txt_file_path = None
        self.save_folder = DEFAULT_SAVE_FOLDER
        self.db = None
        self.monitor_thread = None
        self.total_valid_lines = 0
        self.processed_count = 0

        # 初始化数据库
        self.init_db()

        # 设置样式
        self.setup_style()

        # 初始化UI
        self.init_ui()

    def setup_style(self):
        """设置应用程序样式"""
        # 设置字体以确保中文显示正常
        font = QFont("微软雅黑", 10)
        QApplication.setFont(font)

        # 设置样式表
        style = """
        QMainWindow {
            background-color: #f5f7fa;
        }

        QGroupBox {
            font-weight: bold;
            font-size: 12px;
            border: 1px solid #d1d9e6;
            border-radius: 8px;
            margin-top: 10px;
            padding-top: 10px;
            background-color: white;
        }

        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
            color: #2c3e50;
        }

        QPushButton {
            background-color: #3498db;
            color: white;
            border: none;
            border-radius: 4px;
            padding: 8px 16px;
            font-weight: bold;
            font-size: 11px;
            min-width: 80px;
        }

        QPushButton:hover {
            background-color: #2980b9;
        }

        QPushButton:pressed {
            background-color: #21618c;
        }

        QPushButton:disabled {
            background-color: #bdc3c7;
            color: #7f8c8d;
        }

        QPushButton#startBtn {
            background-color: #2ecc71;
        }

        QPushButton#startBtn:hover {
            background-color: #27ae60;
        }

        QPushButton#stopBtn {
            background-color: #e74c3c;
        }

        QPushButton#stopBtn:hover {
            background-color: #c0392b;
        }

        QLineEdit {
            border: 1px solid #d1d9e6;
            border-radius: 4px;
            padding: 6px;
            font-size: 11px;
        }

        QLineEdit:focus {
            border: 1px solid #3498db;
        }

        QProgressBar {
            border: 1px solid #d1d9e6;
            border-radius: 4px;
            text-align: center;
            font-size: 11px;
            height: 20px;
        }

        QProgressBar::chunk {
            background-color: #3498db;
            border-radius: 3px;
        }

        QTextEdit {
            border: 1px solid #d1d9e6;
            border-radius: 4px;
            font-family: "Consolas", "微软雅黑";
            font-size: 10px;
            background-color: white;
        }

        QLabel {
            color: #2c3e50;
            font-size: 11px;
        }

        QLabel#titleLabel {
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
        }

        QLabel#statusLabel {
            color: #7f8c8d;
            font-size: 10px;
            font-style: italic;
        }
        """
        self.setStyleSheet(style)

    def init_db(self):
        """初始化数据库"""
        create_folder_if_not_exist(self.save_folder)
        db_path = os.path.join(self.save_folder, DB_FILE_NAME)
        self.db = Database(db_path)

    def init_ui(self):
        """初始化用户界面"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 标题区域
        title_layout = QHBoxLayout()
        title_label = QLabel("微信头像工具 v3.0")
        title_label.setObjectName("titleLabel")
        title_layout.addWidget(title_label)
        title_layout.addStretch()

        status_label = QLabel("专业商务版 · 防检测优化")
        status_label.setObjectName("statusLabel")
        title_layout.addWidget(status_label)

        main_layout.addLayout(title_layout)

        # 文件选择区域
        file_group = QGroupBox("文件配置")
        file_layout = QGridLayout(file_group)

        self.file_label = QLabel("未选择TXT文件")
        self.file_label.setStyleSheet("color: #7f8c8d; font-style: italic;")

        file_layout.addWidget(QLabel("TXT文件:"), 0, 0)
        file_layout.addWidget(self.file_label, 0, 1)

        select_btn = QPushButton("选择文件")
        select_btn.clicked.connect(self.select_txt_file)
        file_layout.addWidget(select_btn, 0, 2)

        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet("color: #3498db; font-size: 10px;")
        file_layout.addWidget(self.progress_label, 1, 0, 1, 3)

        main_layout.addWidget(file_group)

        # 参数配置区域
        param_group = QGroupBox("执行参数")
        param_layout = QGridLayout(param_group)

        param_layout.addWidget(QLabel("准备时间:"), 0, 0)
        self.prepare_time_edit = QLineEdit(str(PREPARE_TIME))
        self.prepare_time_edit.setFixedWidth(80)
        param_layout.addWidget(self.prepare_time_edit, 0, 1)
        param_layout.addWidget(QLabel("秒"), 0, 2)

        param_layout.addWidget(QLabel("操作间隔:"), 0, 3)
        self.delay_min_edit = QLineEdit(str(OPERATE_DELAY_MIN))
        self.delay_min_edit.setFixedWidth(60)
        param_layout.addWidget(self.delay_min_edit, 0, 4)
        param_layout.addWidget(QLabel("-"), 0, 5)
        self.delay_max_edit = QLineEdit(str(OPERATE_DELAY_MAX))
        self.delay_max_edit.setFixedWidth(60)
        param_layout.addWidget(self.delay_max_edit, 0, 6)
        param_layout.addWidget(QLabel("秒"), 0, 7)

        # 存储目录
        param_layout.addWidget(QLabel("存储目录:"), 1, 0)
        self.folder_label = QLabel(self.save_folder)
        self.folder_label.setStyleSheet("color: #7f8c8d; font-size: 10px;")
        param_layout.addWidget(self.folder_label, 1, 1, 1, 6)

        open_folder_btn = QPushButton("打开目录")
        open_folder_btn.clicked.connect(self.open_save_folder)
        param_layout.addWidget(open_folder_btn, 1, 7)

        main_layout.addWidget(param_group)

        # 日志区域
        log_group = QGroupBox("执行日志")
        log_layout = QVBoxLayout(log_group)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)

        main_layout.addWidget(log_group)

        # 进度区域
        progress_group = QGroupBox("执行进度")
        progress_layout = QVBoxLayout(progress_group)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        progress_layout.addWidget(self.progress_bar)

        stats_layout = QHBoxLayout()
        self.stats_label = QLabel("已处理: 0/0")
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()

        self.time_label = QLabel("运行时间: 00:00:00")
        stats_layout.addWidget(self.time_label)

        progress_layout.addLayout(stats_layout)
        main_layout.addWidget(progress_group)

        # 按钮区域
        button_layout = QHBoxLayout()

        self.start_btn = QPushButton("开始执行")
        self.start_btn.setObjectName("startBtn")
        self.start_btn.clicked.connect(self.start_execution)
        button_layout.addWidget(self.start_btn)

        self.stop_btn = QPushButton("停止执行")
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.clicked.connect(self.stop_execution)
        self.stop_btn.setEnabled(False)
        button_layout.addWidget(self.stop_btn)

        button_layout.addStretch()

        export_btn = QPushButton("导出Excel")
        export_btn.clicked.connect(self.export_excel)
        button_layout.addWidget(export_btn)

        button_layout.addStretch()

        exit_btn = QPushButton("退出")
        exit_btn.clicked.connect(self.close)
        button_layout.addWidget(exit_btn)

        main_layout.addLayout(button_layout)

        # 定时器用于更新时间
        self.start_time = None
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_time)

        # 连接数据库
        self.log_message("系统初始化完成", "info")
        self.log_message(f"默认保存路径: {self.save_folder}", "info")

    def select_txt_file(self):
        """选择TXT文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择TXT文件", "", "Text Files (*.txt);;All Files (*.*)"
        )

        if file_path:
            self.txt_file_path = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.file_label.setStyleSheet("color: #2c3e50; font-style: normal;")
            self.log_message(f"已选择TXT文件: {file_path}", "info")

            # 查询断点续读进度
            last_line = self.db.get_last_process_line(file_path)
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()

                total_lines = len(lines)
                valid_lines = [line.strip() for line in lines if is_11_digit(line.strip())]
                total_valid = len(valid_lines)

                # 计算已处理的有效行数
                processed_valid = 0
                for i in range(min(last_line, total_lines)):
                    if is_11_digit(lines[i].strip()):
                        processed_valid += 1

                remaining = total_valid - processed_valid
                self.progress_label.setText(
                    f"断点续读: 上次处理到第{last_line}行，已处理{processed_valid}个有效手机号，剩余{remaining}个"
                )
                self.log_message(f"文件分析: 共{total_lines}行，{total_valid}个有效手机号", "info")
            except Exception as e:
                self.log_message(f"读取文件信息失败: {str(e)}", "error")

    def open_save_folder(self):
        """打开保存目录"""
        create_folder_if_not_exist(self.save_folder)
        os.startfile(self.save_folder)

    def log_message(self, message, msg_type="info"):
        """添加日志消息"""
        timestamp = time.strftime("%H:%M:%S")

        if msg_type == "info":
            prefix = "[信息]"
            color = "#3498db"
        elif msg_type == "success":
            prefix = "[成功]"
            color = "#2ecc71"
        elif msg_type == "warning":
            prefix = "[警告]"
            color = "#f39c12"
        elif msg_type == "error":
            prefix = "[错误]"
            color = "#e74c3c"
        else:
            prefix = "[调试]"
            color = "#95a5a6"

        html = f'<span style="color:#7f8c8d;">[{timestamp}]</span> '
        html += f'<span style="color:{color}; font-weight:bold;">{prefix}</span> '
        html += f'<span style="color:#2c3e50;">{message}</span><br>'

        self.log_text.moveCursor(QTextCursor.End)
        self.log_text.insertHtml(html)
        self.log_text.moveCursor(QTextCursor.End)

    def update_progress(self, processed, total):
        """更新进度条"""
        self.processed_count = processed
        self.total_valid_lines = total

        if total > 0:
            progress = int((processed / total) * 100)
            self.progress_bar.setValue(progress)
            self.stats_label.setText(f"已处理: {processed}/{total} ({progress}%)")

    def update_time(self):
        """更新时间显示"""
        if self.start_time:
            elapsed = time.time() - self.start_time
            hours = int(elapsed // 3600)
            minutes = int((elapsed % 3600) // 60)
            seconds = int(elapsed % 60)
            self.time_label.setText(f"运行时间: {hours:02d}:{minutes:02d}:{seconds:02d}")

    def start_execution(self):
        """开始执行"""
        # 验证参数
        if not self.txt_file_path:
            QMessageBox.warning(self, "警告", "请先选择TXT文件！")
            return

        try:
            prepare_time = int(self.prepare_time_edit.text())
            delay_min = float(self.delay_min_edit.text())
            delay_max = float(self.delay_max_edit.text())

            if prepare_time < 0 or delay_min < 0 or delay_max < delay_min:
                raise ValueError
        except ValueError:
            QMessageBox.critical(self, "错误",
                                 "参数错误！\n准备时间需为非负整数\n操作间隔需满足：最小值≥0 且 最大值≥最小值！")
            return

        # 检查Excel文件是否被占用
        excel_path = os.path.join(self.save_folder, EXCEL_FILE_NAME)
        if os.path.exists(excel_path):
            try:
                with open(excel_path, "rb") as f:
                    pass
            except PermissionError:
                QMessageBox.critical(self, "错误",
                                     f"Excel文件正在被其他程序打开！\n请关闭 {os.path.basename(excel_path)} 后重试。")
                return

        # 获取断点续读起始行
        last_line = self.db.get_last_process_line(self.txt_file_path)

        # 确认开始
        confirm_msg = f"即将开始执行！\n准备时间：{prepare_time}秒\n操作间隔：{delay_min}-{delay_max}秒\n"
        if last_line > 0:
            confirm_msg += f"断点续读：将从第{last_line}行开始处理（跳过已处理内容）\n"

        reply = QMessageBox.question(self, "确认", confirm_msg + "\n是否继续？",
                                     QMessageBox.Yes | QMessageBox.No)

        if reply != QMessageBox.Yes:
            return

        # 创建保存文件夹
        create_folder_if_not_exist(self.save_folder)
        self.log_message(f"截图保存路径: {self.save_folder}", "info")

        # 禁用开始按钮，启用停止按钮
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)

        # 启动定时器
        self.start_time = time.time()
        self.timer.start(1000)  # 每秒更新一次

        # 创建工作线程
        self.monitor_thread = MonitorThread(
            self.txt_file_path,
            self.save_folder,
            self.db,
            prepare_time,
            delay_min,
            delay_max,
            last_line
        )

        # 连接信号
        self.monitor_thread.log_signal.connect(self.log_message)
        self.monitor_thread.progress_signal.connect(self.update_progress)
        self.monitor_thread.finished_signal.connect(self.on_execution_finished)
        self.monitor_thread.update_button_signal.connect(self.update_button_state)

        # 启动线程
        self.monitor_thread.start()
        self.log_message("监控线程已启动，请确保目标窗口已打开", "info")

    def stop_execution(self):
        """停止执行"""
        if self.monitor_thread and self.monitor_thread.isRunning():
            reply = QMessageBox.question(self, "确认",
                                         "是否停止当前执行？已处理进度会自动保存！",
                                         QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                self.log_message("正在停止执行...", "warning")
                self.monitor_thread.stop()
                self.stop_btn.setEnabled(False)

    def update_button_state(self, is_running):
        """更新按钮状态"""
        self.start_btn.setEnabled(not is_running)
        self.stop_btn.setEnabled(is_running)

    def on_execution_finished(self, success, message):
        """执行完成处理"""
        # 停止定时器
        self.timer.stop()

        # 更新按钮状态
        self.update_button_state(False)

        # 显示完成消息
        if success:
            self.log_message(message, "success")
            QMessageBox.information(self, "完成", message)
        else:
            self.log_message(message, "warning" if "停止" in message else "error")
            if "失败" in message:
                QMessageBox.warning(self, "警告", message)

    def export_excel(self):
        """导出Excel文件"""
        excel_path = os.path.join(self.save_folder, EXCEL_FILE_NAME)
        if os.path.exists(excel_path):
            try:
                os.startfile(excel_path)
                self.log_message(f"已打开Excel文件: {excel_path}", "info")
            except Exception as e:
                self.log_message(f"打开Excel文件失败: {str(e)}", "error")
        else:
            self.log_message("Excel文件不存在，请先执行监控任务", "warning")

    def closeEvent(self, event):
        """关闭窗口事件"""
        if self.monitor_thread and self.monitor_thread.isRunning():
            reply = QMessageBox.question(self, "确认",
                                         "监控任务正在运行中，确定要退出吗？",
                                         QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                if self.monitor_thread:
                    self.monitor_thread.stop()
                    self.monitor_thread.wait(2000)  # 等待2秒
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


# ===================================== 主程序入口 =====================================
def main():
    # 解决高DPI显示问题
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    # 隐藏鼠标移动痕迹
    try:
        user32 = windll.user32
        user32.SetSystemCursor(None, 0)
    except:
        pass

    # 创建应用程序
    app = QApplication(sys.argv)
    app.setApplicationName("微信头像工具")

    # 设置默认字体以确保中文显示
    font = QFont("Microsoft YaHei", 10)
    app.setFont(font)

    # 创建并显示主窗口
    window = WeChatMonitorGUI()
    window.show()

    # 运行应用程序
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()