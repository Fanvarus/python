import os
import configparser
import datetime
import sys
from typing import List, Optional
from collections import defaultdict

# Pyside2导入
from PySide2.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit,
    QPushButton, QTextEdit, QFileDialog, QMessageBox,
    QGroupBox, QVBoxLayout, QHBoxLayout, QSizePolicy,
    QProgressBar, QFrame, QSpacerItem, QSizePolicy
)
from PySide2.QtCore import Qt, QThread, Signal, Slot
from PySide2.QtGui import QFont, QColor

# Excel依赖导入
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import (
    Alignment, Font, Border, Side, PatternFill
)

# -------------------------- 全局配置 --------------------------
CONFIG_FILE = os.path.join(os.path.expanduser("~"), "头像汇总工具_config.ini")
config = configparser.ConfigParser()
REQUIRED_COLS = ["有效手机号", "法定代表人", "所属省份", "所属城市", "所属区县"]
DESKTOP_PATH = os.path.expanduser("~\\Desktop")


# -------------------------- 工具函数 --------------------------
def init_config():
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE, encoding="utf-8")
        for key in ["single_query", "multi_queries", "source", "output"]:
            if key not in config["PATH"]:
                config["PATH"][key] = ""
    else:
        config["PATH"] = {
            "source": r"G:\优仙 工作\监控头像\头像",
            "output": DESKTOP_PATH,
            "single_query": "",
            "multi_queries": ""
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            config.write(f)
    return (
        config["PATH"]["source"],
        config["PATH"]["output"],
        config["PATH"]["single_query"],
        config["PATH"]["multi_queries"].split(";") if config["PATH"]["multi_queries"] else []
    )


def save_config(source: str, output: str, single_query: str, multi_queries: List[str]):
    config["PATH"]["source"] = source
    config["PATH"]["output"] = output
    config["PATH"]["single_query"] = single_query
    config["PATH"]["multi_queries"] = ";".join([p for p in multi_queries if p])
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        config.write(f)


def clean_data(val) -> str:
    if val is None:
        return ""
    val_str = str(val).strip()
    val_str = ''.join(char for char in val_str if ord(char) >= 32 or char in '\n\t')
    return val_str if val_str else "未匹配"


def split_keywords(keyword_str: str) -> List[str]:
    if not keyword_str.strip():
        return []
    keyword_str = keyword_str.replace('，', ',').strip()
    return [kw.strip() for kw in keyword_str.split(',') if kw.strip()]


def open_folder(folder_path: str):
    if os.path.exists(folder_path):
        os.startfile(folder_path)
    else:
        QMessageBox.warning(None, "警告", f"文件夹不存在：{folder_path}")


# -------------------------- 数据处理线程类 --------------------------
class LoadSingleQueryThread(QThread):
    progress_signal = Signal(int, str)
    log_signal = Signal(str)
    finished_signal = Signal(bool, object)

    def __init__(self, excel_path: str):
        super().__init__()
        self.excel_path = excel_path

    def run(self):
        try:
            self.log_signal.emit(f"开始加载单个查询表：{os.path.basename(self.excel_path)}")
            self.progress_signal.emit(20, "读取表格结构")

            wb = load_workbook(self.excel_path, data_only=True)
            ws = wb.active

            header_row = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                header_val = clean_data(cell.value)
                header_row.append(header_val)
            self.log_signal.emit(f"表格表头：{header_row}")

            col_indices = {}
            missing_cols = []
            for req_col in REQUIRED_COLS:
                if req_col in header_row:
                    col_indices[req_col] = header_row.index(req_col)
                else:
                    col_indices[req_col] = None
                    missing_cols.append(req_col)
            if missing_cols:
                self.log_signal.emit(f"警告：表格缺少列{missing_cols}，对应数据显示'未匹配'")

            self.progress_signal.emit(40, "读取数据内容")
            mapping = {}
            row_count = 0
            valid_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
                phone = ""
                if col_indices["有效手机号"] is not None and col_indices["有效手机号"] < len(row):
                    phone = clean_data(row[col_indices["有效手机号"]])
                if not phone or phone == "未匹配":
                    continue

                valid_count += 1
                contact = clean_data(row[col_indices["法定代表人"]]) if (
                            col_indices["法定代表人"] is not None and col_indices["法定代表人"] < len(row)) else "未匹配"
                province = clean_data(row[col_indices["所属省份"]]) if (
                            col_indices["所属省份"] is not None and col_indices["所属省份"] < len(row)) else "未匹配"
                city = clean_data(row[col_indices["所属城市"]]) if (
                            col_indices["所属城市"] is not None and col_indices["所属城市"] < len(row)) else "未匹配"
                district = clean_data(row[col_indices["所属区县"]]) if (
                            col_indices["所属区县"] is not None and col_indices["所属区县"] < len(row)) else "未匹配"

                mapping[phone] = (contact, province, city, district)

            wb.close()
            self.progress_signal.emit(100, "加载完成")
            self.log_signal.emit(f"加载成功：共{row_count}行数据，有效手机号{valid_count}个")
            self.finished_signal.emit(True, mapping)

        except Exception as e:
            self.log_signal.emit(f"加载单个查询表失败：{str(e)}")
            self.finished_signal.emit(False, {})


class MergeMultiQueryThread(QThread):
    progress_signal = Signal(int, str)
    log_signal = Signal(str)
    finished_signal = Signal(bool, object)

    def __init__(self, excel_paths: List[str]):
        super().__init__()
        self.excel_paths = excel_paths

    def run(self):
        try:
            total_tables = len(self.excel_paths)
            self.log_signal.emit(f"开始合并{total_tables}个查询表（自动合并，无需手动操作）")
            self.progress_signal.emit(10, "初始化合并")

            merged_mapping = {}
            for idx, excel_path in enumerate(self.excel_paths, 1):
                progress = int((idx / total_tables) * 80)
                self.progress_signal.emit(progress, f"处理第{idx}/{total_tables}个表格")

                if not os.path.exists(excel_path):
                    self.log_signal.emit(f"警告：第{idx}个表格不存在，跳过：{os.path.basename(excel_path)}")
                    continue

                try:
                    self.log_signal.emit(f"读取表格：{os.path.basename(excel_path)}")
                    wb = load_workbook(excel_path, data_only=True)
                    ws = wb.active

                    header_row = []
                    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                        header_val = clean_data(cell.value)
                        header_row.append(header_val)

                    col_indices = {}
                    for req_col in REQUIRED_COLS:
                        col_indices[req_col] = header_row.index(req_col) if req_col in header_row else None

                    valid_count = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        phone = ""
                        if col_indices["有效手机号"] is not None and col_indices["有效手机号"] < len(row):
                            phone = clean_data(row[col_indices["有效手机号"]])
                        if not phone or phone == "未匹配" or phone in merged_mapping:
                            continue

                        valid_count += 1
                        contact = clean_data(row[col_indices["法定代表人"]]) if (
                                    col_indices["法定代表人"] is not None and col_indices["法定代表人"] < len(
                                row)) else "未匹配"
                        province = clean_data(row[col_indices["所属省份"]]) if (
                                    col_indices["所属省份"] is not None and col_indices["所属省份"] < len(
                                row)) else "未匹配"
                        city = clean_data(row[col_indices["所属城市"]]) if (
                                    col_indices["所属城市"] is not None and col_indices["所属城市"] < len(
                                row)) else "未匹配"
                        district = clean_data(row[col_indices["所属区县"]]) if (
                                    col_indices["所属区县"] is not None and col_indices["所属区县"] < len(
                                row)) else "未匹配"

                        merged_mapping[phone] = (contact, province, city, district)

                    wb.close()
                    self.log_signal.emit(f"第{idx}个表格处理完成：新增{valid_count}个有效手机号")

                except Exception as e:
                    self.log_signal.emit(f"处理第{idx}个表格失败：{str(e)}")
                    continue

            self.progress_signal.emit(90, "保存合并结果")
            self._save_merged_file(merged_mapping)

            self.progress_signal.emit(100, "合并完成")
            self.log_signal.emit(f"合并成功：共获取{len(merged_mapping)}个唯一有效手机号")
            self.finished_signal.emit(True, merged_mapping)

        except Exception as e:
            self.log_signal.emit(f"合并多个查询表失败：{str(e)}")
            self.finished_signal.emit(False, {})

    def _save_merged_file(self, mapping: dict):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "合并查询表_自动生成"
            ws.append(REQUIRED_COLS)

            # 表头样式（柔和科技风）
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")  # 柔和科技蓝
            # 表头全部边框加粗，外部边框统一
            header_border = Border(
                left=Side(style="thick", color="333333"),
                right=Side(style="thick", color="333333"),
                top=Side(style="thick", color="333333"),
                bottom=Side(style="thick", color="333333")
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            ws.row_dimensions[1].height = 30

            # 数据行样式（柔和交替色+内部细边框）
            data_font = Font(bold=True, size=11, color="333333")  # 深灰字体，降低阅读压力
            # 内部边框用细边框，仅起分隔作用
            data_border = Border(
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD")
            )
            fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色
            fill_even = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")  # 极浅灰

            for row_idx, (phone, (contact, province, city, district)) in enumerate(mapping.items(), start=2):
                # 写入数据
                ws.cell(row=row_idx, column=1, value=phone)
                ws.cell(row=row_idx, column=2, value=contact)
                ws.cell(row=row_idx, column=3, value=province)
                ws.cell(row=row_idx, column=4, value=city)
                ws.cell(row=row_idx, column=5, value=district)

                # 应用样式
                fill = fill_odd if row_idx % 2 == 0 else fill_even
                for col in range(1, 6):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = data_border
                    cell.fill = fill
                ws.row_dimensions[row_idx].height = 25

            # 调整列宽
            column_widths = [18, 18, 15, 15, 15]
            for col_idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = width

            # 表格外部边框加粗（统一包裹）
            self._add_outer_border(ws, 1, len(mapping) + 1, 1, 5)

            merged_path = os.path.join(DESKTOP_PATH,
                                       f"自动合并查询表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(merged_path)
            self.log_signal.emit(f"合并文件已保存到：{merged_path}")
        except Exception as e:
            self.log_signal.emit(f"保存合并文件失败：{str(e)}")

    def _add_outer_border(self, ws, start_row, end_row, start_col, end_col):
        """为表格添加外部加粗边框"""
        thick_border = Border(
            left=Side(style="thick", color="333333"),
            right=Side(style="thick", color="333333"),
            top=Side(style="thick", color="333333"),
            bottom=Side(style="thick", color="333333")
        )
        # 顶部边框（第一行所有列）
        for col in range(start_col, end_col + 1):
            ws.cell(row=start_row, column=col).border = thick_border
        # 底部边框（最后一行所有列）
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=end_row, column=col)
            current_border = cell.border
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=Side(style="thick", color="333333")
            )
        # 左侧边框（所有行第一列）
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=start_col)
            current_border = cell.border
            cell.border = Border(
                left=Side(style="thick", color="333333"),
                right=current_border.right,
                top=current_border.top,
                bottom=current_border.bottom
            )
        # 右侧边框（所有行最后一列）
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=end_col)
            current_border = cell.border
            cell.border = Border(
                left=current_border.left,
                right=Side(style="thick", color="333333"),
                top=current_border.top,
                bottom=current_border.bottom
            )


class GenerateSummaryThread(QThread):
    progress_signal = Signal(int, str)
    log_signal = Signal(str)
    finished_signal = Signal(bool)

    def __init__(self, source_folder: str, output_folder: str,
                 data_mapping: Optional[dict], filter_keywords: List[str]):
        super().__init__()
        self.source_folder = source_folder
        self.output_folder = output_folder
        self.data_mapping = data_mapping or {}
        self.filter_keywords = filter_keywords

    def run(self):
        try:
            self.log_signal.emit("开始生成汇总Excel...")
            self.progress_signal.emit(10, "扫描头像文件")

            all_data = []
            file_count = 0
            total_files = 0
            for root, _, files in os.walk(self.source_folder):
                total_files += sum(1 for f in files if f.lower().endswith(".png"))
            if total_files == 0:
                self.log_signal.emit("警告：未找到任何PNG头像文件")
                self.finished_signal.emit(False)
                return

            for root, _, files in os.walk(self.source_folder):
                for file in files:
                    if file.lower().endswith(".png"):
                        file_count += 1
                        progress = int((file_count / total_files) * 40) + 10
                        self.progress_signal.emit(progress, f"扫描第{file_count}/{total_files}个头像")

                        file_name = os.path.splitext(file)[0]
                        # -------------------------- 核心修改点 --------------------------
                        # 原逻辑：phone, wechat_name = 手机号_微信名
                        # 新逻辑：wechat_name, phone = 微信名_手机号
                        if "_" in file_name:
                            wechat_name, phone = file_name.split("_", 1)  # 拆分顺序反转
                        else:
                            wechat_name = "未命名"
                            phone = file_name  # 无下划线时，手机号为文件名，微信名为未命名
                        # -------------------------- 核心修改点 --------------------------
                        phone_clean = clean_data(phone)
                        contact, province, city, district = self.data_mapping.get(
                            phone_clean, ("未匹配", "未匹配", "未匹配", "未匹配")
                        )

                        all_data.append({
                            "手机号": phone_clean,
                            "微信名": clean_data(wechat_name),
                            "联系人": contact,
                            "所属省份": province,
                            "所属城市": city,
                            "所属区县": district,
                            "头像路径": os.path.join(root, file)
                        })
                        self.log_signal.emit(f"已扫描：{file} | 微信名：{clean_data(wechat_name)} | 手机号：{phone_clean} | 联系人：{contact}")

            self.progress_signal.emit(50, "筛选数据（如需要）")
            filtered_data = []
            if self.filter_keywords:
                self.log_signal.emit(f"筛选关键词：{self.filter_keywords}（任一匹配）")
                for item in all_data:
                    if any(kw.lower() in item["微信名"].lower() for kw in self.filter_keywords):
                        filtered_data.append(item)
                self.log_signal.emit(f"筛选完成：{len(filtered_data)}条符合条件")

            self.progress_signal.emit(60, "生成全量数据文件")
            current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            all_file_path = os.path.join(self.output_folder, f"头像汇总_全量_{current_time}.xlsx")
            self._create_excel(all_data, all_file_path, "全量数据汇总表")
            self.log_signal.emit(f"全量文件生成：{all_file_path}")

            if self.filter_keywords and filtered_data:
                self.progress_signal.emit(80, "生成筛选数据文件")
                filter_file_path = os.path.join(self.output_folder,
                                                f"头像汇总_筛选_{'_'.join(self.filter_keywords)}_{current_time}.xlsx")
                self._create_excel(filtered_data, filter_file_path,
                                   f"筛选数据（关键词：{','.join(self.filter_keywords)}）")
                self.log_signal.emit(f"筛选文件生成：{filter_file_path}")

            self.progress_signal.emit(100, "生成完成")
            self.log_signal.emit("所有汇总文件生成成功！")
            self.finished_signal.emit(True)

        except Exception as e:
            self.log_signal.emit(f"生成汇总失败：{str(e)}")
            self.finished_signal.emit(False)

    def _create_excel(self, data: list, file_path: str, sheet_name: str):
        """创建柔和科技风Excel文件（优化色彩和边框）"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # 定义样式（柔和科技风核心配置）
            headers = ["手机号", "微信名", "联系人", "所属省份", "所属城市", "所属区县", "头像"]

            # 1. 表头样式（柔和科技蓝+全加粗边框）
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")  # 柔和科技蓝
            header_border = Border(
                left=Side(style="thick", color="333333"),
                right=Side(style="thick", color="333333"),
                top=Side(style="thick", color="333333"),
                bottom=Side(style="thick", color="333333")
            )
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 2. 数据行样式（柔和交替色+内部细边框）
            data_font = Font(bold=True, size=11, color="333333")  # 深灰字体，低阅读压力
            data_border = Border(
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD")
            )
            data_alignment = Alignment(horizontal="center", vertical="center")
            fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色
            fill_even = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")  # 极浅灰

            # 写入表头并应用样式
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            ws.row_dimensions[1].height = 30  # 表头行高

            # 写入数据并应用样式
            for idx, item in enumerate(data, start=2):
                # 写入文本数据
                ws.cell(row=idx, column=1, value=item["手机号"]).font = data_font
                ws.cell(row=idx, column=2, value=item["微信名"]).font = data_font
                ws.cell(row=idx, column=3, value=item["联系人"]).font = data_font
                ws.cell(row=idx, column=4, value=item["所属省份"]).font = data_font
                ws.cell(row=idx, column=5, value=item["所属城市"]).font = data_font
                ws.cell(row=idx, column=6, value=item["所属区县"]).font = data_font

                # 应用数据行样式（交替背景色+细边框）
                fill = fill_odd if idx % 2 == 0 else fill_even
                for col in range(1, 7):
                    cell = ws.cell(row=idx, column=col)
                    cell.alignment = data_alignment
                    cell.border = data_border
                    cell.fill = fill

                # 处理头像列（统一样式）
                try:
                    img_cell = ws.cell(row=idx, column=7)
                    img_cell.alignment = data_alignment
                    img_cell.border = data_border
                    img_cell.fill = fill

                    if os.path.exists(item["头像路径"]) and os.path.getsize(item["头像路径"]) > 0:
                        img = Image(item["头像路径"])
                        img.width = 70
                        img.height = 70
                        ws.add_image(img, f"G{idx}")
                    else:
                        img_cell.value = "图片异常"
                        img_cell.font = data_font
                except Exception as e:
                    self.log_signal.emit(f"处理图片{item['头像路径']}失败：{str(e)}")
                    img_cell = ws.cell(row=idx, column=7)
                    img_cell.value = "图片异常"
                    img_cell.font = data_font
                    img_cell.alignment = data_alignment
                    img_cell.border = data_border
                    img_cell.fill = fill

                ws.row_dimensions[idx].height = 85  # 数据行高（适配头像）

            # 调整列宽（适配内容，避免文字截断）
            column_widths = [20, 28, 20, 16, 16, 16, 14]
            column_letters = ["A", "B", "C", "D", "E", "F", "G"]
            for letter, width in zip(column_letters, column_widths):
                ws.column_dimensions[letter].width = width

            # 为整个表格添加外部加粗边框
            self._add_outer_border(ws, 1, len(data) + 1, 1, len(headers))

            # 保存文件
            wb.save(file_path)
            wb.close()
        except Exception as e:
            self.log_signal.emit(f"生成文件{os.path.basename(file_path)}失败：{str(e)}")

    def _add_outer_border(self, ws, start_row, end_row, start_col, end_col):
        """为表格添加外部加粗边框（统一视觉）"""
        thick_border = Side(style="thick", color="333333")
        # 顶部边框（第一行所有列）
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=thick_border,
                bottom=cell.border.bottom
            )
        # 底部边框（最后一行所有列）
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=thick_border
            )
        # 左侧边框（所有行第一列）
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=start_col)
            cell.border = Border(
                left=thick_border,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
        # 右侧边框（所有行最后一列）
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=end_col)
            cell.border = Border(
                left=cell.border.left,
                right=thick_border,
                top=cell.border.top,
                bottom=cell.border.bottom
            )


# -------------------------- 主界面类（添加作者信息） --------------------------
class AvatarSummaryWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("凡先生-头像汇总工具-V3.5")  # 版本号更新为V3.5
        self.setGeometry(100, 100, 1100, 880)  # 增加高度容纳作者信息

        self.source_path, self.output_path, self.single_query_path, self.multi_query_paths = init_config()

        self.current_thread = None
        self.data_mapping = None

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        self.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #1E1E2E;
                color: #E0E0E0;
            }
            QGroupBox {
                font-size: 11pt;
                font-weight: bold;
                color: #00C8FF;
                border: 2px solid #3A3A56;
                border-radius: 8px;
                margin-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                padding: 0 5px;
            }
            QLineEdit {
                background-color: #2D2D44;
                color: #FFFFFF;
                border: 1px solid #4A4A6A;
                border-radius: 4px;
                padding: 6px 10px;
                font-size: 10pt;
            }
            QLineEdit:focus {
                border-color: #00C8FF;
                background-color: #373752;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x1:1, y1:0, stop:0 #007ACC, stop:1 #00C8FF);
                color: #FFFFFF;
                font-size: 10pt;
                font-weight: bold;
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x1:1, y1:0, stop:0 #006BB8, stop:1 #00B5E5);
            }
            QPushButton:disabled {
                background: #3A3A56;
                color: #888888;
            }
            QTextEdit {
                background-color: #2D2D44;
                color: #E0E0E0;
                border: 1px solid #4A4A6A;
                border-radius: 4px;
                font-family: Consolas;
                font-size: 10pt;
            }
            QProgressBar {
                background-color: #2D2D44;
                border: 1px solid #4A4A6A;
                border-radius: 4px;
                text-align: center;
                color: #E0E0E0;
                height: 12px;
            }
            QProgressBar::chunk {
                background-color: #00C8FF;
                border-radius: 3px;
            }
            QLabel {
                font-size: 10pt;
                color: #E0E0E0;
            }
            .author-label {
                font-size: 11pt;
                color: #00C8FF;
                font-weight: bold;
            }
            QFrame {
                border: 1px solid #3A3A56;
                border-radius: 4px;
            }
        """)

        # 源文件夹区域
        source_group = QGroupBox("源文件夹（头像根目录）")
        source_layout = QHBoxLayout(source_group)
        self.source_edit = QLineEdit(self.source_path)
        self.source_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.source_btn = QPushButton("浏览")
        self.source_btn.clicked.connect(self.select_source)
        source_layout.addWidget(self.source_edit)
        source_layout.addWidget(self.source_btn)
        main_layout.addWidget(source_group)

        # 查询表选择区域
        query_group = QGroupBox("查询表选择（可选，选择后自动加载/合并，无需手动操作）")
        query_layout = QVBoxLayout(query_group)

        def create_separator():
            sep = QFrame()
            sep.setFrameShape(QFrame.HLine)
            sep.setStyleSheet("background-color: #3A3A56;")
            return sep

        # 单个查询表区域
        single_layout = QHBoxLayout()
        self.single_query_edit = QLineEdit(self.single_query_path)
        self.single_query_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.single_query_btn = QPushButton("选择单个查询表")
        self.single_query_btn.clicked.connect(self.select_single_query)
        single_layout.addWidget(QLabel("单个查询表（直接使用）："))
        single_layout.addWidget(self.single_query_edit)
        single_layout.addWidget(self.single_query_btn)
        query_layout.addLayout(single_layout)

        query_layout.addWidget(create_separator())

        # 多个查询表区域
        multi_layout = QHBoxLayout()
        self.multi_query_edit = QLineEdit(";".join(self.multi_query_paths) if self.multi_query_paths else "")
        self.multi_query_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.multi_query_btn = QPushButton("选择多个查询表")
        self.multi_query_btn.clicked.connect(self.select_multi_queries)
        multi_layout.addWidget(QLabel("多个查询表（自动合并）："))
        multi_layout.addWidget(self.multi_query_edit)
        multi_layout.addWidget(self.multi_query_btn)
        query_layout.addLayout(multi_layout)

        tip_label = QLabel("提示：单个和多个查询表只能选择一种，优先使用单个查询表")
        tip_label.setStyleSheet("color: #FFA500; font-size: 9pt;")
        query_layout.addWidget(tip_label)

        main_layout.addWidget(query_group)

        # 筛选区域
        filter_group = QGroupBox("微信名筛选（可选）")
        filter_layout = QHBoxLayout(filter_group)
        filter_layout.addWidget(QLabel("关键词（中文/英文逗号分隔）："))
        self.filter_edit = QLineEdit()
        self.filter_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        filter_layout.addWidget(self.filter_edit)
        filter_layout.addWidget(QLabel("示例：张三,李四 或 张三，李四"))
        main_layout.addWidget(filter_group)

        # 输出文件夹区域
        output_group = QGroupBox("输出文件夹")
        output_layout = QHBoxLayout(output_group)
        self.output_edit = QLineEdit(self.output_path)
        self.output_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.output_btn = QPushButton("浏览")
        self.output_btn.clicked.connect(self.select_output)
        self.open_output_btn = QPushButton("打开目录")
        self.open_output_btn.clicked.connect(self.open_output_folder)
        output_layout.addWidget(self.output_edit)
        output_layout.addWidget(self.output_btn)
        output_layout.addWidget(self.open_output_btn)
        main_layout.addWidget(output_group)

        # 进度条区域
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("进度：%p% - 就绪")
        main_layout.addWidget(self.progress_bar)

        # 功能按钮区域
        btn_layout = QHBoxLayout()
        self.generate_btn = QPushButton("生成汇总Excel（柔和科技风）")
        self.generate_btn.clicked.connect(self.start_generate)
        self.open_merged_btn = QPushButton("打开合并文件目录")
        self.open_merged_btn.clicked.connect(self.open_merged_folder)
        btn_layout.addWidget(self.generate_btn)
        btn_layout.addWidget(self.open_merged_btn)
        main_layout.addLayout(btn_layout)

        # 日志区域
        log_group = QGroupBox("操作日志")
        log_layout = QVBoxLayout(log_group)
        self.log_edit = QTextEdit()
        self.log_edit.setReadOnly(True)
        log_layout.addWidget(self.log_edit)
        main_layout.addWidget(log_group, stretch=1)

        # 作者信息区域（新增）
        author_layout = QHBoxLayout()
        author_label = QLabel("作者：凡先生 | 联系电话：17665260321")
        author_label.setObjectName("author-label")
        author_label.setAlignment(Qt.AlignCenter)
        author_layout.addWidget(author_label)
        main_layout.addLayout(author_layout)

        self.init_log()

    def init_log(self):
        self.log("欢迎使用头像汇总工具（科技版 V3.5 - 柔和配色版）")
        self.log("版本更新：PNG文件名解析格式从「手机号_微信名」改为「微信名_手机号」")
        self.log("表格优化特性：")
        self.log("1. 色彩搭配：柔和科技蓝表头 + 白色/极浅灰交替数据行，低阅读压力")
        self.log("2. 边框样式：外部边框加粗 + 表头全加粗边框 + 内部细边框分隔")
        self.log("3. 字体：深灰加粗，对比适中，长时间阅读不疲劳")
        self.log("使用说明：")
        self.log("1. 选择头像源文件夹和输出文件夹")
        self.log("2. 可选：选择单个或多个查询表（自动加载/合并）")
        self.log("3. 可选：输入筛选关键词（中文/英文逗号分隔，任一匹配）")
        self.log("4. 点击「生成汇总Excel」，自动完成所有流程")
        self.log("注意：头像文件名需遵循「微信名_手机号.png」格式（下划线分隔，顺序不可颠倒）\n")

    @Slot(str)
    def log(self, message: str):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_edit.append(f"[{timestamp}] {message}")
        self.log_edit.ensureCursorVisible()

    @Slot(int, str)
    def update_progress(self, value: int, desc: str):
        self.progress_bar.setValue(value)
        self.progress_bar.setFormat(f"进度：{value}% - {desc}")

    # 界面事件
    def select_source(self):
        folder = QFileDialog.getExistingDirectory(self, "选择头像源文件夹", self.source_path)
        if folder:
            self.source_edit.setText(folder)
            self.source_path = folder

    def select_single_query(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "选择单个查询表",
            os.path.dirname(self.single_query_path) if self.single_query_path else os.path.expanduser("~"),
            "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        if file:
            self.single_query_edit.setText(file)
            self.single_query_path = file
            self.multi_query_edit.setText("")
            self.multi_query_paths = []

    def select_multi_queries(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择多个查询表",
            os.path.dirname(self.multi_query_paths[0]) if self.multi_query_paths else os.path.expanduser("~"),
            "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        if files:
            self.multi_query_paths = files
            self.multi_query_edit.setText(";".join(files))
            self.single_query_edit.setText("")
            self.single_query_path = ""

    def select_output(self):
        folder = QFileDialog.getExistingDirectory(self, "选择输出文件夹", self.output_path)
        if folder:
            self.output_edit.setText(folder)
            self.output_path = folder

    def open_output_folder(self):
        folder = self.output_edit.text().strip() or self.output_path
        self.log(f"打开输出文件夹：{folder}")
        open_folder(folder)

    def open_merged_folder(self):
        self.log(f"打开合并文件目录：{DESKTOP_PATH}")
        open_folder(DESKTOP_PATH)

    # 核心生成流程
    def start_generate(self):
        source_folder = self.source_edit.text().strip()
        output_folder = self.output_edit.text().strip()
        if not os.path.exists(source_folder):
            QMessageBox.critical(self, "错误", "源文件夹不存在！")
            self.log("错误：源文件夹不存在")
            return
        if not os.path.exists(output_folder):
            QMessageBox.critical(self, "错误", "输出文件夹不存在！")
            self.log("错误：输出文件夹不存在")
            return

        save_config(
            source_folder, output_folder,
            self.single_query_path, self.multi_query_paths
        )

        self.generate_btn.setEnabled(False)
        self.open_output_btn.setEnabled(False)
        self.open_merged_btn.setEnabled(False)

        if self.single_query_path and os.path.exists(self.single_query_path):
            self.log("检测到单个查询表，自动加载...")
            self.current_thread = LoadSingleQueryThread(self.single_query_path)
            self.current_thread.progress_signal.connect(self.update_progress)
            self.current_thread.log_signal.connect(self.log)
            self.current_thread.finished_signal.connect(self.on_query_process_finished)
            self.current_thread.start()
        elif self.multi_query_paths:
            self.log("检测到多个查询表，自动合并...")
            self.current_thread = MergeMultiQueryThread(self.multi_query_paths)
            self.current_thread.progress_signal.connect(self.update_progress)
            self.current_thread.log_signal.connect(self.log)
            self.current_thread.finished_signal.connect(self.on_query_process_finished)
            self.current_thread.start()
        else:
            self.log("未选择查询表，联系人/地区显示'未匹配'")
            self.data_mapping = {}
            self.start_summary_generate()

    @Slot(bool, object)
    def on_query_process_finished(self, success: bool, mapping: dict):
        if success:
            self.data_mapping = mapping
        else:
            self.data_mapping = {}
            QMessageBox.warning(self, "警告", "查询表处理失败，联系人/地区将显示'未匹配'")
        self.start_summary_generate()

    def start_summary_generate(self):
        source_folder = self.source_edit.text().strip()
        output_folder = self.output_edit.text().strip()
        filter_keywords = split_keywords(self.filter_edit.text().strip())

        self.log("开始生成汇总Excel（柔和科技风表格...）")
        self.current_thread = GenerateSummaryThread(
            source_folder, output_folder, self.data_mapping, filter_keywords
        )
        self.current_thread.progress_signal.connect(self.update_progress)
        self.current_thread.log_signal.connect(self.log)
        self.current_thread.finished_signal.connect(self.on_generate_finished)
        self.current_thread.start()

    @Slot(bool)
    def on_generate_finished(self, success: bool):
        if success:
            QMessageBox.information(self, "成功", "汇总Excel（柔和科技风）生成完成！")
        else:
            QMessageBox.critical(self, "失败", "汇总Excel生成失败，请查看日志")

        self.generate_btn.setEnabled(True)
        self.open_output_btn.setEnabled(True)
        self.open_merged_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("进度：%p% - 就绪")


# -------------------------- 程序入口 --------------------------
if __name__ == "__main__":
    try:
        import openpyxl
        from PySide2.QtWidgets import QApplication
    except ImportError:
        print("缺少依赖库，请执行：pip install openpyxl pyside2 numpy==1.26.4")
        sys.exit(1)

    if sys.platform == "win32":
        import _locale

        try:
            _locale._getdefaultlocale = (lambda *args: ['en_US', 'utf8'])
        except Exception:
            pass

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = AvatarSummaryWindow()
    window.show()
    sys.exit(app.exec_())