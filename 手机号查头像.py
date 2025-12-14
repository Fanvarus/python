import os
import re
import time
import random
import threading
import sqlite3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pyautogui
import pyperclip
import mss
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from pathlib import Path
from ctypes import windll, POINTER, c_int, byref

# ===================================== 全局配置（可修改）=====================================
DEFAULT_SAVE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "监控头像")  # 截图保存路径
EXCEL_FILE_NAME = "微信监控数据.xlsx"  # Excel表格文件名
DB_FILE_NAME = "wechat_monitor.db"  # 数据库文件名
PREPARE_TIME = 5  # 准备时间（秒）
OPERATE_DELAY_MIN = 0.3  # 操作间隔最小值（秒）
OPERATE_DELAY_MAX = 0.8  # 操作间隔最大值（秒）
CELL_PADDING = 5  # 图片与单元格边框的内边距（像素）

# 拟人化配置
MOUSE_MOVE_DURATION_MIN = 0.2  # 鼠标移动时间最小值（秒）
MOUSE_MOVE_DURATION_MAX = 0.5  # 鼠标移动时间最大值（秒）
INPUT_INTERVAL_MIN = 0.03  # 输入字符间隔最小值（秒）
INPUT_INTERVAL_MAX = 0.1  # 输入字符间隔最大值（秒）
BATCH_REST_COUNT = 5  # 每处理N个手机号后休息
BATCH_REST_MIN = 3  # 批量休息最小值（秒）
BATCH_REST_MAX = 6  # 批量休息最大值（秒）
CLICK_OFFSET_RANGE = 5  # 点击坐标随机偏移范围（像素）

# 表格样式配置（专业风）
HEADER_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")  # 表头字体：白色14号加粗
CONTENT_FONT = Font(name="微软雅黑", size=11, bold=True, color="333333")  # 正文字体：深灰11号加粗
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 表头背景：深蓝色
BORDER = Border(  # 表格边框：细黑边
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# 坐标配置（改为「目标区域」而非固定点，容错率更高）
COORDS = {
    "input_activate": (1250, 625, 1260, 635),  # 左上x,左上y,右下x,右下y（区域）
    "after_enter": (1150, 670, 1155, 675),  # 回车后点击区域
    "copy_double_click": (1170, 765, 1180, 770),  # 双击复制区域
    "screenshot_activate": (1420, 720, 1428, 728),  # 截图前激活区域
    "fail_click": (1385, 745, 1393, 750),  # 复制失败时点击区域
    "capture_left_top": (1275, 783),  # 截图左上角（固定）
    "capture_right_bottom": (1561, 1067)  # 截图右下角（固定）
}

# 隐藏工具特征：禁用pyautogui的失败安全（避免触发弹窗）
pyautogui.FAILSAFE = False
# 隐藏鼠标移动痕迹（Windows系统）
try:
    user32 = windll.user32
    # 设置鼠标移动时不显示轨迹
    user32.SetSystemCursor(None, 0)
except:
    pass


# ===================================== 数据库工具类=====================================
class Database:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        self.init_db()  # 初始化数据库（创建表）

    def connect(self):
        """连接数据库"""
        try:
            self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
            self.cursor = self.conn.cursor()
            return True
        except Exception as e:
            print(f"数据库连接失败：{str(e)}")
            return False

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            self.conn = None
            self.cursor = None

    def init_db(self):
        """初始化数据库表结构"""
        if not self.connect():
            return

        # 1. 业务数据表：存储手机号、微信名、头像路径
        self.cursor.execute('''
                            CREATE TABLE IF NOT EXISTS phone_data
                            (
                                phone
                                TEXT
                                PRIMARY
                                KEY
                                NOT
                                NULL, -- 手机号（主键，唯一）
                                wechat_name
                                TEXT
                                NOT
                                NULL, -- 微信名
                                avatar_path
                                TEXT
                                NOT
                                NULL, -- 头像路径
                                create_time
                                TEXT
                                NOT
                                NULL  -- 处理时间
                            )
                            ''')

        # 2. 进度表：记录每个TXT文件的最后处理行数（断点续读用）
        self.cursor.execute('''
                            CREATE TABLE IF NOT EXISTS process_progress
                            (
                                txt_path
                                TEXT
                                PRIMARY
                                KEY
                                NOT
                                NULL, -- TXT文件路径（主键，唯一标识文件）
                                last_line
                                INT
                                NOT
                                NULL
                                DEFAULT
                                0,    -- 最后处理的行号（从0开始）
                                update_time
                                TEXT
                                NOT
                                NULL  -- 最后更新时间
                            )
                            ''')

        self.conn.commit()
        self.close()

    def get_last_process_line(self, txt_path: str) -> int:
        """获取指定TXT文件的最后处理行号（断点续读）"""
        if not self.connect():
            return 0

        try:
            self.cursor.execute('''
                                SELECT last_line
                                FROM process_progress
                                WHERE txt_path = ?
                                ''', (txt_path,))
            result = self.cursor.fetchone()
            return result[0] if result else 0
        except Exception as e:
            print(f"获取进度失败：{str(e)}")
            return 0
        finally:
            self.close()

    def update_process_line(self, txt_path: str, last_line: int):
        """更新指定TXT文件的最后处理行号"""
        if not self.connect():
            return

        try:
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            # 存在则更新，不存在则插入
            self.cursor.execute('''
                INSERT OR REPLACE INTO process_progress 
                (txt_path, last_line, update_time) 
                VALUES (?, ?, ?)
            ''', (txt_path, last_line, current_time))
            self.conn.commit()
        except Exception as e:
            print(f"更新进度失败：{str(e)}")
        finally:
            self.close()

    def insert_phone_data(self, phone: str, wechat_name: str, avatar_path: str):
        """插入业务数据到数据库"""
        if not self.connect():
            return

        try:
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            # 存在则更新（避免重复插入），不存在则插入
            self.cursor.execute('''
                INSERT OR REPLACE INTO phone_data 
                (phone, wechat_name, avatar_path, create_time) 
                VALUES (?, ?, ?, ?)
            ''', (phone, wechat_name, avatar_path, current_time))
            self.conn.commit()
        except Exception as e:
            print(f"插入数据失败：{str(e)}")
        finally:
            self.close()

    def is_phone_processed(self, phone: str) -> bool:
        """判断手机号是否已处理过（避免重复）"""
        if not self.connect():
            return False

        try:
            self.cursor.execute('''
                                SELECT 1
                                FROM phone_data
                                WHERE phone = ? LIMIT 1
                                ''', (phone,))
            return self.cursor.fetchone() is not None
        except Exception as e:
            print(f"查询手机号状态失败：{str(e)}")
            return False
        finally:
            self.close()


# ===================================== 工具函数（拟人化优化版）=====================================
def is_11_digit(text: str) -> bool:
    """验证是否为11位纯数字"""
    return re.fullmatch(r"\d{11}", text.strip()) is not None


def replace_invalid_filename_chars(text: str) -> str:
    """替换文件名中的非法字符"""
    invalid_chars = r'[\\/:*?"<>|]'
    return re.sub(invalid_chars, "_", text)


def create_folder_if_not_exist(folder_path: str):
    """创建文件夹（不存在则创建）"""
    Path(folder_path).mkdir(parents=True, exist_ok=True)
    return folder_path


def capture_screen(save_path: str) -> bool:
    """截取指定区域并保存图片（降低截图频率，优化参数）"""
    try:
        x1, y1 = COORDS["capture_left_top"]
        x2, y2 = COORDS["capture_right_bottom"]
        with mss.mss() as sct:
            monitor = {
                "top": y1, "left": x1,
                "width": x2 - x1, "height": y2 - y1
            }
            sct_img = sct.grab(monitor)
            # 优化：降低图片质量，减少特征
            mss.tools.to_png(sct_img.rgb, sct_img.size, output=save_path, compress_level=6)
        # 截图后随机停顿
        time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))
        return True
    except Exception as e:
        print(f"截图失败：{str(e)}")
        return False


def clear_input_box() -> None:
    """清空输入框（模拟人类可能的重复操作）"""
    # 随机选择清空方式：Ctrl+A+Delete 或 多次Backspace
    if random.random() > 0.3:
        pyautogui.keyDown("ctrl")
        pyautogui.press("a")
        pyautogui.keyUp("ctrl")
        time.sleep(random.uniform(0.1, 0.2))
        pyautogui.press("delete")
    else:
        pyautogui.keyDown("ctrl")
        pyautogui.press("a")
        pyautogui.keyUp("ctrl")
        time.sleep(random.uniform(0.1, 0.2))
        pyautogui.press("backspace")
        # 偶尔多按一次Backspace（人类失误）
        if random.random() > 0.7:
            time.sleep(0.1)
            pyautogui.press("backspace")
    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))


def get_random_point_in_area(area: tuple) -> tuple:
    """在目标区域内获取随机坐标（避免固定点点击）"""
    x1, y1, x2, y2 = area
    x = random.randint(x1, x2)
    y = random.randint(y1, y2)
    return (x, y)


def human_like_move_click(area: tuple) -> None:
    """拟人化移动并点击（随机路径、随机偏移、模拟犹豫）"""
    # 1. 获取目标区域内随机点
    target_x, target_y = get_random_point_in_area(area)
    # 2. 随机偏移目标点（模拟人类点击不准）
    offset_x = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    offset_y = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    target_x += offset_x
    target_y += offset_y
    # 3. 模拟人类鼠标移动曲线（非匀速）
    move_duration = random.uniform(MOUSE_MOVE_DURATION_MIN, MOUSE_MOVE_DURATION_MAX)
    pyautogui.moveTo(target_x, target_y, duration=move_duration, tween=pyautogui.easeInOutQuad)
    # 4. 点击前随机停顿（犹豫）
    time.sleep(random.uniform(0.05, 0.2))
    # 5. 偶尔双击（人类操作误差）
    if random.random() > 0.85:
        pyautogui.doubleClick()
    else:
        pyautogui.click()
    # 6. 点击后随机停顿
    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))


def human_like_double_click(area: tuple) -> None:
    """拟人化双击（避免机械双击）"""
    target_x, target_y = get_random_point_in_area(area)
    offset_x = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    offset_y = random.randint(-CLICK_OFFSET_RANGE, CLICK_OFFSET_RANGE)
    target_x += offset_x
    target_y += offset_y
    move_duration = random.uniform(MOUSE_MOVE_DURATION_MIN, MOUSE_MOVE_DURATION_MAX)
    pyautogui.moveTo(target_x, target_y, duration=move_duration, tween=pyautogui.easeInOutQuad)
    time.sleep(random.uniform(0.05, 0.15))
    # 双击间隔随机
    pyautogui.click()
    time.sleep(random.uniform(0.03, 0.08))
    pyautogui.click()
    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))


def copy_selected_content() -> str:
    """复制选中内容（模拟人类可能的重复复制）"""
    pyperclip.copy("")
    time.sleep(random.uniform(0.05, 0.15))

    # 随机选择复制方式：Ctrl+C 或 右键复制（降低快捷键频率）
    if random.random() > 0.4:
        pyautogui.keyDown("ctrl")
        # 偶尔多按一次C（人类失误）
        if random.random() > 0.8:
            pyautogui.press("c")
            time.sleep(0.05)
        pyautogui.press("c")
        pyautogui.keyUp("ctrl")
    else:
        pyautogui.rightClick()
        time.sleep(random.uniform(0.1, 0.2))
        pyautogui.press("c")  # 右键菜单中选择复制（需确保菜单位置正确）

    time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))
    content = pyperclip.paste().strip()
    return content if content else ""


def human_like_typewrite(text: str) -> None:
    """拟人化输入（随机间隔、偶尔回退重输）"""
    for char in text:
        pyautogui.press(char)
        # 输入间隔随机
        time.sleep(random.uniform(INPUT_INTERVAL_MIN, INPUT_INTERVAL_MAX))
        # 10%概率回退重输（人类输入错误修正）
        if random.random() < 0.1 and len(text) > 3:
            time.sleep(random.uniform(0.1, 0.2))
            pyautogui.press("backspace")
            time.sleep(random.uniform(0.05, 0.1))
            pyautogui.press(char)
            time.sleep(random.uniform(INPUT_INTERVAL_MIN, INPUT_INTERVAL_MAX))


def random_rest() -> None:
    """随机休息（模拟人类走神）"""
    if random.random() > 0.7:
        rest_time = random.uniform(0.5, 1.5)
        time.sleep(rest_time)


def batch_rest(count: int, log_func) -> None:
    """批量处理后休息（避免高强度操作）"""
    if count % BATCH_REST_COUNT == 0 and count != 0:
        rest_time = random.uniform(BATCH_REST_MIN, BATCH_REST_MAX)
        log_func(f"批量处理{count}个，休息{rest_time:.1f}秒...")
        time.sleep(rest_time)


# ===================================== Excel表格生成函数（图片嵌入版）=====================================
class ExcelExporter:
    def __init__(self, save_path: str):
        self.save_path = os.path.join(save_path, EXCEL_FILE_NAME)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "微信监控数据"
        self.init_table_style()

    def init_table_style(self):
        """初始化表格样式（专业美化+嵌入适配）"""
        # 列名：手机号、微信名、头像
        headers = ["手机号", "微信名", "头像"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            # 表头样式：加大加粗、白色文字、蓝色背景、居中对齐、边框
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

            # 调整列宽（适配内容和嵌入图片）
            if col == 1:  # 手机号列（11位数字+边距）
                self.ws.column_dimensions[get_column_letter(col)].width = 18
            elif col == 2:  # 微信名列（适配较长昵称）
                self.ws.column_dimensions[get_column_letter(col)].width = 28
            elif col == 3:  # 头像列（适配图片+内边距，1列≈8px）
                self.ws.column_dimensions[get_column_letter(col)].width = 12  # 12*8=96px

        # 表头行高（适配14号字体）
        self.ws.row_dimensions[1].height = 35

        # 设置工作表默认字体
        self.ws.font = CONTENT_FONT

    def add_row(self, phone: str, wechat_name: str, avatar_path: str):
        """添加一行数据（图片嵌入单元格）"""
        next_row = self.ws.max_row + 1

        # 头像列单元格尺寸（用于图片适配）
        cell_width_px = self.ws.column_dimensions["C"].width * 8  # 1列≈8px
        cell_height_px = 70  # 固定行高，适配图片嵌入
        self.ws.row_dimensions[next_row].height = cell_height_px

        # 1. 手机号列：加粗、居中、边框
        phone_cell = self.ws.cell(row=next_row, column=1, value=phone)
        phone_cell.alignment = Alignment(horizontal="center", vertical="center")
        phone_cell.border = BORDER

        # 2. 微信名列：加粗、居中、边框、无昵称显示
        wechat_name = wechat_name if wechat_name.strip() else "无昵称"
        wechat_cell = self.ws.cell(row=next_row, column=2, value=wechat_name)
        wechat_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wechat_cell.border = BORDER

        # 3. 头像列：嵌入图片+边框
        avatar_cell = self.ws.cell(row=next_row, column=3, value="")
        avatar_cell.alignment = Alignment(horizontal="center", vertical="center")
        avatar_cell.border = BORDER

        # 有头像时嵌入单元格（图片尺寸适配单元格，带内边距）
        if avatar_path and avatar_path.strip() and os.path.exists(avatar_path):
            try:
                # 计算图片最大可用尺寸（单元格尺寸 - 2*内边距）
                max_img_width = cell_width_px - 2 * CELL_PADDING
                max_img_height = cell_height_px - 2 * CELL_PADDING

                # 插入图片并按单元格尺寸等比例缩放
                img = Image(avatar_path)
                # 等比例缩放，确保图片不超出单元格
                img_ratio = img.width / img.height
                max_ratio = max_img_width / max_img_height

                if img_ratio > max_ratio:
                    # 按宽度适配
                    img.width = max_img_width
                    img.height = max_img_width / img_ratio
                else:
                    # 按高度适配
                    img.height = max_img_height
                    img.width = max_img_height * img_ratio

                # 绑定到单元格，实现嵌入效果（随单元格移动）
                img.anchor = f"C{next_row}"

                # 设置图片位置，实现带内边距的居中
                img.left = int(CELL_PADDING * 9525)  # 内边距，EMU单位（1px≈9525）
                img.top = int(CELL_PADDING * 9525)

                self.ws.add_image(img)
            except Exception as e:
                print(f"插入图片失败：{str(e)}")
                # 失败时仍保持单元格空状态，不显示错误文字

    def save(self):
        """保存Excel文件（处理权限问题）"""
        try:
            # 先检查文件是否被占用
            if os.path.exists(self.save_path):
                with open(self.save_path, "rb") as f:
                    pass  # 能打开说明未被占用

            self.wb.save(self.save_path)
            return self.save_path
        except PermissionError:
            messagebox.showerror("错误",
                                 f"Excel文件正在被其他程序打开！\n请关闭 {os.path.basename(self.save_path)} 后重试。")
            return None
        except Exception as e:
            print(f"Excel保存失败：{str(e)}")
            messagebox.showerror("错误", f"Excel保存失败：{str(e)}")
            return None


# ===================================== TK界面主类=====================================
class WeChatMonitorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("微信监控工具 v3.0（防检测优化版）")
        self.root.geometry("950x650")
        self.root.resizable(False, False)

        # 全局变量
        self.txt_file_path = None
        self.save_folder = DEFAULT_SAVE_FOLDER
        self.db = None  # 数据库实例
        self.is_running = False
        self.thread = None
        self.total_valid_lines = 0  # 总有效手机号数
        self.processed_count = 0  # 已处理数
        self.all_lines = []  # 所有行（含无效内容）
        self.valid_tasks = []  # 格式：[(行号, 手机号), ...]
        self.start_line = 0  # 断点续读起始行号

        # 初始化数据库和界面
        self.init_db()
        self.init_ui()

    def init_db(self):
        """初始化数据库"""
        db_path = os.path.join(self.save_folder, DB_FILE_NAME)
        self.db = Database(db_path)
        self.log(f"数据库初始化完成，路径：{db_path}")

    def init_ui(self):
        """初始化界面"""
        # 1. 顶部文件选择区域
        file_frame = ttk.Frame(self.root, padding="10")
        file_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(file_frame, text="TXT文件：", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W)
        self.file_label = ttk.Label(file_frame, text="未选择文件", font=("Arial", 10), foreground="gray")
        self.file_label.grid(row=0, column=1, padx=10, sticky=tk.W)
        ttk.Button(file_frame, text="选择TXT", command=self.select_txt_file).grid(row=0, column=2, padx=10)

        # 断点续读提示
        self.progress_label = ttk.Label(file_frame, text="", font=("Arial", 9), foreground="blue")
        self.progress_label.grid(row=0, column=3, padx=20, sticky=tk.W)

        # 2. 配置区域
        config_frame = ttk.Frame(self.root, padding="10")
        config_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(config_frame, text="准备时间：", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W)
        self.prepare_time_var = tk.StringVar(value=str(PREPARE_TIME))
        ttk.Entry(config_frame, textvariable=self.prepare_time_var, width=10).grid(row=0, column=1, padx=10)
        ttk.Label(config_frame, text="秒", font=("Arial", 10)).grid(row=0, column=2)

        # 操作间隔改为「最小-最大」范围
        ttk.Label(config_frame, text="操作间隔：", font=("Arial", 10)).grid(row=0, column=3, padx=20, sticky=tk.W)
        self.operate_delay_min_var = tk.StringVar(value=str(OPERATE_DELAY_MIN))
        self.operate_delay_max_var = tk.StringVar(value=str(OPERATE_DELAY_MAX))
        ttk.Entry(config_frame, textvariable=self.operate_delay_min_var, width=8).grid(row=0, column=4, padx=5)
        ttk.Label(config_frame, text="-", font=("Arial", 10)).grid(row=0, column=5)
        ttk.Entry(config_frame, textvariable=self.operate_delay_max_var, width=8).grid(row=0, column=6, padx=5)
        ttk.Label(config_frame, text="秒", font=("Arial", 10)).grid(row=0, column=7)

        # 3. 日志显示区域
        log_frame = ttk.Frame(self.root, padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        ttk.Label(log_frame, text="执行日志：", font=("Arial", 10)).pack(anchor=tk.W)
        self.log_text = tk.Text(log_frame, width=105, height=22, font=("Arial", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)  # 禁止编辑

        # 4. 进度条和统计
        progress_frame = ttk.Frame(self.root, padding="10")
        progress_frame.pack(fill=tk.X, padx=10, pady=5)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, side=tk.LEFT, expand=True, padx=5)

        self.stat_label = ttk.Label(progress_frame, text="已处理：0/0", font=("Arial", 10))
        self.stat_label.pack(side=tk.RIGHT, padx=10)

        # 5. 底部按钮区域
        btn_frame = ttk.Frame(self.root, padding="10")
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        self.start_btn = ttk.Button(btn_frame, text="开始执行", command=self.start_execution)
        self.start_btn.pack(side=tk.LEFT, padx=10)
        self.stop_btn = ttk.Button(btn_frame, text="停止执行", command=self.stop_execution, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="打开存储目录", command=self.open_save_folder).pack(side=tk.RIGHT, padx=10)
        ttk.Button(btn_frame, text="退出", command=self.root.quit).pack(side=tk.RIGHT, padx=10)

    def select_txt_file(self):
        """选择TXT文件并显示断点续读进度"""
        file_path = filedialog.askopenfilename(
            title="选择TXT文件",
            filetypes=[("TXT文件", "*.txt"), ("所有文件", "*.*")]
        )
        if file_path:
            self.txt_file_path = file_path
            self.file_label.config(text=os.path.basename(file_path), foreground="black")
            self.log(f"已选择TXT文件：{file_path}")

            # 查询该文件的最后处理行号
            last_line = self.db.get_last_process_line(file_path)
            # 读取文件总行数和有效行数，计算剩余待处理数
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                total_lines = len(lines)
                valid_lines = [line.strip() for line in lines if is_11_digit(line.strip())]
                total_valid = len(valid_lines)

                # 计算已处理的有效行数（从进度中推导）
                processed_valid = 0
                for i in range(min(last_line, total_lines)):
                    if is_11_digit(lines[i].strip()):
                        processed_valid += 1

                self.progress_label.config(
                    text=f"断点续读：上次处理到第{last_line}行，已处理{processed_valid}个有效手机号，剩余{total_valid - processed_valid}个"
                )
            except Exception as e:
                self.log(f"读取文件进度失败：{str(e)}")
                self.progress_label.config(text="断点续读：无法获取文件信息")

    def open_save_folder(self):
        """打开存储目录（截图+Excel+数据库）"""
        create_folder_if_not_exist(self.save_folder)
        os.startfile(self.save_folder)

    def log(self, msg: str):
        """在日志框中添加信息（线程安全）"""
        self.root.after(0, self._update_log, msg)

    def _update_log(self, msg: str):
        """更新日志（内部调用）"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {msg}\n")
        self.log_text.see(tk.END)  # 自动滚动到底部
        self.log_text.config(state=tk.DISABLED)

    def update_progress(self, processed: int, total: int):
        """更新进度条和统计信息（线程安全）"""
        self.root.after(0, self._update_progress_ui, processed, total)

    def _update_progress_ui(self, processed: int, total: int):
        """更新进度UI（内部调用）"""
        self.processed_count = processed
        self.total_valid_lines = total
        progress = (processed / total) * 100 if total > 0 else 0
        self.progress_var.set(progress)
        self.stat_label.config(text=f"已处理：{processed}/{total}")

    def start_execution(self):
        """开始执行（子线程）"""
        # 验证参数
        if not self.txt_file_path:
            messagebox.showwarning("警告", "请先选择TXT文件！")
            return

        try:
            prepare_time = int(self.prepare_time_var.get())
            operate_delay_min = float(self.operate_delay_min_var.get())
            operate_delay_max = float(self.operate_delay_max_var.get())
            if prepare_time < 0 or operate_delay_min < 0 or operate_delay_max < operate_delay_min:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "准备时间需为非负整数，操作间隔需满足：最小值≥0 且 最大值≥最小值！")
            return

        # 检查Excel文件是否被占用
        excel_path = os.path.join(self.save_folder, EXCEL_FILE_NAME)
        if os.path.exists(excel_path):
            try:
                with open(excel_path, "rb") as f:
                    pass
            except PermissionError:
                messagebox.showerror("错误",
                                     f"Excel文件正在被其他程序打开！\n请关闭 {os.path.basename(excel_path)} 后重试。")
                return

        # 确认开始
        last_line = self.db.get_last_process_line(self.txt_file_path)
        confirm_msg = f"即将开始执行！\n准备时间：{prepare_time}秒\n操作间隔：{operate_delay_min}-{operate_delay_max}秒\n"
        if last_line > 0:
            confirm_msg += f"断点续读：将从第{last_line}行开始处理（跳过已处理内容）\n"
        if not messagebox.askyesno("确认", confirm_msg + "是否继续？"):
            return

        # 初始化全局参数
        global PREPARE_TIME, OPERATE_DELAY_MIN, OPERATE_DELAY_MAX
        PREPARE_TIME = prepare_time
        OPERATE_DELAY_MIN = operate_delay_min
        OPERATE_DELAY_MAX = operate_delay_max

        # 创建保存文件夹
        create_folder_if_not_exist(self.save_folder)
        self.log(f"截图保存路径：{self.save_folder}")

        # 读取TXT文件
        try:
            with open(self.txt_file_path, "r", encoding="utf-8") as f:
                self.all_lines = f.readlines()  # 所有行（含无效内容）
            total_lines = len(self.all_lines)
            self.log(f"成功读取TXT文件，总行数：{total_lines}")

            # 获取断点续读的起始行号
            self.start_line = self.db.get_last_process_line(self.txt_file_path)
            if self.start_line > 0:
                self.log(f"断点续读：从第{self.start_line}行开始处理（前{self.start_line}行已跳过）")
            if self.start_line >= total_lines:
                messagebox.showinfo("提示", "该文件已全部处理完成，无需重复执行！")
                return

            # 筛选有效手机号（从起始行开始）
            self.valid_tasks = []  # 格式：[(行号, 手机号), ...]
            for line_num in range(self.start_line, total_lines):
                line_content = self.all_lines[line_num].strip()
                if is_11_digit(line_content) and not self.db.is_phone_processed(line_content):
                    self.valid_tasks.append((line_num, line_content))

            self.total_valid_lines = len(self.valid_tasks)
            if self.total_valid_lines == 0:
                messagebox.showinfo("提示", "当前文件无未处理的有效手机号！")
                return
            self.log(f"筛选出未处理的有效手机号：{self.total_valid_lines}个")
        except Exception as e:
            messagebox.showerror("错误", f"读取TXT文件失败：{str(e)}")
            return

        # 初始化Excel（如果文件不存在则创建，存在则追加）
        if os.path.exists(excel_path):
            # 已存在，读取现有数据行数（避免重复添加列名）
            self.excel_exporter = ExcelExporter(self.save_folder)
            self.excel_exporter.wb = load_workbook(excel_path)
            self.excel_exporter.ws = self.excel_exporter.wb.active
            self.log("Excel文件已存在，将追加数据（图片嵌入格式）")
        else:
            self.excel_exporter = ExcelExporter(self.save_folder)
            self.log("Excel表格初始化完成（专业美化+图片嵌入）")

        # 开始倒计时
        self.log(f"准备倒计时：{PREPARE_TIME}秒，请打开目标窗口...")
        for i in range(PREPARE_TIME, 0, -1):
            self.log(f"倒计时：{i}秒")
            time.sleep(1)
        self.log("倒计时结束，开始执行操作！")

        # 启动子线程执行任务
        self.is_running = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.update_progress(0, self.total_valid_lines)
        self.thread = threading.Thread(target=self.execute_task)
        self.thread.daemon = True
        self.thread.start()

    def execute_task(self):
        """执行核心任务（子线程，拟人化优化版）"""
        processed_count = 0
        last_processed_line = self.start_line  # 记录最后处理的行号

        for (line_num, phone) in self.valid_tasks:
            if not self.is_running:
                self.log("执行已停止，当前进度已保存到数据库")
                break

            processed_count += 1
            self.log(
                f"\n===== 正在处理第{processed_count}/{self.total_valid_lines}个手机号（行号：{line_num}）：{phone} =====")
            avatar_path = ""
            wechat_name = ""

            try:
                # 模拟人类随机休息（走神）
                random_rest()

                # 1. 激活输入框 + 清空内容（拟人化操作）
                human_like_move_click(COORDS["input_activate"])
                self.log("激活输入框并清空内容")
                clear_input_box()

                # 2. 拟人化输入手机号 + 回车（随机间隔、可能修正）
                human_like_typewrite(phone)
                time.sleep(random.uniform(OPERATE_DELAY_MIN, OPERATE_DELAY_MAX))
                # 回车前随机停顿
                time.sleep(random.uniform(0.1, 0.3))
                pyautogui.press("enter")
                self.log("输入手机号并回车")
                # 回车后等待时间随机（模拟网络延迟）
                time.sleep(random.uniform(OPERATE_DELAY_MIN * 2, OPERATE_DELAY_MAX * 2))

                # 3. 点击坐标2（拟人化）
                human_like_move_click(COORDS["after_enter"])
                self.log("点击坐标2")

                # 4. 双击复制微信名（拟人化双击）
                human_like_double_click(COORDS["copy_double_click"])
                wechat_name = copy_selected_content()
                wechat_name = replace_invalid_filename_chars(wechat_name) if wechat_name else ""
                if wechat_name:
                    self.log(f"获取微信名：{wechat_name}")
                else:
                    self.log("未获取到微信名（无昵称）")

                # 5. 判断是否复制成功
                if wechat_name:
                    # 模拟人类随机移动鼠标（无意义操作，增加真实性）
                    if random.random() > 0.6:
                        random_x = random.randint(100, 300)
                        random_y = random.randint(200, 400)
                        pyautogui.moveTo(random_x, random_y, duration=random.uniform(0.2, 0.4))
                        time.sleep(random.uniform(0.1, 0.2))
                        # 移回目标区域附近
                        human_like_move_click(COORDS["screenshot_activate"])
                    else:
                        human_like_move_click(COORDS["screenshot_activate"])

                    self.log("点击坐标4，准备截图")
                    # 保存截图
                    avatar_filename = f"{phone}_{wechat_name}.png"
                    avatar_path = os.path.join(self.save_folder, avatar_filename)
                    if capture_screen(avatar_path):
                        self.log(f"截图保存成功：{avatar_filename}")
                    else:
                        self.log("截图失败！")
                        avatar_path = ""
                else:
                    # 复制失败，点击备用坐标（拟人化）
                    human_like_move_click(COORDS["fail_click"])
                    self.log("未获取到微信名，点击备用坐标，跳过截图")
                    avatar_path = ""

                # 6. 保存数据到Excel和数据库
                db_wechat_name = wechat_name if wechat_name else "无昵称"
                db_avatar_path = avatar_path if avatar_path else "无截图"
                self.excel_exporter.add_row(phone, wechat_name, avatar_path)
                self.db.insert_phone_data(phone, db_wechat_name, db_avatar_path)
                self.log("数据已保存到Excel（图片嵌入）和数据库")

                # 7. 更新进度（每处理一行就更新）
                last_processed_line = line_num
                self.db.update_process_line(self.txt_file_path, last_processed_line)
                self.log(f"进度更新：已处理到第{last_processed_line}行")

                # 8. 批量处理后休息（避免高强度操作）
                batch_rest(processed_count, self.log)

            except Exception as e:
                error_msg = f"处理失败：{str(e)}"
                self.log(error_msg)
                # 错误数据保存
                self.excel_exporter.add_row(phone, "", "")
                self.db.insert_phone_data(phone, error_msg, "处理失败")
                # 更新进度
                self.db.update_process_line(self.txt_file_path, line_num)
                # 失败后延长休息时间（模拟人类排查问题）
                time.sleep(random.uniform(1, 2))

            # 更新UI进度
            self.update_progress(processed_count, self.total_valid_lines)
            # 操作间隔随机化
            time.sleep(random.uniform(OPERATE_DELAY_MIN * 1.5, OPERATE_DELAY_MAX * 1.5))

        # 任务结束：保存Excel文件
        excel_path = self.excel_exporter.save()
        if excel_path:
            self.log(f"Excel表格保存成功：{excel_path}（图片嵌入+专业美化）")
        else:
            self.log("Excel表格保存失败！")

        self.log("\n===== 所有任务处理完成 =====")
        self.is_running = False
        self.root.after(0, self.reset_ui)
        messagebox.showinfo("完成",
                            f"任务处理完成！\n共处理{processed_count}个手机号\n存储路径：{self.save_folder}\nExcel表格已实现图片嵌入+专业美化")

    def stop_execution(self):
        """停止执行"""
        if messagebox.askyesno("确认", "是否停止当前执行？已处理进度会自动保存！"):
            self.is_running = False
            self.log("正在停止执行...")
            self.stop_btn.config(state=tk.DISABLED)

    def reset_ui(self):
        """重置UI状态"""
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.update_progress(0, self.total_valid_lines)


if __name__ == "__main__":
    # 解决高DPI屏幕界面模糊问题（Windows）
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass

    root = tk.Tk()
    app = WeChatMonitorGUI(root)
    root.mainloop()