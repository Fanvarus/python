import requests
import random
import time
import os
from openpyxl import Workbook

# ===================== 内置配置 =====================
CONFIG = {
    "api": {
        "login_url": "http://123.56.58.202:8085/user/login",
        "balance_url": "http://123.56.58.202:8085/profit/profitcanwithdraw",
        "bill_list_url": "http://123.56.58.202:8085/profit/list",
        "verify_code_url": "",
        "timeout": 10,
        "verify_code_expire": 10
    },
    "accounts": [
        {
            "username": "超凡威视",
            "password": "525231314."
        },
        {
            "username": "塘厦益雅贸易",
            "password": "112233"
        },
        {
            "username": "小姜安防",
            "password": "Wu5626480"
        },
        {
            "username": "南鲁集镇",
            "password": "525231314."
        }
    ],
    "request_headers": {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "Host": "123.56.58.202:8085",
        "Origin": "http://iot.xiaotaifeng.cn",
        "Referer": "http://iot.xiaotaifeng.cn/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36"
    },
    "data_clean": {
        "balance_fields": ["balance"],
        "bill_fields": ["id", "amount", "profit", "createtime", "account", "productname", "yunyingshang"],
        "output_format": "both",
        "excel_path": "./account_summary.xlsx"
    }
}

# ===================== 全局汇总数据 =====================
SUMMARY_DATA = {
    "total_accounts": len(CONFIG["accounts"]),
    "success_accounts": 0,
    "balance_summary": [],
    "bill_summary": [],
    "error_logs": []
}


# ===================== 工具函数 =====================
def get_nested_value(data, field):
    if not isinstance(data, dict):
        return "-"
    fields = field.split(".")
    val = data
    for f in fields:
        f_lower = f.lower()
        matched_key = None
        for key in val.keys():
            if key.lower() == f_lower:
                matched_key = key
                break
        if matched_key is None:
            return "-"
        val = val[matched_key]
        if val is None:
            return "-"
    return val


# ===================== 核心登录类 =====================
class LoginHandler:
    def __init__(self):
        self.api_config = CONFIG["api"]
        self.request_headers = CONFIG["request_headers"]
        self.session = requests.Session()
        self.session.headers.update(self.request_headers)
        self.token = None
        self.username = None
        self.login_status = False

    def auto_get_verify_code(self):
        """自动获取验证码（不打印具体内容）"""
        if self.api_config["verify_code_url"].strip():
            try:
                resp = self.session.get(self.api_config["verify_code_url"], timeout=self.api_config["timeout"])
                resp.raise_for_status()
                code = resp.json().get("code")
                if code and len(code) == 4:
                    print(f"[自动验证码] 接口获取成功")
                    return code
            except Exception as e:
                print(f"[自动验证码] 接口获取失败，使用前端生成逻辑")

        # 前端生成4位数字验证码，不打印具体值
        code = "".join([str(random.randint(0, 9)) for _ in range(4)])
        print(f"[自动验证码] 前端生成成功")
        return code

    def login(self, username, password):
        """自动登录（不显示验证码）"""
        self.username = username
        verify_code = self.auto_get_verify_code()
        print(f"[登录准备] 账号：{username}，验证码已自动填充")  # 移除验证码具体数值

        try:
            login_data = {"username": username, "password": password}
            resp = self.session.post(
                self.api_config["login_url"],
                json=login_data,
                timeout=self.api_config["timeout"]
            )
            resp.raise_for_status()
            result = resp.json()
            print(f"[登录响应] 状态：{result.get('message')}")

            if result.get("code") == "0" and result.get("message") == "登录成功":
                self.token = result.get("data", {}).get("token")
                if self.token:
                    self.login_status = True
                    self.session.headers["X-Token"] = self.token
                    return True, {"msg": "登录成功", "token": self.token}
            return False, {"msg": f"登录失败：{result.get('message', '未知错误')}"}
        except Exception as e:
            return False, {"msg": f"登录异常：{str(e)}"}

    def query_balance(self):
        if not self.login_status:
            return False, {"msg": "未登录", "data": None}
        try:
            resp = self.session.get(self.api_config["balance_url"], timeout=self.api_config["timeout"])
            resp.raise_for_status()
            raw_data = resp.json()
            print(f"[余额查询] 账号：{self.username}，原始数据：{raw_data}")

            if raw_data.get("code") == "0" and isinstance(raw_data.get("data"), (int, float)):
                return True, raw_data["data"]
            return False, {"msg": "余额格式异常", "data": raw_data}
        except Exception as e:
            return False, {"msg": f"余额查询失败：{str(e)}", "data": None}

    def query_bill_list(self, page=1, limit=10):
        if not self.login_status:
            return False, {"msg": "未登录", "data": None}
        try:
            params = {
                "paytype": "", "account": "", "productid": "", "name": "",
                "page": page, "limit": limit, "sort": "-d.ID"
            }
            resp = self.session.get(
                self.api_config["bill_list_url"],
                params=params,
                timeout=self.api_config["timeout"]
            )
            resp.raise_for_status()
            raw_data = resp.json()
            print(f"[账单查询] 账号：{self.username}，总条数：{raw_data.get('data', {}).get('total', 0)}")

            if raw_data.get("code") == "0" and "data" in raw_data and "items" in raw_data["data"]:
                return True, raw_data["data"]["items"]
            return False, {"msg": "账单格式异常", "data": raw_data}
        except Exception as e:
            return False, {"msg": f"账单查询失败：{str(e)}", "data": None}

    def logout(self):
        self.login_status = False
        self.token = None
        username = self.username
        self.username = None
        self.session.headers.pop("X-Token", None)
        print(f"[退出登录] 账号：{username}")


# ===================== 数据清洗 =====================
def clean_balance_data(balance_value, username):
    if balance_value is None or not isinstance(balance_value, (int, float)):
        print(f"[余额清洗] 账号：{username}，无有效余额")
        return None

    cleaned = {"username": username, "balance": balance_value}
    print(f"[余额清洗] 账号：{username}，清洗后：{cleaned}")
    return cleaned


def clean_bill_data(bill_list, username):
    if not isinstance(bill_list, list):
        print(f"[账单清洗] 账号：{username}，无有效账单列表")
        return []

    cleaned_list = []
    fields = CONFIG["data_clean"]["bill_fields"]
    for bill in bill_list:
        if not isinstance(bill, dict):
            continue
        bill_cleaned = {"username": username}
        for field in fields:
            bill_cleaned[field] = get_nested_value(bill, field)
        cleaned_list.append(bill_cleaned)

    print(f"[账单清洗] 账号：{username}，清洗后条数：{len(cleaned_list)}")
    return cleaned_list


def export_excel():
    wb = Workbook()
    dc = CONFIG["data_clean"]

    ws_balance = wb.active
    ws_balance.title = "余额汇总"
    if SUMMARY_DATA["balance_summary"]:
        headers = ["username"] + dc["balance_fields"]
        ws_balance.append(headers)
        for item in SUMMARY_DATA["balance_summary"]:
            ws_balance.append([item.get(h, "-") for h in headers])
    else:
        ws_balance.append(["无有效余额数据"])

    ws_bill = wb.create_sheet("账单汇总")
    if SUMMARY_DATA["bill_summary"]:
        headers = ["username"] + dc["bill_fields"]
        ws_bill.append(headers)
        for item in SUMMARY_DATA["bill_summary"]:
            ws_bill.append([item.get(h, "-") for h in headers])
    else:
        ws_bill.append(["无有效账单数据"])

    ws_error = wb.create_sheet("错误日志")
    if SUMMARY_DATA["error_logs"]:
        ws_error.append(["username", "error_type", "error_msg", "time"])
        for err in SUMMARY_DATA["error_logs"]:
            ws_error.append([err["username"], err["error_type"], err["error_msg"], err["time"]])
    else:
        ws_error.append(["无错误日志"])

    wb.save(dc["excel_path"])
    print(f"\n📄 Excel已保存：{os.path.abspath(dc['excel_path'])}")


def generate_summary():
    print("\n" + "=" * 120)
    print("📊 多账户数据汇总报告")
    print("=" * 120)

    print(f"\n📈 核心统计：")
    print(f"   总账户数：{SUMMARY_DATA['total_accounts']}")
    print(f"   登录成功数：{SUMMARY_DATA['success_accounts']}")
    print(f"   有效余额账户数：{len(SUMMARY_DATA['balance_summary'])}")
    print(f"   有效账单总数：{len(SUMMARY_DATA['bill_summary'])}")

    print(f"\n💰 余额详情：")
    total_balance = 0.0
    for item in SUMMARY_DATA["balance_summary"]:
        bal = item["balance"]
        print(f"   账号：{item['username']} | 余额：{bal:.2f} 元")
        total_balance += bal
    print(f"   🎯 所有账户总余额：{total_balance:.2f} 元")

    print(f"\n📋 账单统计：")
    if SUMMARY_DATA["bill_summary"]:
        bill_by_user = {}
        for bill in SUMMARY_DATA["bill_summary"]:
            user = bill["username"]
            bill_by_user[user] = bill_by_user.get(user, []) + [bill]

        for user, bills in bill_by_user.items():
            total_profit = sum(float(b.get("profit", 0)) for b in bills if b.get("profit") != "-")
            print(f"   账号：{user} | 账单数：{len(bills)} | 总利润：{total_profit:.2f} 元")
    else:
        print("   无有效账单数据")

    if SUMMARY_DATA["error_logs"]:
        print(f"\n❌ 错误详情：")
        for err in SUMMARY_DATA["error_logs"]:
            print(f"   账号：{err['username']} | 类型：{err['error_type']} | 信息：{err['error_msg']}")


# ===================== 主程序 =====================
def process_single_account(account):
    username = account["username"]
    password = account["password"]
    error_log = {
        "username": username,
        "error_type": "",
        "error_msg": "",
        "time": time.strftime("%Y-%m-%d %H:%M:%S")
    }

    handler = LoginHandler()
    login_ok, login_res = handler.login(username, password)
    if not login_ok:
        error_log["error_type"] = "登录失败"
        error_log["error_msg"] = login_res["msg"]
        SUMMARY_DATA["error_logs"].append(error_log)
        print(f"❌ 账号【{username}】登录失败：{login_res['msg']}")
        return

    print(f"✅ 账号【{username}】登录成功")
    SUMMARY_DATA["success_accounts"] += 1

    balance_ok, balance_data = handler.query_balance()
    if balance_ok:
        cleaned_balance = clean_balance_data(balance_data, username)
        if cleaned_balance:
            SUMMARY_DATA["balance_summary"].append(cleaned_balance)
    else:
        error_log["error_type"] = "余额查询失败"
        error_log["error_msg"] = balance_data["msg"]
        SUMMARY_DATA["error_logs"].append(error_log)
        print(f"❌ 账号【{username}】余额查询失败：{balance_data['msg']}")

    bill_ok, bill_data = handler.query_bill_list()
    if bill_ok:
        cleaned_bills = clean_bill_data(bill_data, username)
        if cleaned_bills:
            SUMMARY_DATA["bill_summary"].extend(cleaned_bills)
    else:
        error_log["error_type"] = "账单查询失败"
        error_log["error_msg"] = bill_data["msg"]
        SUMMARY_DATA["error_logs"].append(error_log)
        print(f"❌ 账号【{username}】账单查询失败：{bill_data['msg']}")

    handler.logout()
    time.sleep(1)


def main():
    print("===== 多账户自动登录+数据清洗+汇总系统 =====")
    print(f"📌 开始时间：{time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📌 账户数量：{len(CONFIG['accounts'])}")
    print("===========================================\n")

    for idx, account in enumerate(CONFIG["accounts"], 1):
        username = account["username"]
        print(f"\n{'=' * 100}")
        print(f"[处理进度 {idx}/{len(CONFIG['accounts'])}] 账号：{username}")
        print(f"{'=' * 100}")
        process_single_account(account)

    generate_summary()
    if CONFIG["data_clean"]["output_format"] in ["excel", "both"]:
        export_excel()

    print(f"\n🎉 全部处理完成！结束时间：{time.strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("⚠️  安装依赖库...")
        os.system("pip install requests openpyxl -q")
        import openpyxl

    main()