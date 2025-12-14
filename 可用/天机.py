import re
import json
import time
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from colorama import Fore
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
# 关闭urllib3的HTTPS警告
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ======================== 1. 内置配置（替代外部ini文件）========================
CONFIG = {
    "common": {
        "save_root_path": "桌面",  # 结果保存根路径
        "request_timeout": 15,  # 请求超时时间（秒）
        "platform_delay": 1,  # 账号间延迟（秒）
        "query_all_bills": False,  # 是否查询全部账单（False=仅第一页）
        "bill_page_size": 10  # 单次账单查询条数
    },
    "tianji": {
        "base_url": "https://sys.szlaina.com",
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
        "referer": "https://sys.szlaina.com/Index/index",
        # 账号配置（格式：账号名 = 加密密码）
        "accounts": {
            "Wdy": "90535de091e878a11a3e1724ab22bc10",
            "CFWS": "a71a5ba407b3e4333d1a89689779446b",
            "九五AA": "a71a5ba407b3e4333d1a89689779446b",
            "臻鼎视界": "a71a5ba407b3e4333d1a89689779446b",
            "JCKJ": "a71a5ba407b3e4333d1a89689779446b",
            "hmr": "a71a5ba407b3e4333d1a89689779446b",
            "dengweiqiang": "a71a5ba407b3e4333d1a89689779446b",
            "wangyingqi": "a71a5ba407b3e4333d1a89689779446b",
            "晨阳科技": "a71a5ba407b3e4333d1a89689779446b",
            "弘毅威视": "a71a5ba407b3e4333d1a89689779446b"
        }
    }
}

# ======================== 2. 全局数据容器（指定存储容器）========================
# 余额容器：存储所有账号的余额汇总数据
BALANCE_CONTAINER = []
# 账单容器：存储所有账号的账单明细数据（按账号分组）
BILL_CONTAINER = {}


# ======================== 工具函数 ========================
def get_save_path(common_cfg):
    """获取文件保存路径"""
    if common_cfg["save_root_path"].lower() == "桌面":
        from os.path import expanduser
        desktop = Path(expanduser("~")) / "Desktop"
    else:
        desktop = Path(common_cfg["save_root_path"])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_path = desktop / f"平台查询结果_{timestamp}"
    save_path.mkdir(parents=True, exist_ok=True)
    return save_path


def init_excel_style(ws, headers, column_widths):
    """初始化Excel表格样式（美化）"""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = column_widths[col - 1]

    content_font = Font(name='微软雅黑', size=10)
    content_align = Alignment(horizontal='center', vertical='center')
    return thin_border, content_font, content_align


def generate_excel():
    """生成包含天机平台结果的Excel（从全局容器读取数据）"""
    save_path = get_save_path(CONFIG["common"])
    excel_path = save_path / f"平台余额与账单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()

    # 天机余额表（从BALANCE_CONTAINER读取）
    tj_balance_ws = wb.active
    tj_balance_ws.title = "天机平台-余额"
    tj_balance_headers = ["平台", "账号", "原始余额", "汇总余额", "账单数", "状态"]
    tj_balance_widths = [10, 20, 15, 15, 8, 20]
    border, font, align = init_excel_style(tj_balance_ws, tj_balance_headers, tj_balance_widths)
    for row_idx, res in enumerate(BALANCE_CONTAINER, 2):
        for col_idx, key in enumerate(tj_balance_headers, 1):
            cell = tj_balance_ws.cell(row=row_idx, column=col_idx, value=res[key])
            cell.font = font
            cell.alignment = align
            cell.border = border

    # 天机账单表（从BILL_CONTAINER读取）
    all_bills = []
    for account, bills in BILL_CONTAINER.items():
        all_bills.extend(bills)
    if all_bills:
        tj_bill_ws = wb.create_sheet(title="天机平台-账单")
        tj_bill_headers = list(all_bills[0].keys()) if all_bills else []
        tj_bill_widths = [15] * len(tj_bill_headers) if tj_bill_headers else []
        if tj_bill_headers:
            border, font, align = init_excel_style(tj_bill_ws, tj_bill_headers, tj_bill_widths)
            for row_idx, bill in enumerate(all_bills, 2):
                for col_idx, key in enumerate(tj_bill_headers, 1):
                    cell = tj_bill_ws.cell(row=row_idx, column=col_idx, value=bill[key])
                    cell.font = font
                    cell.alignment = align
                    cell.border = border

    wb.save(excel_path)
    print(f"✅ Excel文件已保存：{excel_path}")


def generate_txt_report():
    """生成天机平台TXT报告（从全局容器读取数据）"""
    save_path = get_save_path(CONFIG["common"])
    txt_path = save_path / f"平台汇总报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(f"多平台查询汇总报告（{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}）\n")
        f.write("=" * 50 + "\n\n")

        # 天机平台汇总（从BALANCE_CONTAINER读取）
        f.write("【天机平台】\n")
        f.write(f"账号总数：{len(BALANCE_CONTAINER)}\n")
        total_raw = sum([r["原始余额"] for r in BALANCE_CONTAINER])
        total_summary = sum([r["汇总余额"] for r in BALANCE_CONTAINER])
        f.write(f"原始总余额（含负数）：{total_raw:.2f} 元\n")
        f.write(f"汇总总余额（负数计0）：{total_summary:.2f} 元\n")
        f.write("详细账号信息：\n")
        for res in BALANCE_CONTAINER:
            f.write(
                f"  - {res['账号']}：原始余额 {res['原始余额']:.2f} 元，汇总余额 {res['汇总余额']:.2f} 元，状态：{res['状态']}\n")
        f.write("\n" + "=" * 50 + "\n")

    print(f"✅ TXT报告已保存：{txt_path}")


# ======================== 账单翻译工具函数 ========================
def translate_bill_field(field_name):
    """账单字段名翻译"""
    field_map = {
        "bill_detail_id": "账单ID",
        "company_id": "公司ID",
        "income_money": "收入金额",
        "cost_id": "支出类型ID",
        "cost_money": "成本金额",
        "fee": "手续费",
        "cost_name": "交易类型",
        "create_time": "创建时间（时间戳）",
        "trans_time": "交易时间（时间戳）",
        "bill_status": "账单状态",
        "withdraw_num": "提现单号",
        "iccid": "物联网卡ICCID号",
        "seller_id": "商家/销售ID",
        "remarks": "备注",
        "second_operator_code": "二级运营商编码",
        "table_name": "数据表名",
        "table_id": "数据表ID",
        "create_time_format": "交易时间",
        "trans_time_format": "实际交易时间",
        "bill_time_format": "账单时间"
    }
    return field_map.get(field_name, field_name)


def translate_bill_value(field_name, value):
    """账单字段值翻译/格式化"""
    # 空值处理
    if value == "" or value is None:
        return "无"

    # 交易类型翻译（Unicode转中文）
    if field_name == "cost_name":
        trans_map = {
            "\u63d0\u73b0": "提现",
            "\u5145\u503c": "充值",
            "\u6263\u9664": "扣除",
            "\u8fd4\u6b3e": "退款",
            "\u62a5\u9500": "报销"
        }
        return trans_map.get(value, value)

    # 账单状态翻译
    if field_name == "bill_status":
        status_map = {"1": "成功", "0": "失败", "2": "处理中"}
        return status_map.get(str(value), f"未知状态({value})")

    # 时间戳转格式化时间
    if field_name in ["create_time", "trans_time"]:
        try:
            timestamp = int(value)
            return f"{timestamp}（{datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')}）"
        except:
            return value

    # 金额字段加单位
    if field_name in ["income_money", "cost_money", "fee"]:
        try:
            return f"{float(value)} 元"
        except:
            return value

    # 默认返回原值
    return value


def print_translated_bill(u_name, bill_json):
    """打印翻译后的账单内容（逐行）"""
    print(f"\n📝 【天机-{u_name}】账单数据（翻译后）：")
    # 提取账单列表
    bill_list = bill_json.get("list", [])
    if not bill_list:
        print(f"  无账单数据")
        return

    # 打印前3条（避免内容过多）
    for idx, bill in enumerate(bill_list[:3], 1):
        print(f"  第{idx}条账单：")
        for field, value in bill.items():
            trans_field = translate_bill_field(field)
            trans_value = translate_bill_value(field, value)
            print(f"    - {trans_field}：{trans_value}")
        if idx < len(bill_list[:3]):
            print("    ---")

    # 提示剩余条数
    if len(bill_list) > 3:
        print(f"  （共{len(bill_list)}条账单，仅展示前3条）")


# ======================== 天机平台客户端 ========================
class TianjiClient:
    def __init__(self):
        self.cfg = CONFIG["tianji"]
        self.common_cfg = CONFIG["common"]
        # 清空全局容器（避免重复运行时数据残留）
        BALANCE_CONTAINER.clear()
        BILL_CONTAINER.clear()

        # 请求头
        self.headers = {
            "User-Agent": self.cfg["user_agent"],
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Referer": self.cfg["referer"],
            "Upgrade-Insecure-Requests": "1"
        }
        # 账单字段映射（用于Excel/容器存储）
        self.BILL_FIELD_MAP = {
            "bill_detail_id": "账单ID",
            "cost_name": "交易类型",
            "profit": "交易金额(利润)",
            "create_time_format": "交易时间",
            "trans_time_format": "实际交易时间",
            "order_no": "订单编号",
            "company_name": "公司名称",
            "income_money": "收入金额",
            "cost_money": "成本金额",
            "fee": "手续费",
            "remarks": "备注",
            "withdraw_num": "提现单号",
            "bill_status": "账单状态"
        }
        # 余额提取正则（支持负数）
        self.BALANCE_PATTERNS = [
            r"余额[:：]\s*(-?\d+\.?\d*)",
            r"利润[:：]\s*(-?\d+\.?\d*)",
            r"¥\s*(-?\d+\.?\d*)",
            r"(-?\d+\.?\d*)\s*元",
            r"可用余额[:：]\s*(-?\d+\.?\d*)"
        ]

    def login_single_account(self, u_name, encrypted_pwd):
        """单个账号登录（3次重试）"""
        session = requests.Session()
        login_url = f"{self.cfg['base_url']}/Login/doLogin"
        login_data = {
            "u_name": u_name,
            "pwd": encrypted_pwd,
            "encry": "1"
        }

        # 登录重试（最多3次）
        for retry in range(3):
            try:
                resp = session.post(
                    url=login_url,
                    data=login_data,
                    headers=self.headers,
                    verify=False,
                    timeout=self.common_cfg["request_timeout"]
                )
                phpsessid = session.cookies.get("PHPSESSID")
                if phpsessid:
                    print(f"\n✅ 【天机-{u_name}】登录成功！PHPSESSID：{phpsessid}")
                    return session, True
                else:
                    if retry < 2:
                        print(f"\n⚠️ 【天机-{u_name}】登录失败（重试{retry + 1}/3），等待1秒后重试...")
                        time.sleep(1)
                        continue
                    else:
                        print(f"\n❌ 【天机-{u_name}】登录失败（已重试3次）：{resp.text[:200]}")
                        return session, False
            except Exception as e:
                if retry < 2:
                    print(f"\n⚠️ 【天机-{u_name}】登录异常（重试{retry + 1}/3）：{str(e)}，等待1秒后重试...")
                    time.sleep(1)
                    continue
                else:
                    print(f"\n❌ 【天机-{u_name}】登录异常（已重试3次）：{str(e)}")
                    return session, False

    def extract_balance(self, session, u_name):
        """提取余额（正则+BeautifulSoup兜底）"""
        profit_url = f"{self.cfg['base_url']}/Profit/companyProfit"
        try:
            resp = session.get(
                url=profit_url,
                headers=self.headers,
                verify=False,
                timeout=self.common_cfg["request_timeout"]
            )
            resp.encoding = "utf-8"

            if resp.status_code != 200:
                print(f"⚠️ 【天机-{u_name}】余额接口请求失败，状态码：{resp.status_code}")
                return 0.0

            # 正则提取（支持负数）
            html = resp.text
            for pattern in self.BALANCE_PATTERNS:
                match = re.search(pattern, html, re.IGNORECASE)
                if match:
                    balance = float(match.group(1))
                    print(f"📌 【天机-{u_name}】余额提取成功：{balance} 元")
                    return balance

            # BeautifulSoup兜底提取
            soup = BeautifulSoup(html, "html.parser")
            all_text = soup.get_text()
            for pattern in self.BALANCE_PATTERNS:
                match = re.search(pattern, all_text, re.IGNORECASE)
                if match:
                    balance = float(match.group(1))
                    print(f"📌 【天机-{u_name}】余额提取成功（兜底）：{balance} 元")
                    return balance

            # 提取失败：保存HTML排查
            print(f"⚠️ 【天机-{u_name}】未提取到余额，已保存HTML到本地")
            phpsessid = session.cookies.get("PHPSESSID", "unknown")
            html_path = get_save_path(self.common_cfg) / f"companyProfit_{u_name}_{phpsessid}.html"
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"📁 【天机-{u_name}】HTML保存路径：{html_path}")
            return 0.0
        except Exception as e:
            print(f"❌ 【天机-{u_name}】余额提取异常：{str(e)}")
            return 0.0

    def get_bill_detail(self, session, u_name, page=1, limit=10):
        """查询账单详情（分页）"""
        bill_url = f"{self.cfg['base_url']}/Profit/billDetail"
        bill_data = {
            "page": page,
            "limit": limit,
            "start_time": "",
            "end_time": "",
            "type": ""
        }

        # AJAX请求头
        bill_headers = self.headers.copy()
        bill_headers.update({
            "Accept": "*/*",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://sys.szlaina.com/Profit/listBillDetail"
        })

        try:
            resp = session.post(
                url=bill_url,
                data=bill_data,
                headers=bill_headers,
                verify=False,
                timeout=self.common_cfg["request_timeout"]
            )
            resp.encoding = "utf-8"

            if resp.status_code == 200:
                bill_json = resp.json()
                # 打印翻译后的账单内容
                print_translated_bill(u_name, bill_json)

                # 兼容list为空但状态为1的情况
                if bill_json.get("status") in [1, "1"] and bill_json.get("message") == "成功":
                    bill_list = bill_json.get("list", [])
                    print(f"✅ 【天机-{u_name}】提取到 {len(bill_list)} 条账单数据")
                    normalized_bills = self.normalize_bill_data(bill_list)
                    # 将账单存入全局容器（按账号分组）
                    if u_name not in BILL_CONTAINER:
                        BILL_CONTAINER[u_name] = []
                    BILL_CONTAINER[u_name].extend(normalized_bills)
                    return normalized_bills
                else:
                    print(f"❌ 【天机-{u_name}】账单接口状态异常：{bill_json.get('status')}, {bill_json.get('message')}")
                    return []
            else:
                print(f"❌ 【天机-{u_name}】账单接口请求失败，状态码：{resp.status_code}")
                return []
        except json.JSONDecodeError:
            print(f"❌ 【天机-{u_name}】账单接口返回非JSON格式：{resp.text[:200]}")
            return []
        except Exception as e:
            print(f"❌ 【天机-{u_name}】账单查询异常：{str(e)}")
            return []

    def normalize_bill_data(self, bill_list):
        """规范化账单数据（金额转数值，用于存储）"""
        normalized = []
        for bill in bill_list:
            norm_bill = {"账号": bill.get("company_name", "")}
            for en_field, cn_field in self.BILL_FIELD_MAP.items():
                value = bill.get(en_field, "无")
                # 金额字段转数值（便于计算）
                if en_field in ["profit", "income_money", "cost_money", "fee"] and value != "无":
                    try:
                        value = float(value)
                    except:
                        value = 0.0
                # 交易类型翻译（存储时直接存中文）
                if en_field == "cost_name":
                    value = translate_bill_value(en_field, value)
                # 账单状态翻译
                if en_field == "bill_status":
                    value = translate_bill_value(en_field, value)
                norm_bill[cn_field] = value
            normalized.append(norm_bill)
        return normalized

    def get_all_bills(self, session, u_name):
        """查询全部账单（分页遍历）"""
        all_bills = []
        page = 1
        page_size = self.common_cfg["bill_page_size"]

        while True:
            bills = self.get_bill_detail(session, u_name, page, page_size)
            if not bills:
                break
            all_bills.extend(bills)
            # 不查询全部时只取第一页
            if not self.common_cfg["query_all_bills"]:
                break
            page += 1
            time.sleep(0.5)  # 分页延迟
        return len(all_bills)

    def run(self):
        """执行天机平台查询（数据存入全局容器）"""
        print(f"\n{Fore.BLUE}===== 开始查询【天机平台】（共{len(self.cfg['accounts'])}个账号）=====")
        if not self.cfg["accounts"]:
            print(f"{Fore.YELLOW}⚠️ 天机平台无账号配置，跳过")
            return

        for idx, (u_name, encrypted_pwd) in enumerate(self.cfg["accounts"].items()):
            # 账号间延迟
            if idx > 0:
                print(f"\n⏳ 等待{self.common_cfg['platform_delay']}秒后处理下一个账号...")
                time.sleep(self.common_cfg["platform_delay"])

            # 登录
            session, login_ok = self.login_single_account(u_name, encrypted_pwd)
            if not login_ok:
                # 登录失败数据存入余额容器
                balance_data = {
                    "平台": "天机", "账号": u_name, "原始余额": 0.0, "汇总余额": 0.0,
                    "账单数": 0, "状态": "登录失败"
                }
                BALANCE_CONTAINER.append(balance_data)
                continue

            # 提取余额
            raw_balance = self.extract_balance(session, u_name)
            summary_balance = raw_balance if raw_balance >= 0 else 0.0  # 负数计0

            # 查询账单
            bill_count = self.get_all_bills(session, u_name)

            # 余额数据存入全局容器
            balance_data = {
                "平台": "天机", "账号": u_name,
                "原始余额": raw_balance, "汇总余额": summary_balance,
                "账单数": bill_count, "状态": "成功"
            }
            BALANCE_CONTAINER.append(balance_data)

            print(
                f"{Fore.CYAN}📌 【天机-{u_name}】| 原始余额：{raw_balance:.2f} 元 | 汇总余额：{summary_balance:.2f} 元 | 账单数：{bill_count}")

        # 平台汇总
        total_raw = sum([r["原始余额"] for r in BALANCE_CONTAINER])
        total_summary = sum([r["汇总余额"] for r in BALANCE_CONTAINER])
        print(
            f"{Fore.MAGENTA}===== 天机平台汇总 | 原始总余额：{total_raw:.2f} 元 | 汇总总余额：{total_summary:.2f} 元 =====\n")


# ======================== 主程序入口 ========================
def main():
    print(f"{Fore.CYAN}===== 天机平台查询系统启动（{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}）=====\n")

    # 初始化客户端
    tianji_client = TianjiClient()

    # 执行查询（数据自动存入全局容器）
    tianji_client.run()

    # 生成结果文件（从全局容器读取数据）
    print(f"\n{Fore.CYAN}===== 开始生成汇总文件 =====\n")
    generate_excel()
    generate_txt_report()

    # 打印容器数据示例（验证存储效果）
    print(f"\n{Fore.GREEN}===== 数据容器存储示例 =====\n")
    print(f"📊 余额容器前2条数据：{BALANCE_CONTAINER[:2]}")
    print(f"📋 账单容器-晨阳科技账单数：{len(BILL_CONTAINER.get('晨阳科技', []))} 条")

    # 结束提示
    print(f"\n{Fore.GREEN}===== 所有操作完成！结果文件已保存至：{get_save_path(CONFIG['common'])} =====\n")
    print(f"{Fore.YELLOW}⚠️ 数据容器说明：")
    print(f"  - BALANCE_CONTAINER：存储所有账号余额，类型为列表，每个元素是字典")
    print(f"  - BILL_CONTAINER：存储所有账单，类型为字典（账号为键，账单列表为值）")
    print(f"  - 可直接调用这两个容器进行后续数据处理/分析")


if __name__ == "__main__":
    main()