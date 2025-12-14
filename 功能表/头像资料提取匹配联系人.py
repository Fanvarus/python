import pandas as pd
import os
import re
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def normalize_phone(phone):
    """标准化手机号格式，提取纯数字并处理常见格式问题"""
    if pd.isna(phone):
        return ""
    # 转换为字符串并提取所有数字
    phone_str = str(phone)
    digits = re.sub(r'\D', '', phone_str)
    # 处理可能的手机号长度问题
    if len(digits) == 11:
        return digits
    elif len(digits) > 11:
        # 尝试截取最后11位
        return digits[-11:]
    return digits


def get_legal_contacts_map(file_path):
    """从全国资料Excel中获取手机号到法定联系人的映射，增强容错性"""
    try:
        if not os.path.exists(file_path):
            print(f"错误：未找到全国资料文件 - {file_path}")
            return None

        # 尝试读取Excel文件，支持不同格式
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except Exception as e:
            print(f"尝试用openpyxl引擎读取失败，尝试默认引擎: {str(e)}")
            df = pd.read_excel(file_path)

        print(f"成功读取全国资料，共 {len(df)} 行数据")

        # 更灵活地查找手机号列
        phone_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if '手机' in col_lower or '电话' in col_lower:
                phone_col = col
                break

        if phone_col is None:
            print("警告：全国资料中未找到手机号相关列")
            return None
        print(f"使用 '{phone_col}' 作为手机号列")

        # 更灵活地查找法定联系人列
        contact_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if '法定' in col_lower and '联系' in col_lower:
                contact_col = col
                break
        if contact_col is None:
            for col in df.columns:
                col_lower = str(col).lower()
                if '法人' in col_lower or '代表' in col_lower:
                    contact_col = col
                    break

        if contact_col is None:
            print("警告：全国资料中未找到法定联系人相关列")
            return None
        print(f"使用 '{contact_col}' 作为法定联系人列")

        # 处理并创建映射
        df['normalized_phone'] = df[phone_col].apply(normalize_phone)
        # 过滤无效手机号
        valid_entries = df[df['normalized_phone'].str.len() >= 7]
        # 去重，保留最后一个出现的记录
        contact_map = dict(valid_entries.drop_duplicates(
            subset=['normalized_phone'], keep='last')[['normalized_phone', contact_col]].values)

        print(f"成功创建 {len(contact_map)} 条有效的手机号-法定联系人映射")
        return contact_map

    except Exception as e:
        print(f"读取全国资料时出错: {str(e)}")
        return None


def beautify_excel(file_path):
    """美化Excel文件"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active

        # 设置标题行样式
        title_font = Font(bold=True, color="FFFFFF", size=11)
        title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 设置标题行
        for cell in ws[1]:
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = align_center
            cell.border = thin_border

        # 设置数据行样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                # 手机号列右对齐
                if cell.column_letter == get_column_letter(2):  # 手机号是第2列
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # 调整列宽
        column_widths = {
            "A": 25,  # 昵称
            "B": 15,  # 手机号
            "C": 20,  # 法定联系人
            "D": 10,  # 省份
            "E": 15  # 城市
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(file_path)
        print("Excel文件美化完成")

    except Exception as e:
        print(f"美化Excel文件时出错: {str(e)}")


def process_excel_files(folder_path, keywords, legal_contacts_map=None):
    """处理文件夹中的所有Excel文件并汇总结果"""
    all_results = []

    # 可能的昵称列名
    nickname_columns = ["昵称", "微信昵称", "用户昵称", "名字", "姓名", "顾客名", "客户名"]
    # 可能的手机号列名
    phone_columns = ["手机号", "电话", "联系电话", "手机", "电话号码", "手机号码", "有效手机号", "联系手机"]

    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):  # 排除临时文件
            # 从文件名提取省份和城市信息（去掉文件扩展名）
            name_without_ext = os.path.splitext(filename)[0]
            # 按空格分割文件名（处理"省份 城市"格式）
            parts = [p.strip() for p in name_without_ext.split() if p.strip()]

            # 确定省份和城市
            if len(parts) >= 2:
                province, city = parts[0], ' '.join(parts[1:])  # 处理城市名可能包含空格的情况
            else:
                province, city = "未知省份", "未知城市"

            file_path = os.path.join(folder_path, filename)
            print(f"正在处理文件: {filename} -> 省份: {province}, 城市: {city}")

            try:
                # 读取Excel文件中的所有工作表
                xls = pd.ExcelFile(file_path)
                for sheet_name in xls.sheet_names:
                    print(f"  处理工作表: {sheet_name}")

                    # 读取工作表数据
                    df = pd.read_excel(xls, sheet_name=sheet_name)

                    # 查找昵称列
                    nickname_col = next((col for col in nickname_columns if col in df.columns), None)
                    if nickname_col is None:
                        print(f"  警告: 工作表 {sheet_name} 中未找到昵称相关列，跳过此表")
                        continue

                    # 将昵称列转换为字符串，避免类型错误
                    df[nickname_col] = df[nickname_col].astype(str)

                    # 筛选包含任何关键词的行（不区分大小写）
                    mask = df[nickname_col].apply(
                        lambda x: any(keyword.lower() in x.lower() for keyword in keywords)
                    )
                    filtered_df = df[mask].copy()

                    if not filtered_df.empty:
                        # 提取需要的列
                        result_df = pd.DataFrame()
                        result_df["昵称"] = filtered_df[nickname_col]

                        # 查找手机号列
                        phone_col = next((col for col in phone_columns if col in filtered_df.columns), None)

                        if phone_col:
                            # 标准化手机号格式
                            result_df["手机号"] = filtered_df[phone_col].apply(normalize_phone)
                            # 统计有效手机号数量
                            valid_count = sum(1 for p in result_df["手机号"] if len(p) >= 7)
                            print(f"  提取到 {valid_count} 个有效手机号")
                        else:
                            result_df["手机号"] = ""
                            print(f"  警告: 工作表 {sheet_name} 中未找到手机号相关列")

                        # 添加省份和城市信息
                        result_df["省份"] = province
                        result_df["城市"] = city

                        # 添加到总结果
                        all_results.append(result_df)
                        print(f"  找到 {len(filtered_df)} 条匹配记录")
                    else:
                        print(f"  工作表 {sheet_name} 中没有找到包含指定关键词的记录")

            except Exception as e:
                print(f"  处理文件 {filename} 时出错: {str(e)}")

    # 合并所有结果
    if all_results:
        final_df = pd.concat(all_results, ignore_index=True)
        total_records = len(final_df)
        print(f"\n共汇总 {total_records} 条记录")

        # 匹配法定联系人
        if legal_contacts_map:
            # 使用标准化手机号进行匹配
            final_df["normalized_phone"] = final_df["手机号"].apply(normalize_phone)
            final_df["法定联系人"] = final_df["normalized_phone"].map(legal_contacts_map)

            # 填充空值并统计匹配情况
            final_df["法定联系人"] = final_df["法定联系人"].fillna("未匹配")
            matched_count = sum(1 for c in final_df["法定联系人"] if c != "未匹配")
            print(f"成功匹配 {matched_count} 条法定联系人记录 (匹配率: {matched_count / total_records:.2%})")

            # 调整列顺序，移除临时列
            final_df = final_df[["昵称", "手机号", "法定联系人", "省份", "城市"]]
        else:
            final_df = final_df[["昵称", "手机号", "省份", "城市"]]

        return final_df
    else:
        return None


if __name__ == "__main__":
    # 文件夹路径（已保留你设置的路径）
    folder_path = r"G:\优仙 工作\微信头像资料\已经筛选好2"

    # 全国资料文件路径
    legal_contacts_file = r"C:\Users\Administrator\Desktop\全国.xlsx"

    # 筛选关键词（已保留你设置的关键词）
    keywords = ["克丽缇娜","百莲凯"]

    # 获取法定联系人映射
    print("正在加载全国资料中的法定联系人信息...")
    legal_contacts_map = get_legal_contacts_map(legal_contacts_file)

    # 处理文件并获取结果
    print("\n开始处理Excel文件...")
    result_df = process_excel_files(folder_path, keywords, legal_contacts_map)

    # 保存结果到桌面
    if result_df is not None and not result_df.empty:
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        output_file = os.path.join(desktop_path, "汇总结果.xlsx")
        result_df.to_excel(output_file, index=False)

        # 美化Excel文件
        beautify_excel(output_file)

        print(f"\n处理完成！共找到 {len(result_df)} 条匹配记录。")
        print(f"结果已保存至: {output_file}")
    else:
        print("\n未找到任何匹配的记录。")
