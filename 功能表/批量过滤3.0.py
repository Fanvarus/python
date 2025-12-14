import pandas as pd
import re
import os
import sys
import glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from contextlib import contextmanager
from typing import List, Optional, Tuple, Dict, Set

# 配置区域 - 可根据需要修改
base_dir = r'C:\Users\Administrator\Desktop'
filter_filename = '筛选文件过滤词.txt'  # 过滤词文件名，可修改
input_folder = os.path.join(base_dir, 'input-gl')
output_folder = os.path.join(base_dir, 'output-gl')
filtered_output_folder = os.path.join(base_dir, 'output-filtered')  # 被过滤内容的输出目录

# 修复：添加ExcelWriter兼容处理（避免openpyxl版本差异报错）
@contextmanager
def excel_writer(file_path: str) -> pd.ExcelWriter:
    """上下文管理器确保正确释放Excel资源"""
    writer = None
    try:
        # 兼容openpyxl不同版本的引擎参数
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='w')
        yield writer
    except Exception as e:
        raise RuntimeError(f"Excel写入器创建失败: {str(e)}")
    finally:
        if writer is not None:
            try:
                writer.close()
            except:
                writer.save()  # 降级处理，确保文件保存

def load_filter_words(file_path: str) -> Tuple[List[str], Dict[str, List[str]]]:
    """加载过滤词文件（按空格分隔）并去重，返回去重后的列表和重复词字典"""
    encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16', 'utf-16-le']
    all_words = []

    # 读取文件内容
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                content = file.read()
                all_words = [word.strip() for word in content.split() if word.strip()]
                print(f"成功以{encoding}编码读取过滤词文件")
                break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            raise RuntimeError(f"读取过滤词文件失败: {str(e)}")

    if not all_words:
        raise ValueError(f"无法解码文件 {file_path} 或文件内容为空，尝试过的编码: {', '.join(encodings)}")

    # 显示去重前统计信息
    original_count = len(all_words)
    print(f"过滤词去重前: {original_count} 个")

    # 使用字典进行去重（保留原始大小写），并记录重复词
    unique_words = {}
    duplicate_words = {}

    for word in all_words:
        key = word.lower()
        if key not in unique_words:
            unique_words[key] = word
            duplicate_words[key] = [word]
        else:
            duplicate_words[key].append(word)

    # 过滤出有重复的词
    actual_duplicates = {k: v for k, v in duplicate_words.items() if len(v) > 1}

    # 获取去重后的列表
    unique_list = list(unique_words.values())
    final_count = len(unique_list)

    # 显示去重统计和具体重复词
    print(f"发现重复过滤词: {len(actual_duplicates)} 组")
    for key, duplicates in actual_duplicates.items():
        print(f"  - 重复词 '{key}': {duplicates}")
    print(f"过滤词去重后: {final_count} 个")

    return unique_list, actual_duplicates

def clean_phone_number(phone: str) -> str:
    """清理单个手机号"""
    if not phone or pd.isna(phone):
        return ""

    phone = str(phone).strip()
    # 提取所有可能的手机号（修复：支持带区号或分隔符的手机号）
    phones = re.findall(r'1[3-9]\d{9}', re.sub(r'[^\d]', '', phone))
    return phones[0] if phones else ""

def process_phone_column(series: pd.Series) -> pd.Series:
    """处理整个手机号列"""
    return series.apply(clean_phone_number)

def count_chinese_chars(text: str) -> int:
    """计算字符串中的中文字符数量"""
    if not text or pd.isna(text):
        return 0
    return len(re.findall(
        r'[\u4e00-\u9fff\u3400-\u4dbf\U00020000-\U0002a6df\U0002a700-\U0002b73f\U0002b740-\U0002b81f\U0002b820-\U0002ceaf]',
        str(text)))

def filter_companies(df: pd.DataFrame, filter_words: List[str], company_col: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """过滤公司名称，包括过滤词和短名称，返回过滤后的数据和被过滤掉的数据"""
    if company_col not in df.columns:
        return df, pd.DataFrame()

    original_count = len(df)

    # 过滤包含过滤词的公司（修复：空过滤词列表时跳过）
    if filter_words:
        pattern = '|'.join(map(re.escape, filter_words))
        mask_keywords = df[company_col].astype(str).str.contains(
            pattern, case=False, na=False, regex=True
        )
    else:
        mask_keywords = pd.Series([False] * len(df))

    # 过滤公司名称长度在3个中文字以内的
    mask_short_names = df[company_col].apply(
        lambda x: count_chinese_chars(str(x)) <= 3 if pd.notna(x) else False
    )

    # 过滤包含"分公司"的公司
    mask_branch = df[company_col].astype(str).str.contains(
        "分公司", case=False, na=False, regex=False
    )

    # 为被过滤的行添加过滤原因
    df_with_reason = df.copy()
    df_with_reason['过滤原因'] = ''
    df_with_reason.loc[mask_keywords, '过滤原因'] += '包含过滤词; '
    df_with_reason.loc[mask_short_names, '过滤原因'] += '公司名称过短; '
    df_with_reason.loc[mask_branch, '过滤原因'] += '包含分公司; '
    df_with_reason['过滤原因'] = df_with_reason['过滤原因'].str.rstrip('; ')

    # 组合三种过滤条件
    mask_total = mask_keywords | mask_short_names | mask_branch

    # 计算过滤数量
    keyword_filtered = mask_keywords.sum()
    short_name_filtered = mask_short_names.sum()
    branch_filtered = mask_branch.sum()
    total_filtered = mask_total.sum()

    # 打印过滤统计
    print(f"   - 包含过滤词的公司: {keyword_filtered} 个")
    print(f"   - 公司名称≤3个中文字: {short_name_filtered} 个")
    print(f"   - 包含'分公司'的公司: {branch_filtered} 个")
    print(f"   - 总过滤公司数量: {total_filtered} 个")

    return df[~mask_total], df_with_reason[mask_total]

def find_target_column(df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
    """查找包含所有关键词的列"""
    for col in df.columns:
        col_lower = col.lower()
        if all(keyword.lower() in col_lower for keyword in keywords):
            return col
    return None

def format_excel_worksheet(ws):
    """美化Excel工作表格式"""
    # 设置标题行格式
    header_font = Font(name='微软雅黑', size=14, bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据行格式
    data_alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment

    # 调整行高
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

    # 优化列宽调整（修复：处理空值导致的报错）
    max_rows_to_check = min(100, ws.max_row)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col[:max_rows_to_check + 1]:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

def save_filtered_data(filtered_data: Dict[str, pd.DataFrame], output_path: str):
    """保存被过滤掉的数据到Excel文件"""
    if not filtered_data:
        return

    try:
        with excel_writer(output_path) as writer:
            for sheet_name, df in filtered_data.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 美化格式
            wb = writer.book
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                format_excel_worksheet(ws)
        print(f"√ 已保存被过滤数据到: {output_path}")
    except Exception as e:
        print(f"保存被过滤数据失败: {str(e)}")

def process_excel_file(input_path: str, output_path: str, filtered_output_path: str,
                       filter_words: List[str]) -> Tuple[int, Dict[str, int]]:
    """处理单个Excel文件，返回处理的工作表数量和统计信息"""
    processed_sheets = 0
    sheet_stats = {}
    filtered_data = {}

    try:
        # 修复：支持老版本Excel格式（.xls）
        with pd.ExcelFile(input_path, engine='xlrd' if input_path.endswith('.xls') else None) as xls:
            with excel_writer(output_path) as writer:
                for sheet_name in xls.sheet_names:
                    df = xls.parse(sheet_name)
                    original_rows = len(df)
                    print(f"\n正在处理工作表 [{sheet_name}] - 原始行数: {original_rows}")

                    # 公司名称过滤
                    company_col = find_target_column(df, ['企业名称'])
                    filtered_df = pd.DataFrame()

                    if company_col:
                        df, filtered_df = filter_companies(df, filter_words, company_col)
                        filtered_rows = len(df)
                        filtered_count = original_rows - filtered_rows
                        print(f"√ 已过滤公司名称列：{company_col}，过滤掉 {filtered_count} 行")

                        if not filtered_df.empty:
                            filtered_data[sheet_name] = filtered_df
                    else:
                        filtered_rows = original_rows
                        filtered_count = 0
                        print("※ 未找到公司名称列，跳过过滤")

                    # 手机号处理
                    phone_col = find_target_column(df, ['手机号', '有效'])
                    if phone_col:
                        df[phone_col] = process_phone_column(df[phone_col])
                        print(f"√ 已处理手机号列：{phone_col}")
                    else:
                        print("※ 未找到有效手机号列")

                    # 去重处理
                    duplicate_removed = 0
                    if phone_col:
                        before_duplicate = len(df)
                        non_empty_mask = df[phone_col] != ""
                        df_non_empty = df[non_empty_mask]
                        df_empty = df[~non_empty_mask]

                        # 对非空手机号去重
                        duplicates_mask = df_non_empty.duplicated(subset=[phone_col], keep='first')
                        duplicates_removed = df_non_empty[duplicates_mask]
                        if not duplicates_removed.empty:
                            duplicates_removed = duplicates_removed.copy()
                            duplicates_removed['过滤原因'] = duplicates_removed.get('过滤原因', '') + '手机号重复; '
                            if sheet_name in filtered_data:
                                filtered_data[sheet_name] = pd.concat([filtered_data[sheet_name], duplicates_removed])
                            else:
                                filtered_data[sheet_name] = duplicates_removed

                        df_non_empty = df_non_empty.drop_duplicates(subset=[phone_col])
                        df = pd.concat([df_non_empty, df_empty])
                        duplicate_removed = before_duplicate - len(df)
                        print(f"√ 已根据手机号去重，移除重复行: {duplicate_removed} 行")

                        # 公司名称+手机号去重
                        if company_col:
                            before_duplicate_combined = len(df)
                            combined_duplicates_mask = df.duplicated(subset=[company_col, phone_col], keep='first')
                            combined_duplicates_removed = df[combined_duplicates_mask]
                            if not combined_duplicates_removed.empty:
                                combined_duplicates_removed = combined_duplicates_removed.copy()
                                combined_duplicates_removed['过滤原因'] = combined_duplicates_removed.get('过滤原因', '') + '公司名称+手机号重复; '
                                if sheet_name in filtered_data:
                                    filtered_data[sheet_name] = pd.concat([filtered_data[sheet_name], combined_duplicates_removed])
                                else:
                                    filtered_data[sheet_name] = combined_duplicates_removed

                            df = df.drop_duplicates(subset=[company_col, phone_col])
                            additional_removed = before_duplicate_combined - len(df)
                            duplicate_removed += additional_removed
                            print(f"√ 已根据公司名称+手机号去重，额外移除重复行: {additional_removed} 行")
                    elif company_col:
                        # 仅公司名称去重
                        before_duplicate = len(df)
                        duplicates_mask = df.duplicated(subset=[company_col], keep='first')
                        duplicates_removed = df[duplicates_mask]
                        if not duplicates_removed.empty:
                            duplicates_removed = duplicates_removed.copy()
                            duplicates_removed['过滤原因'] = duplicates_removed.get('过滤原因', '') + '公司名称重复; '
                            if sheet_name in filtered_data:
                                filtered_data[sheet_name] = pd.concat([filtered_data[sheet_name], duplicates_removed])
                            else:
                                filtered_data[sheet_name] = duplicates_removed

                        df = df.drop_duplicates(subset=[company_col])
                        duplicate_removed = before_duplicate - len(df)
                        print(f"√ 已根据公司名称去重，移除重复行: {duplicate_removed} 行")
                    else:
                        print("※ 未找到公司名称列和手机号列，跳过去重")

                    # 保存结果
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    processed_sheets += 1

                    # 记录统计
                    sheet_stats[sheet_name] = {
                        'original': original_rows,
                        'filtered': len(df),
                        'removed': filtered_count,
                        'duplicates_removed': duplicate_removed
                    }

                # 美化格式
                wb = writer.book
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    format_excel_worksheet(ws)

        # 保存被过滤数据
        save_filtered_data(filtered_data, filtered_output_path)

    except Exception as e:
        print(f"处理Excel文件时出错: {str(e)}")
        raise

    return processed_sheets, sheet_stats

def process_folder(input_folder: str, output_folder: str, filtered_output_folder: str, filter_words: List[str]):
    """处理整个文件夹中的Excel文件"""
    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(filtered_output_folder, exist_ok=True)

    # 获取所有Excel文件（修复：区分.xlsx和.xls）
    excel_files = glob.glob(os.path.join(input_folder, '*.xlsx')) + glob.glob(os.path.join(input_folder, '*.xls'))

    if not excel_files:
        print(f"在文件夹 {input_folder} 中未找到Excel文件")
        return

    total_files = len(excel_files)
    print(f"\n找到 {total_files} 个Excel文件需要处理")

    for i, input_path in enumerate(excel_files, 1):
        filename = os.path.basename(input_path)
        output_path = os.path.join(output_folder, filename)
        filtered_filename = f"filtered_{filename}"
        filtered_output_path = os.path.join(filtered_output_folder, filtered_filename)

        print(f"\n{'=' * 60}")
        print(f"处理文件 [{i}/{total_files}]: {filename}")
        print(f"输出文件: {output_path}")
        print(f"被过滤数据文件: {filtered_output_path}")

        try:
            sheets_processed, sheet_stats = process_excel_file(
                input_path, output_path, filtered_output_path, filter_words
            )
            print(f"√ 文件处理完成! 共处理 {sheets_processed} 个工作表")

            # 打印统计
            for sheet, stats in sheet_stats.items():
                print(f"  - {sheet}:")
                print(f"     原始行数: {stats['original']}")
                print(f"     过滤后行数: {stats['filtered']}")
                print(f"     过滤移除行数: {stats['removed']}")
                print(f"     去重移除行数: {stats['duplicates_removed']}")
        except Exception as e:
            print(f"处理文件 {filename} 失败: {str(e)}")

    print(f"\n{'=' * 60}")
    print(f"所有文件处理完成!")
    print(f"处理结果保存在: {output_folder}")
    print(f"被过滤数据保存在: {filtered_output_folder}")

def main():
    """主处理程序"""
    try:
        print("=== Excel数据处理程序 ===")
        print("功能说明:")
        print("1. 处理文件夹中的所有Excel文件（支持.xlsx/.xls）")
        print("2. 过滤包含敏感词的公司名称（敏感词来自filter_words.txt，按空格分隔）")
        print("3. 过滤公司名称长度≤3个中文字的公司")
        print("4. 过滤包含'分公司'的公司")
        print("5. 清理有效手机号列（保留第一个有效号码）")
        print("6. 根据公司名称和手机号进行去重")
        print("7. 美化工作表格式（标题加粗加大、内容居中、调整行高列宽）")
        print("8. 输出文件保留原始文件名")
        print("9. 额外生成包含被过滤信息的文件\n")

        # 路径配置说明
        print("路径配置:")
        print(f"  - 基础目录: {base_dir}")
        print(f"  - 过滤词文件: {filter_filename}")
        print(f"  - 输入文件夹: {input_folder}")
        print(f"  - 输出文件夹: {output_folder}")
        print(f"  - 被过滤数据文件夹: {filtered_output_folder}\n")

        # 路径验证
        if not os.path.exists(input_folder):
            raise FileNotFoundError(f"输入文件夹不存在: {input_folder}")

        # 修正路径拼接
        filter_path = os.path.join(base_dir, '模板', filter_filename)
        if not os.path.exists(filter_path):
            os.makedirs(os.path.dirname(filter_path), exist_ok=True)
            with open(filter_path, 'w', encoding='utf-8') as f:
                f.write("")
            print(f"已创建空白过滤词文件: {filter_path}")
            print("请在该文件中添加过滤词，用空格分隔，然后重新运行程序")
            return

        # 加载过滤词
        filter_words, duplicate_words = load_filter_words(filter_path)
        if not filter_words:
            print("警告: 过滤词文件为空，将不会过滤任何公司名称")
        else:
            print(f"已加载 {len(filter_words)} 个过滤词")

        # 处理文件夹
        process_folder(input_folder, output_folder, filtered_output_folder, filter_words)

    except Exception as e:
        print(f"\n错误发生：{str(e)}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()