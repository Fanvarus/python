import requests
import json
import pandas as pd
import os
import re
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import time
import math

# 高德地图API密钥
AMAP_KEY = "35261fa1a7eae9be266b813b7de7b4fa"

# 文件保存路径设置
BASE_DIR = "C:\\Users\\Administrator\\Desktop\\高德商家信息"
RECENT_DIR = os.path.join(BASE_DIR, "最近生成的表")
MASTER_DIR = os.path.join(BASE_DIR, "总表")
MASTER_FILE = os.path.join(MASTER_DIR, "高德商家信息库.xlsx")

# 确保目录存在
os.makedirs(RECENT_DIR, exist_ok=True)
os.makedirs(MASTER_DIR, exist_ok=True)

# 配置参数
API_CALL_DELAY = 1.5  # API调用间隔
MAX_RETRIES = 5  # 最大重试次数
BATCH_SIZE = 30  # 每页获取数量
EMPTY_PAGE_THRESHOLD = 5  # 连续空页阈值


def is_mobile_number(phone):
    """判断是否为手机号（中国大陆11位数字）"""
    return len(phone) == 11 and phone.isdigit() and phone.startswith(('13', '14', '15', '16', '17', '18', '19'))


def split_phone_numbers(phone_str):
    """拆分电话号码字符串"""
    if not phone_str or phone_str in ["未提供", "详情获取失败"]:
        return {"手机号": "", "更多号码": ""}

    # 分割可能的分隔符
    separators = [',', '，', ';', '；', '、', ' ', '/', '\\']
    for sep in separators:
        phone_str = phone_str.replace(sep, '|')

    phones = [p.strip() for p in phone_str.split('|') if p.strip()]

    # 提取手机号
    mobile = ""
    others = []

    for phone in phones:
        cleaned = re.sub(r'[^\d]', '', phone)
        if is_mobile_number(cleaned) and not mobile:
            mobile = phone  # 保留原始格式
        else:
            others.append(phone)

    return {
        "手机号": mobile if mobile else "",
        "更多号码": '; '.join(others) if others else ""
    }


def ensure_string(value):
    """确保值是字符串类型，处理可能的列表或其他类型"""
    if isinstance(value, list):
        return '; '.join(map(str, value))
    elif pd.isna(value):
        return ""
    return str(value).strip()


def load_master_table():
    """加载总表数据"""
    if os.path.exists(MASTER_FILE):
        try:
            df = pd.read_excel(MASTER_FILE)
            # 确保关键列是字符串类型
            for col in ['名称', '地址', '此次筛选的关键字', '城市']:
                if col in df.columns:
                    df[col] = df[col].apply(ensure_string)
            print(f"成功加载总表，当前包含 {len(df)} 条记录")
            return df
        except Exception as e:
            print(f"加载总表失败: {str(e)}，将创建新的总表")
            return pd.DataFrame()
    else:
        print("总表不存在，将在本次查询后创建")
        return pd.DataFrame()


def save_to_master_table(new_data, keyword):
    """将新数据添加到总表，去重处理"""
    master_df = load_master_table()
    new_data_with_keyword = new_data.copy()

    # 为新数据添加关键字列
    new_data_with_keyword["此次筛选的关键字"] = keyword

    # 确保关键列是字符串类型
    for col in ['名称', '地址', '城市']:
        if col in new_data_with_keyword.columns:
            new_data_with_keyword[col] = new_data_with_keyword[col].apply(ensure_string)

    # 合并数据并去重
    if not master_df.empty:
        combined = pd.concat([master_df, new_data_with_keyword])
        combined = combined.drop_duplicates(subset=['名称', '地址'], keep='last')
        added_count = len(combined) - len(master_df)
        print(f"已向总表添加 {added_count} 条新记录，总表现在共有 {len(combined)} 条记录")
    else:
        combined = new_data_with_keyword
        added_count = len(combined)
        print(f"已创建总表并添加 {added_count} 条记录")

    # 保存合并后的数据
    try:
        combined.to_excel(MASTER_FILE, index=False, engine='openpyxl')
        format_excel(MASTER_FILE, is_master=True)
        return added_count
    except Exception as e:
        print(f"保存总表失败: {str(e)}")
        return 0


def search_poi(keyword, city="", page=1, page_size=BATCH_SIZE):
    """搜索兴趣点(POI)，带增强重试机制"""
    url = "https://restapi.amap.com/v3/place/text"
    params = {
        "key": AMAP_KEY,
        "keywords": keyword,
        "city": city,
        "page": page,
        "offset": page_size,
        "output": "json",
        "extensions": "all"
    }

    for retry in range(MAX_RETRIES):
        try:
            time.sleep(API_CALL_DELAY * (1 + retry * 0.5))  # 指数退避策略
            response = requests.get(url, params=params)
            result = json.loads(response.text)

            if result["status"] == "1":
                return result
            else:
                error_msg = f"搜索失败({retry + 1}/{MAX_RETRIES}): {result.get('info', '未知错误')}"
                print(error_msg)
                # 如果是配额不足，尝试等待更长时间
                if "配额" in result.get('info', ''):
                    time.sleep(10)
        except Exception as e:
            print(f"搜索发生错误({retry + 1}/{MAX_RETRIES}): {str(e)}")

    return None


def get_poi_details(poi_id):
    """获取POI详细信息"""
    url = "https://restapi.amap.com/v3/place/detail"
    params = {
        "key": AMAP_KEY,
        "id": poi_id,
        "output": "json"
    }

    for retry in range(MAX_RETRIES):
        try:
            time.sleep(API_CALL_DELAY)
            response = requests.get(url, params=params)
            result = json.loads(response.text)

            if result["status"] == "1":
                return result
            else:
                print(f"获取详情失败({retry + 1}/{MAX_RETRIES}) ID:{poi_id}: {result.get('info', '未知错误')}")
                if "配额" in result.get('info', ''):
                    time.sleep(10)
        except Exception as e:
            print(f"获取详情发生错误({retry + 1}/{MAX_RETRIES}) ID:{poi_id}: {str(e)}")

    return None


def extract_contacts(details, basic_info=None):
    """从POI详情中提取联系方式"""
    if not details or "pois" not in details or len(details["pois"]) == 0:
        if basic_info:
            phone_str = basic_info.get("tel", "详情获取失败")
            phones = split_phone_numbers(phone_str)
            city = ensure_string(
                f"{basic_info.get('pname', '')}{basic_info.get('cityname', '')}{basic_info.get('adname', '')}")
            return {
                "名称": ensure_string(basic_info.get("name", "未知")),
                "地址": ensure_string(basic_info.get("address", "未知")),
                "手机号": phones["手机号"],
                "更多号码": phones["更多号码"],
                "类型": ensure_string(basic_info.get("type", "未知")),
                "坐标": ensure_string(basic_info.get("location", "未知")),
                "城市": city,
                "评分": "未知"
            }
        return None

    poi_info = details["pois"][0]
    phone_str = poi_info.get("tel", "未提供")
    phones = split_phone_numbers(phone_str)
    city = ensure_string(f"{poi_info.get('pname', '')}{poi_info.get('cityname', '')}{poi_info.get('adname', '')}")

    return {
        "名称": ensure_string(poi_info.get("name", "未知")),
        "地址": ensure_string(poi_info.get("address", "未知")),
        "手机号": phones["手机号"],
        "更多号码": phones["更多号码"],
        "类型": ensure_string(poi_info.get("type", "未知")),
        "坐标": ensure_string(poi_info.get("location", "未知")),
        "城市": city,
        "评分": ensure_string(poi_info.get("biz_ext", {}).get("rating", "未评分"))
    }


def get_all_poi_details(keyword, city="", master_df=None):
    """获取所有POI的详细信息，优化API限制处理"""
    all_contacts = []
    page = 1
    total_count = 0
    processed_count = 0
    new_count = 0
    empty_page_count = 0
    api_call_count = 0  # 单独计数

    print(f"正在搜索 {keyword} 在 {city if city else '全国'} 的商家信息...")

    # 先获取第一页结果确定总数量
    search_result = search_poi(keyword, city, page, BATCH_SIZE)
    api_call_count += 1
    if not search_result or "pois" not in search_result:
        print("未找到相关商家信息")
        return [], api_call_count

    total_count = int(search_result.get("count", 0))
    if total_count == 0:
        print("未找到相关商家信息")
        return [], api_call_count

    total_pages = math.ceil(total_count / BATCH_SIZE)
    print(f"共找到 {total_count} 个相关商家，理论上分布在 {total_pages} 页，正在获取详细信息...")

    # 处理所有页面
    while empty_page_count < EMPTY_PAGE_THRESHOLD and len(all_contacts) < total_count:
        # 显示当前进度百分比
        progress = min(int(len(all_contacts) / total_count * 100), 100)
        print(f"处理进度: {progress}% | 第 {page}/{total_pages} 页 | 已获取: {len(all_contacts)}/{total_count} 条")

        # 获取当前页数据
        if page == 1 and search_result:
            current_page_data = search_result
        else:
            current_page_data = search_poi(keyword, city, page, BATCH_SIZE)
            api_call_count += 1

        # 检查当前页是否有效
        if not current_page_data or "pois" not in current_page_data or len(current_page_data["pois"]) == 0:
            empty_page_count += 1
            print(f"第 {page} 页无有效数据，已连续 {empty_page_count} 页")

            # 尝试跳过几页再试
            if empty_page_count >= 2:
                jump_pages = min(3, total_pages - page)
                if jump_pages > 0:
                    print(f"尝试跳过 {jump_pages} 页")
                    page += jump_pages
                    continue
        else:
            empty_page_count = 0
            # 处理当前页的POI
            for poi in current_page_data["pois"]:
                processed_count += 1

                # 检查是否已在总表中存在
                poi_name = ensure_string(poi.get("name", "未知"))
                poi_address = ensure_string(poi.get("address", "未知"))

                if master_df is not None and not master_df.empty:
                    exists = not master_df[(master_df["名称"] == poi_name) &
                                           (master_df["地址"] == poi_address)].empty
                    if exists:
                        continue

                # 获取详细信息
                details = get_poi_details(poi["id"])
                api_call_count += 1
                contact = extract_contacts(details, poi)

                if contact:
                    all_contacts.append(contact)
                    new_count += 1

        page += 1
        # 防止无限循环
        if page > total_pages + 10:
            print("已超出理论页数过多，停止查询")
            break

    print(f"处理完成，共获取到 {len(all_contacts)} 条新的有效商家信息")
    if len(all_contacts) < total_count:
        print(f"注意：实际获取数量({len(all_contacts)})少于理论数量({total_count})")

    return all_contacts, api_call_count


def format_excel(file_path, is_master=False):
    """优化表格美化，自适应列宽"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 设置字体和边框
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        content_font = Font(name="微软雅黑", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 设置表头
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = openpyxl.styles.PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 设置内容单元格格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = content_font
                cell.border = thin_border
                # 根据列内容设置对齐方式
                if cell.column_letter in ['B', 'D', 'I']:  # 地址、更多号码、关键字列自动换行
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # 智能调整列宽（根据内容长度）
        column_widths = {}
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row)):  # 取前100行计算
            for cell in row:
                col_letter = cell.column_letter
                # 计算内容长度，中文算2个字符，英文和数字算1个
                if cell.value:
                    value = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in value) + 2  # 加2个字符的边距
                    if col_letter not in column_widths or length > column_widths[col_letter]:
                        column_widths[col_letter] = length

        # 设置最小和最大列宽限制
        min_widths = {
            'A': 15, 'B': 25, 'C': 13, 'D': 20,
            'E': 20, 'F': 18, 'G': 15, 'H': 6, 'I': 15
        }
        max_widths = {
            'B': 40, 'D': 35, 'I': 25
        }

        for col, width in column_widths.items():
            # 应用最小宽度限制
            if col in min_widths:
                width = max(width, min_widths[col])
            else:
                width = max(width, 10)

            # 应用最大宽度限制
            if col in max_widths:
                width = min(width, max_widths[col])

            ws.column_dimensions[col].width = width

        # 冻结表头
        ws.freeze_panes = 'A2'

        # 总表添加筛选功能
        if is_master:
            ws.auto_filter.ref = ws.dimensions

        wb.save(file_path)
        print(f"表格美化完成：{os.path.basename(file_path)}")
    except Exception as e:
        print(f"美化Excel表格时出错: {str(e)}")


def save_to_recent_excel(contacts, keyword, city=""):
    """保存联系信息到最近生成的表文件夹，仅包含指定城市的信息"""
    if not contacts and city:
        print(f"没有找到 {city} 的相关商家信息")
        return None

    # 加载总表数据
    master_df = load_master_table()

    # 创建新数据DataFrame并添加关键字
    new_data_df = pd.DataFrame(contacts)
    for col in ['名称', '地址', '城市']:
        if col in new_data_df.columns:
            new_data_df[col] = new_data_df[col].apply(ensure_string)
    new_data_df["此次筛选的关键字"] = keyword

    # 合并总表数据和新数据，然后按城市筛选
    if not master_df.empty:
        # 合并数据并去重
        combined_df = pd.concat([master_df, new_data_df]).drop_duplicates(
            subset=['名称', '地址'], keep='last')

        # 筛选出指定城市的数据（如果指定了城市）
        if city:
            # 处理可能的城市名称变体，进行模糊匹配
            city_lower = city.lower()
            combined_df = combined_df[
                combined_df['城市'].apply(lambda x: city_lower in x.lower())
            ]
            print(f"生成的文件包含 {len(combined_df)} 条 {city} 的记录（总表中该城市数据 + 新增数据）")
        else:
            print(f"生成的文件包含 {len(combined_df)} 条全国记录（总表数据 + 新增数据）")
    else:
        # 没有总表数据，直接使用新数据
        combined_df = new_data_df
        print(f"生成的文件包含 {len(combined_df)} 条记录")

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    city_suffix = f"_{city}" if city else ""
    filename = f"高德地图商家信息_{keyword}{city_suffix}_{timestamp}.xlsx"
    file_path = os.path.join(RECENT_DIR, filename)

    # 保存到Excel
    combined_df.to_excel(file_path, index=False, engine='openpyxl')

    # 美化表格
    format_excel(file_path)

    print(f"\n已成功将信息保存至：")
    print(f"文件路径：{file_path}")
    return file_path


def main():
    print("高德地图商家信息提取与导出工具（按城市筛选版）")
    print("-" * 60)

    # 加载总表数据
    master_df = load_master_table()

    # 获取搜索参数
    keyword = input("请输入要搜索的商家关键字: ").strip()
    if not keyword:
        print("关键字不能为空！")
        return

    city = input("请输入城市名(可选，直接回车则全国搜索): ").strip()

    # 获取所有新的商家信息
    all_contacts, api_call_count = get_all_poi_details(keyword, city, master_df)

    # 保存到总表
    if all_contacts:
        added_count = save_to_master_table(pd.DataFrame(all_contacts), keyword)
    else:
        added_count = 0
        print("没有新的商家信息需要添加到总表")

    # 保存到最近生成的表
    if not all_contacts and (
            master_df.empty or (city and len(master_df[master_df['城市'].str.contains(city, case=False)]) == 0)):
        print(f"未获取到 {city if city else '全国'} 的任何商家信息，不生成新文件")
        return

    save_to_recent_excel(all_contacts, keyword, city)

    # 显示汇总信息
    print("\n" + "-" * 60)
    print("汇总信息：")
    print(f"总表当前共包含 {len(master_df) + added_count} 条记录")
    print(f"本次查询新增 {added_count} 条记录")
    print(f"本次操作共调用API {api_call_count} 次")

    # 统计有手机号的商家数量
    if not master_df.empty:
        master_has_mobile = sum(1 for _, row in master_df.iterrows() if str(row.get("手机号", "")).strip())
        print(f"总表中包含手机号的商家：{master_has_mobile} 家")

    if all_contacts:
        new_has_mobile = sum(1 for contact in all_contacts if contact.get("手机号", "").strip())
        print(f"本次新增包含手机号的商家：{new_has_mobile} 家")


if __name__ == "__main__":
    main()
