import os
import pandas as pd
import time
from tqdm import tqdm
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# 配置日志记录
def setup_logger(log_file):
    """配置日志记录器"""
    logger = logging.getLogger('excel_merger')
    logger.setLevel(logging.INFO)

    # 创建文件处理器
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)

    # 创建控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # 创建格式化器并添加到处理器
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # 清除旧的处理器并添加新的处理器
    if logger.hasHandlers():
        logger.handlers.clear()
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


# 读取文件函数，保留格式信息
def read_file(file_path, logger, required_columns):
    """读取不同格式的表格文件并保留格式信息，只保留需要的列"""
    file_ext = os.path.splitext(file_path)[1].lower()

    try:
        if file_ext in ['.xlsx', '.xls']:
            # 读取Excel文件数据，使用只读模式提高性能
            df = pd.read_excel(file_path, dtype=object)

            # 只保留需要的列，不存在的列会被忽略
            existing_columns = [col for col in required_columns if col in df.columns]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                logger.warning(f"文件 {os.path.basename(file_path)} 缺少以下列: {missing_columns}")
            df = df[existing_columns]

            # 尝试加载格式信息，使用非只读模式
            format_info = None
            try:
                # 使用非只读模式加载工作簿以获取格式信息
                wb_format = load_workbook(file_path, read_only=False, data_only=True)
                ws_format = wb_format.active

                # 获取列格式
                column_formats = {}
                for col_idx in range(1, ws_format.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    if ws_format[f'{col_letter}1'].value in required_columns:  # 只关注需要的列
                        if ws_format[f'{col_letter}1'].number_format:
                            column_formats[col_letter] = ws_format[f'{col_letter}1'].number_format

                # 获取表头样式
                header_styles = {}
                for col_idx in range(1, ws_format.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    header_value = ws_format[f'{col_letter}1'].value
                    if header_value in required_columns:  # 只关注需要的列
                        cell = ws_format[f'{col_letter}1']
                        header_style = {
                            'font': {
                                'name': cell.font.name,
                                'size': cell.font.size,
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'color': cell.font.color.rgb if cell.font.color else None
                            },
                            'alignment': {
                                'horizontal': cell.alignment.horizontal,
                                'vertical': cell.alignment.vertical,
                                'wrap_text': cell.alignment.wrap_text
                            },
                            'border': {
                                'left': cell.border.left.style,
                                'right': cell.border.right.style,
                                'top': cell.border.top.style,
                                'bottom': cell.border.bottom.style
                            },
                            'fill': {
                                'fill_type': cell.fill.fill_type,
                                'start_color': cell.fill.start_color.rgb if cell.fill.start_color else None
                            }
                        }
                        header_styles[header_value] = header_style  # 使用表头值作为键，而不是列字母

                wb_format.close()
                format_info = {'column_formats': column_formats, 'header_styles': header_styles}
            except Exception as e:
                logger.warning(f"获取文件 {os.path.basename(file_path)} 格式信息失败: {str(e)}")

            return df, format_info

        elif file_ext == '.csv':
            # 读取CSV文件，无法保留格式信息
            df = pd.read_csv(file_path, dtype=object)
            # 只保留需要的列
            existing_columns = [col for col in required_columns if col in df.columns]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                logger.warning(f"文件 {os.path.basename(file_path)} 缺少以下列: {missing_columns}")
            df = df[existing_columns]
            return df, None

        elif file_ext == '.tsv':
            # 读取TSV文件，无法保留格式信息
            df = pd.read_csv(file_path, sep='\t', dtype=object)
            # 只保留需要的列
            existing_columns = [col for col in required_columns if col in df.columns]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                logger.warning(f"文件 {os.path.basename(file_path)} 缺少以下列: {missing_columns}")
            df = df[existing_columns]
            return df, None
        else:
            logger.error(f"不支持的文件格式: {file_ext}")
            return None, None

    except Exception as e:
        logger.error(f"读取文件 {os.path.basename(file_path)} 失败: {str(e)}")
        return None, None


# 应用格式到Excel工作表
def apply_formatting(ws, format_info, headers):
    """应用格式信息到工作表"""
    if not format_info:
        return

    # 应用表头样式
    header_styles = format_info.get('header_styles', {})
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        cell = ws[f'{column_letter}1']

        if header in header_styles:  # 使用表头值查找样式
            style = header_styles[header]

            # 应用字体样式
            font = style['font']
            cell.font = Font(
                name=font['name'],
                size=font['size'],
                bold=font['bold'],
                italic=font['italic'],
                color=font['color']
            )

            # 应用对齐方式
            alignment = style['alignment']
            cell.alignment = Alignment(
                horizontal=alignment['horizontal'],
                vertical=alignment['vertical'],
                wrap_text=alignment['wrap_text']
            )

            # 应用边框
            border = style['border']
            cell.border = Border(
                left=Side(style=border['left']),
                right=Side(style=border['right']),
                top=Side(style=border['top']),
                bottom=Side(style=border['bottom'])
            )

            # 应用填充
            fill = style['fill']
            if fill['fill_type']:
                cell.fill = PatternFill(
                    fill_type=fill['fill_type'],
                    start_color=fill['start_color']
                )

    # 应用列格式
    column_formats = format_info.get('column_formats', {})
    for col_letter, fmt in column_formats.items():
        for row in range(2, ws.max_row + 1):
            ws[f'{col_letter}{row}'].number_format = fmt


# 主函数
def main():
    # 定义需要保留的表头
    REQUIRED_COLUMNS = [
        '企业名称',
        '登记状态',
        '法定代表人',
        '所属省份',
        '所属城市',
        '所属区县',
        '有效手机号'
    ]

    # 文件夹路径
    folder_path = r"C:\Users\Administrator\Desktop\汇总原件"
    base_dir = os.path.dirname(folder_path)

    # 创建输出文件夹
    output_dir = os.path.join(base_dir, "汇总结果")
    os.makedirs(output_dir, exist_ok=True)

    # 输出文件路径
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"汇总表格_{timestamp}.xlsx")
    log_path = os.path.join(output_dir, f"处理日志_{timestamp}.log")

    # 支持的文件扩展名
    SUPPORTED_EXTS = ['.xlsx', '.xls', '.csv', '.tsv']

    # 配置日志
    logger = setup_logger(log_path)

    # 记录开始时间
    start_time = time.time()
    total_rows = 0
    processed_files = 0
    failed_files = []

    logger.info(f"开始处理文件夹: {folder_path}")
    logger.info(f"输出文件: {output_path}")
    logger.info(f"将只保留以下表头: {REQUIRED_COLUMNS}")

    # 获取所有文件路径
    all_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if os.path.splitext(file)[1].lower() in SUPPORTED_EXTS:
                all_files.append(os.path.join(root, file))

    logger.info(f"找到 {len(all_files)} 个表格文件")

    if not all_files:
        logger.info("没有找到可处理的表格文件，程序退出")
        return

    # 创建ExcelWriter对象
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        current_sheet = 1
        current_sheet_rows = 0

        # 设置表头为我们需要的列
        headers = REQUIRED_COLUMNS

        # 创建进度条
        with tqdm(total=len(all_files), desc="处理进度") as pbar:
            # 处理所有文件
            for file_path in all_files:
                file_start_time = time.time()

                # 读取文件并只保留需要的列
                df, format_info = read_file(file_path, logger, REQUIRED_COLUMNS)
                if df is None:
                    failed_files.append(file_path)
                    pbar.update(1)
                    continue

                # 确保列的顺序与我们定义的一致
                existing_columns = df.columns.tolist()
                # 添加缺失的列并填充为None
                for col in REQUIRED_COLUMNS:
                    if col not in existing_columns:
                        df[col] = None
                # 重新排序列以匹配我们需要的顺序
                df = df[REQUIRED_COLUMNS]

                # 检查是否需要创建新工作表
                if current_sheet_rows + len(df) > 800000:  # 每个工作表最大行数
                    current_sheet += 1
                    current_sheet_rows = 0
                    sheet_name = f'Sheet{current_sheet}'
                    logger.info(f"创建新工作表: {sheet_name}")

                # 写入数据到当前工作表
                sheet_name = f'Sheet{current_sheet}'
                startrow = current_sheet_rows + 1  # +1 是因为Excel行从1开始

                # 如果是新工作表，需要写入表头
                write_header = (current_sheet_rows == 0)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=write_header, startrow=startrow)

                # 应用格式
                ws = writer.sheets[sheet_name]
                if write_header:  # 如果是新工作表，应用表头格式
                    apply_formatting(ws, format_info, headers)
                else:  # 否则只应用列格式
                    if format_info and 'column_formats' in format_info:
                        column_formats = format_info['column_formats']
                        for col_letter, fmt in column_formats.items():
                            for row in range(startrow + 1, startrow + len(df) + 1):
                                ws[f'{col_letter}{row}'].number_format = fmt

                # 更新统计信息
                file_rows = len(df)
                total_rows += file_rows
                current_sheet_rows += file_rows
                processed_files += 1

                # 打印处理信息
                file_size = os.path.getsize(file_path) / (1024 * 1024)
                file_elapsed = time.time() - file_start_time
                pbar.set_postfix({
                    "文件": os.path.basename(file_path),
                    "行数": f"{file_rows:,}",
                    "大小": f"{file_size:.2f}MB",
                    "耗时": f"{file_elapsed:.2f}s"
                })
                pbar.update(1)

    # 输出汇总统计
    total_elapsed = time.time() - start_time
    logger.info("\n===== 处理完成 =====")
    logger.info(f"输出文件: {output_path}")
    logger.info(f"处理文件数: {processed_files}/{len(all_files)}")
    logger.info(f"失败文件数: {len(failed_files)}")
    logger.info(f"汇总总行数: {total_rows:,}")
    logger.info(f"工作表数量: {current_sheet}")
    logger.info(f"总耗时: {total_elapsed:.2f} 秒 ({total_elapsed / 60:.2f} 分钟)")
    logger.info(f"处理速度: {total_rows / total_elapsed:.0f} 行/秒")

    # 输出失败文件列表
    if failed_files:
        logger.info("\n===== 处理失败的文件 =====")
        for file in failed_files:
            logger.info(f"  {os.path.basename(file)}")

    logger.info("✅ 已按要求只保留指定的表头信息")


if __name__ == "__main__":
    main()
