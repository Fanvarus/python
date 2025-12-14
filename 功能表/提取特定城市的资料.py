import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl


def create_border():
    """创建清晰的单元格边框样式"""
    thin = Side(style='medium')  # 使用中等粗细边框，更清晰
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def beautify_excel(file_path):
    """美化Excel文件格式，仅表头加粗，优化列宽"""
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 定义字体样式
        header_font = Font(name="微软雅黑", size=12, bold=True, color="004B23")  # 表头加粗
        data_font = Font(name="微软雅黑", size=12)  # 数据不加粗

        # 设置标题行样式
        header_fill = PatternFill(start_color="C2E0C6", end_color="C2E0C6", fill_type="solid")

        # 处理标题行
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = create_border()

        # 针对特定列设置最小宽度
        column_min_widths = {
            '公司名称': 30,
            '手机号码': 15,
            '法定代表人': 15,
            '所属城市': 12,
            '所属区县': 15,
            '注册地址': 40,
            '所属省份': 10
        }

        # 自动调整列宽，结合最小宽度设置
        header_row = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(header_row, 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            # 检查该列是否有预设的最小宽度
            min_width = column_min_widths.get(header, 12)

            # 计算该列最大内容长度
            for cell in ws[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # 确定最终宽度
            adjusted_width = max(max_length + 2, min_width)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 处理数据行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font  # 数据行不加粗
                cell.border = create_border()
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 调整行高
        ws.row_dimensions[1].height = 25  # 标题行更高
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 22  # 数据行高度

        # 保存美化后的文件
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"美化Excel文件时出错: {str(e)}")
        return False


def get_column_letter(col_idx):
    """获取列索引对应的字母"""
    try:
        from openpyxl.utils import get_column_letter as xl_get_col_letter
        return xl_get_col_letter(col_idx)
    except:
        # 简单的列字母转换实现
        letters = []
        while col_idx > 0:
            col_idx -= 1
            letters.append(chr(col_idx % 26 + ord('A')))
            col_idx = col_idx // 26
        return ''.join(reversed(letters))


def extract_data():
    # 原文件路径
    input_file = r"C:\Users\Administrator\Desktop\全国资料.xlsx"

    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误：文件 '{input_file}' 不存在，请检查路径是否正确。")
        return

    # 确保输出文件夹存在
    output_dir = r"G:\优仙 工作\指定城市资料生成历史记录"
    os.makedirs(output_dir, exist_ok=True)

    try:
        # 读取Excel文件（只读取一次，提高效率）
        df = pd.read_excel(input_file)

        # 打印表头信息
        print("\n检测到的表头：")
        print(df.columns.tolist())
        print()  # 空行分隔

        while True:
            # 让用户选择提取方式
            print("请选择提取方式：")
            print("1 - 按城市提取")
            print("2 - 按省份提取")
            print("3 - 结束操作")

            choice = input("请输入选项(1/2/3)：").strip()

            if choice == '3':
                print("已结束操作。")
                return
            elif choice not in ['1', '2']:
                print("无效选项，请重新输入。")
                continue

            # 获取用户输入的名称（城市或省份）
            if choice == '1':
                data_type = "城市"
                column_name = "所属城市"
            else:
                data_type = "省份"
                column_name = "所属省份"

            name = input(f"请输入要提取的{data_type}名：").strip()

            if not name:
                print(f"{data_type}名不能为空，请重新输入。")
                continue

            # 提取指定数据
            filtered_df = df[df[column_name] == name]

            if filtered_df.empty:
                print(f"没有找到{data_type}'{name}'的数据。")
                continue  # 继续下一次循环，让用户重新输入

            # 获取当前月日
            current_date = datetime.now().strftime("%m%d")

            # 构建新文件名和路径
            if choice == '1':
                # 按城市提取时，文件名仍使用省份+城市+日期
                province = filtered_df['所属省份'].iloc[0]
                output_filename = f"{province}{name}{current_date}.xlsx"
            else:
                # 按省份提取时，文件名使用省份+日期
                output_filename = f"{name}{current_date}.xlsx"

            output_file = os.path.join(output_dir, output_filename)

            # 保存为新的Excel文件
            filtered_df.to_excel(output_file, index=False, sheet_name=name)

            # 美化Excel文件
            print("正在美化文件格式...")
            if beautify_excel(output_file):
                print("文件美化成功！")
            else:
                print("文件美化失败，将使用默认格式。")

            print(f"成功提取数据！文件已保存至：\n{output_file}")
            print(f"共提取到 {len(filtered_df)} 条记录。\n")

    except Exception as e:
        print(f"\n处理过程中发生错误：{str(e)}")


if __name__ == "__main__":
    # 确保导入必要的库
    try:
        import openpyxl
        import pandas
    except ImportError as e:
        missing_lib = str(e).split("'")[1]
        print(f"检测到缺少必要的库 {missing_lib}，正在安装...")
        os.system(f"pip install {missing_lib}")
        if missing_lib == "openpyxl":
            import openpyxl
        else:
            import pandas

    extract_data()
    input("\n操作完成，按回车键退出...")
