#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多平台数据校验抓取工具 - 纯真实数据库版
功能：仅针对真实数据库进行校验、抓取，无任何模拟数据逻辑
"""

import pandas as pd
import os
import configparser
import time
import traceback
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import platform
import sys

# ======================== 核心配置（用户必须修改这里） ========================
# 真实数据库配置（请严格按照实际环境修改）
DB_CONFIG = {
    "tianji": {
        "host": "192.168.1.100",  # 替换：天机数据库IP/域名（禁止用localhost，填真实IP）
        "port": 3306,  # 替换：数据库端口
        "user": "root",  # 替换：数据库用户名
        "password": "your_password",  # 替换：数据库密码
        "database": "tianji_bill",  # 替换：天机账单数据库名
        "test_account": "wdy"  # 替换：用于测试的天机账号
    },
    "xiaotaifeng": {
        "host": "192.168.1.100",  # 替换：小台风数据库IP/域名
        "port": 3306,  # 替换：数据库端口
        "user": "root",  # 替换：数据库用户名
        "password": "your_password",  # 替换：数据库密码
        "database": "xiaotaifeng_bill",  # 替换：小台风账单数据库名
        "test_account": "超凡威视"  # 替换：用于测试的小台风账号
    },
    "miaoyue": {
        "host": "192.168.1.100",  # 替换：妙月数据库IP/域名
        "port": 3306,  # 替换：数据库端口
        "user": "root",  # 替换：数据库用户名
        "password": "your_password",  # 替换：数据库密码
        "database": "miaoyue_bill",  # 替换：妙月账单数据库名
        "test_account": "jiweishidai"  # 替换：用于测试的妙月账号
    }
}

# 账号配置文件路径（自动生成）
ACCOUNT_CONFIG_PATH = os.path.join(os.path.expanduser("~"), "Desktop", "account_config.ini")
# 输出路径（固定桌面）
OUTPUT_PATH = os.path.join(os.path.expanduser("~"), "Desktop")

# ======================== 固定配置（无需修改） ========================
# 字段映射（与数据库表字段对应）
FIELD_MAPPING = {
    "tianji": {
        "order_no": "订单号", "iccid": "ICCID", "card_number": "卡号",
        "trans_time": "交易时间", "income_money": "售价（元）", "cost_money": "成本（元）",
        "profit": "佣金（元）", "company_name": "客户名称", "order_name": "套餐/产品名称",
        "second_operator_name": "运营商", "remarks": "备注", "income_type": "收入类型"
    },
    "xiaotaifeng": {
        "orderid": "订单号", "iccid": "ICCID", "msisdn": "卡号",
        "purchasetime": "交易时间", "amount": "售价（元）", "cost_money": "成本（元）",
        "profit": "佣金（元）", "custom": "客户名称", "account": "客户名称备用",
        "mpname": "套餐/产品名称", "yunyingshang": "运营商", "incometype": "收入类型",
        "remark": "备注"
    },
    "miaoyue": {
        "orderNo": "订单号", "cardIccid": "ICCID", "cardNumber": "卡号",
        "settleTime": "交易时间", "salePrice": "售价（元）", "costPrice": "成本（元）",
        "commissionAmount": "佣金基数（元）", "deCommissionAmount": "佣金扣减（元）",
        "extraAmount": "佣金额外（元）", "final_profit": "佣金（元）",
        "customer_name": "客户名称", "orderContent": "套餐/产品名称",
        "operator": "运营商", "order_type": "收入类型", "mark": "备注"
    }
}

# 样式配置
STYLE = {
    "primary_color": "FF0078D7",
    "font_main": "微软雅黑" if platform.system() == "Windows" else "Arial",
    "font_number": "Consolas" if platform.system() == "Windows" else "Monaco",
    "border_thin": Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
}

# 全局数据存储
DATA = {
    "check": {
        "platform": {"tianji": {"login": False, "capture": False, "data": False, "error": ""},
                     "xiaotaifeng": {"login": False, "capture": False, "data": False, "error": ""},
                     "miaoyue": {"login": False, "capture": False, "data": False, "error": ""}},
        "account": {"tianji": {}, "xiaotaifeng": {}, "miaoyue": {}}
    },
    "bills": {"tianji": {}, "xiaotaifeng": {}, "miaoyue": {}}
}


# ======================== 工具函数 ========================
def init_env():
    """初始化运行环境（解决中文编码）"""
    if platform.system() == "Windows":
        # 强制控制台UTF8编码
        os.system("chcp 65001 >nul 2>&1")
        os.environ["PYTHONIOENCODING"] = "utf-8"

    # 创建输出目录
    os.makedirs(OUTPUT_PATH, exist_ok=True)

    # 打印启动信息
    print("=" * 70)
    print(f"🚀 纯真实数据库版数据校验工具 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📁 输出路径：{OUTPUT_PATH}")
    print("⚠️  请确认已修改DB_CONFIG中的数据库配置！")
    print("=" * 70)


def print_log(msg: str, level: str = "info"):
    """打印日志"""
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {"info": "[ℹ️ ]", "success": "[✅ ]", "warning": "[⚠️ ]", "error": "[❌ ]", "process": "[🔧 ]"}
    print(f"{ts} {prefix.get(level, '[ℹ️ ]')} {msg}")


def create_account_config():
    """创建账号配置文件（首次运行自动生成）"""
    if os.path.exists(ACCOUNT_CONFIG_PATH):
        return

    # 初始账号列表（可手动修改配置文件添加/删除）
    init_accounts = {
        "TIANJI_ACCOUNTS": ["wdy", "cfws", "九五aa", "臻鼎视界", "jckj", "hmr", "dengweiqiang", "wangyingqi",
                            "晨阳科技", "弘毅威视"],
        "XIAOTAIENG_ACCOUNTS": ["超凡威视", "南鲁集镇", "小姜安防", "塘厦益雅贸易"],
        "MIAOYUE_ACCOUNTS": ["jiweishidai", "huangfangyi", "wudeyou", "wudeyou01", "蓝硕商贸科技", "tpchengze",
                             "wu0321", "chaofan"]
    }

    config = configparser.ConfigParser()
    for section, accounts in init_accounts.items():
        config[section] = {acc: "" for acc in accounts}

    with open(ACCOUNT_CONFIG_PATH, "w", encoding="utf-8-sig") as f:
        config.write(f)

    print_log(f"首次运行，自动生成账号配置文件：{ACCOUNT_CONFIG_PATH}", "info")
    print_log("可手动修改该文件添加/删除需要校验的账号", "warning")


def load_accounts() -> Dict[str, List[str]]:
    """加载需要校验的账号列表"""
    create_account_config()

    config = configparser.ConfigParser()
    config.read(ACCOUNT_CONFIG_PATH, encoding="utf-8-sig")

    accounts = {
        "tianji": list(config["TIANJI_ACCOUNTS"].keys()) if "TIANJI_ACCOUNTS" in config else [],
        "xiaotaifeng": list(config["XIAOTAIENG_ACCOUNTS"].keys()) if "XIAOTAIENG_ACCOUNTS" in config else [],
        "miaoyue": list(config["MIAOYUE_ACCOUNTS"].keys()) if "MIAOYUE_ACCOUNTS" in config else []
    }

    print_log(
        f"加载账号 - 天机：{len(accounts['tianji'])} | 小台风：{len(accounts['xiaotaifeng'])} | 妙月：{len(accounts['miaoyue'])}",
        "info")
    return accounts


# ======================== 数据库核心操作 ========================
def connect_db(platform: str) -> Optional[Any]:
    """连接数据库（终极编码修复）"""
    try:
        import pymysql
        cfg = DB_CONFIG[platform]

        # 解决中文乱码+连接拒绝问题的核心配置
        conn = pymysql.connect(
            host=cfg["host"],
            port=cfg["port"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset="utf8",
            use_unicode=True,
            connect_timeout=10,  # 超时时间
            init_command="""
                SET NAMES utf8mb4;
                SET CHARACTER SET utf8mb4;
                SET character_set_connection=utf8mb4;
                SET character_set_results=utf8mb4;
                SET character_set_client=utf8mb4;
            """
        )
        return conn
    except ImportError:
        err_msg = "未安装pymysql，请执行：pip install pymysql"
        DATA["check"]["platform"][platform]["error"] = err_msg
        print_log(err_msg, "error")
        return None
    except Exception as e:
        err_msg = f"连接失败：{str(e)}"
        DATA["check"]["platform"][platform]["error"] = err_msg

        # 针对性错误提示
        if "10061" in str(e):
            err_msg += " → 原因：数据库服务未启动/IP端口错误/防火墙拦截"
        elif "Access denied" in str(e):
            err_msg += " → 原因：用户名/密码错误/无数据库访问权限"
        elif "Unknown database" in str(e):
            err_msg += " → 原因：数据库名错误/数据库不存在"

        print_log(f"{platform} - {err_msg}", "error")
        return None


def check_db_structure(conn: Any, platform: str) -> str:
    """检查数据库表结构"""
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        # 1. 查找账单表
        table_candidates = [f"{platform}_bills", f"{platform}_bill", f"bill_{platform}"]
        target_table = ""
        for tbl in table_candidates:
            cursor.execute(f"SHOW TABLES LIKE '{tbl}'")
            if cursor.fetchone():
                target_table = tbl
                break

        if not target_table:
            return f"未找到账单表（尝试表名：{','.join(table_candidates)}）"

        # 2. 检查核心字段
        cursor.execute(f"DESCRIBE {target_table}")
        db_fields = [col["Field"] for col in cursor.fetchall()]
        core_fields = list(FIELD_MAPPING[platform].keys())[:5]
        missing_fields = [f for f in core_fields if f not in db_fields]

        if missing_fields:
            return f"表{target_table}缺少核心字段：{','.join(missing_fields)}"

        # 3. 测试数据读取
        test_acc = DB_CONFIG[platform]["test_account"]
        cursor.execute(f"SELECT * FROM {target_table} WHERE username = %s LIMIT 1", (test_acc,))
        if not cursor.fetchone():
            return f"测试账号{test_acc}在表{target_table}中无数据"

        cursor.close()
        return ""  # 无错误返回空字符串
    except Exception as e:
        return f"表结构检查失败：{str(e)}"


def fetch_account_data(conn: Any, platform: str, username: str) -> Optional[List[Dict]]:
    """抓取单个账号的账单数据"""
    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        # 查找目标表
        table_candidates = [f"{platform}_bills", f"{platform}_bill", f"bill_{platform}"]
        target_table = ""
        for tbl in table_candidates:
            cursor.execute(f"SHOW TABLES LIKE '{tbl}'")
            if cursor.fetchone():
                target_table = tbl
                break

        if not target_table:
            raise Exception(f"未找到账单表")

        # 抓取数据（限制100条，可根据需要调整）
        cursor.execute(f"SELECT * FROM {target_table} WHERE username = %s LIMIT 100", (username,))
        raw_data = cursor.fetchall()
        cursor.close()

        if not raw_data:
            return None

        # 标准化数据（处理编码+字段映射）
        standardized = []
        for row in raw_data:
            bill = {}
            for raw_field, std_field in FIELD_MAPPING[platform].items():
                val = row.get(raw_field, "")

                # 强制UTF8编码处理，避免中文乱码
                if isinstance(val, str):
                    bill[std_field] = val.encode("utf-8", errors="ignore").decode("utf-8")
                # 金额字段格式化
                elif "（元）" in std_field and val is not None:
                    bill[std_field] = round(float(val), 2)
                else:
                    bill[std_field] = val if val is not None else ""

            # 补充平台名称
            bill["平台"] = {"tianji": "天机", "xiaotaifeng": "小台风", "miaoyue": "妙月"}[platform]
            standardized.append(bill)

        return standardized
    except Exception as e:
        print_log(f"{platform}-{username} 数据抓取失败：{str(e)[:50]}", "error")
        return None


# ======================== 全流程校验 ========================
def run_full_check():
    """执行全流程数据库校验和数据抓取"""
    accounts = load_accounts()

    # 逐个平台处理
    for platform in ["tianji", "xiaotaifeng", "miaoyue"]:
        print_log(f"\n开始处理【{platform}】平台", "process")

        # 1. 数据库连接校验
        conn = connect_db(platform)
        if not conn:
            continue
        DATA["check"]["platform"][platform]["login"] = True
        print_log(f"{platform} - 数据库连接成功", "success")

        # 2. 表结构校验
        struct_err = check_db_structure(conn, platform)
        if struct_err:
            DATA["check"]["platform"][platform]["error"] = struct_err
            print_log(f"{platform} - 表结构校验失败：{struct_err}", "error")
            conn.close()
            continue
        DATA["check"]["platform"][platform]["capture"] = True
        print_log(f"{platform} - 表结构校验通过", "success")

        # 3. 账号数据抓取
        DATA["check"]["platform"][platform]["data"] = True
        DATA["bills"][platform] = {}

        for username in accounts[platform]:
            bills = fetch_account_data(conn, platform, username)
            if bills:
                DATA["bills"][platform][username] = bills
                DATA["check"]["account"][platform][username] = {"ok": True, "count": len(bills)}
                print_log(f"{platform}-{username} - 成功抓取{len(bills)}条数据", "success")
            else:
                DATA["check"]["account"][platform][username] = {"ok": False, "error": "无数据/抓取失败"}
                print_log(f"{platform}-{username} - 无有效数据", "warning")

        conn.close()


# ======================== 报告生成 ========================
def generate_check_report():
    """生成数据库校验报告"""
    # 1. TXT报告（详细日志）
    txt_path = os.path.join(OUTPUT_PATH, f"数据库校验报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("多平台数据库校验报告\n")
        f.write(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n\n")

        # 平台级校验结果
        f.write("📊 平台级校验结果：\n")
        for platform, res in DATA["check"]["platform"].items():
            f.write(f"\n{platform}平台：\n")
            f.write(f"  - 数据库连接：{'✅ 成功' if res['login'] else '❌ 失败'}\n")
            f.write(f"  - 表结构校验：{'✅ 成功' if res['capture'] else '❌ 失败'}\n")
            f.write(f"  - 数据抓取：{'✅ 成功' if res['data'] else '❌ 失败'}\n")
            if res["error"]:
                f.write(f"  - 错误信息：{res['error']}\n")

        # 账号级结果
        f.write("\n\n📋 账号级数据抓取结果：\n")
        for platform, accs in DATA["check"]["account"].items():
            f.write(f"\n{platform}平台账号：\n")
            for acc, res in accs.items():
                if res["ok"]:
                    f.write(f"  - {acc}：✅ 成功（{res['count']}条）\n")
                else:
                    f.write(f"  - {acc}：❌ 失败（{res.get('error', '未知错误')}）\n")

    # 2. Excel报告（可视化）
    excel_path = os.path.join(OUTPUT_PATH, f"数据库校验报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb = Workbook()

    # 平台级工作表
    ws_platform = wb.active
    ws_platform.title = "平台级校验"
    headers = ["平台", "数据库连接", "表结构校验", "数据抓取", "错误信息"]
    ws_platform.append(headers)

    for platform, res in DATA["check"]["platform"].items():
        ws_platform.append([
            platform,
            "成功" if res["login"] else "失败",
            "成功" if res["capture"] else "失败",
            "成功" if res["data"] else "失败",
            res["error"][:100] if res["error"] else ""
        ])

    # 账号级工作表
    ws_account = wb.create_sheet(title="账号级抓取")
    headers = ["平台", "账号名称", "抓取状态", "数据条数", "错误信息"]
    ws_account.append(headers)

    for platform, accs in DATA["check"]["account"].items():
        for acc, res in accs.items():
            ws_account.append([
                platform,
                acc,
                "成功" if res["ok"] else "失败",
                res.get("count", 0) if res["ok"] else 0,
                res.get("error", "")[:100]
            ])

    # 美化Excel
    for ws in [ws_platform, ws_account]:
        # 设置列宽
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

        # 表头样式
        for cell in ws[1]:
            cell.font = Font(name=STYLE["font_main"], size=11, bold=True, color="white")
            cell.fill = PatternFill(start_color=STYLE["primary_color"], end_color=STYLE["primary_color"],
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = STYLE["border_thin"]

        # 数据行样式
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center")
                cell.border = STYLE["border_thin"]
                # 失败项标红
                if cell.value == "失败":
                    cell.font = Font(color="red")

    wb.save(excel_path)
    print_log(f"\n校验报告生成完成：", "success")
    print_log(f"TXT报告：{txt_path}", "info")
    print_log(f"Excel报告：{excel_path}", "info")


def generate_bill_excel():
    """生成账单数据Excel"""
    # 汇总所有有效数据
    all_bills = []
    platform_bills = {"天机账单": [], "小台风账单": [], "妙月账单": []}

    for platform, accs in DATA["bills"].items():
        for acc, bills in accs.items():
            all_bills.extend(bills)
            if platform == "tianji":
                platform_bills["天机账单"].extend(bills)
            elif platform == "xiaotaifeng":
                platform_bills["小台风账单"].extend(bills)
            elif platform == "miaoyue":
                platform_bills["妙月账单"].extend(bills)

    # 无数据时生成提示文件
    if not all_bills:
        print_log("无有效账单数据，生成排查指引", "warning")
        excel_path = os.path.join(OUTPUT_PATH, f"数据抓取失败指引_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="排查指引")

        ws["A1"] = "📢 未抓取到任何账单数据"
        ws["A2"] = "请按以下步骤排查："
        ws["A3"] = "1. 检查DB_CONFIG中的数据库IP/端口/用户名/密码是否正确"
        ws["A4"] = "2. 确认数据库服务已启动，且网络可通（关闭防火墙/开放端口）"
        ws["A5"] = "3. 确认数据库名和表名正确，表字段与FIELD_MAPPING匹配"
        ws["A6"] = "4. 确认账号配置文件中的账号在数据库中存在且有数据"
        ws["A7"] = f"排查时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # 样式
        ws.column_dimensions["A"].width = 60
        ws["A1"].font = Font(size=14, bold=True, color=STYLE["primary_color"])
        for row in range(2, 8):
            ws[f"A{row}"].font = Font(size=11, name=STYLE["font_main"])

        wb.save(excel_path)
        print_log(f"排查指引已生成：{excel_path}", "info")
        return

    # 有数据时生成汇总Excel
    excel_path = os.path.join(OUTPUT_PATH, f"账单数据汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # 多平台汇总表
        df_all = pd.DataFrame(all_bills)
        df_all = df_all.sort_values(by=["平台", "客户名称", "交易时间"], ascending=[True, True, False])
        df_all.to_excel(writer, sheet_name="多平台汇总", index=False)

        # 各平台单独表
        for sheet_name, bills in platform_bills.items():
            if bills:
                df_platform = pd.DataFrame(bills)
                df_platform = df_platform.sort_values(by=["客户名称", "交易时间"], ascending=[True, False])
                df_platform.to_excel(writer, sheet_name=sheet_name, index=False)

    # 美化Excel
    wb = load_workbook(excel_path)
    for ws in wb.worksheets:
        # 列宽自适应
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            header = ws.cell(row=1, column=col).value
            ws.column_dimensions[col_letter].width = 20 if "（元）" in str(header) else 18

        # 表头样式
        for cell in ws[1]:
            cell.font = Font(name=STYLE["font_main"], size=11, bold=True, color="white")
            cell.fill = PatternFill(start_color=STYLE["primary_color"], end_color=STYLE["primary_color"],
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = STYLE["border_thin"]

        # 数据行样式
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = STYLE["border_thin"]
                # 金额列右对齐+数字字体
                if "（元）" in str(ws.cell(row=1, column=col).value):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.font = Font(name=STYLE["font_number"])

    wb.save(excel_path)
    print_log(f"账单数据Excel生成完成：{excel_path}", "success")


# ======================== 主函数 ========================
def main():
    """主执行函数"""
    try:
        # 初始化环境
        init_env()

        # 执行全流程校验
        run_full_check()

        # 生成校验报告
        generate_check_report()

        # 生成账单Excel
        generate_bill_excel()

        # 完成提示
        print_log("\n" + "=" * 70, "info")
        print_log("🎉 全流程执行完成！所有文件已保存至桌面", "success")
        print_log("=" * 70, "info")

    except Exception as e:
        # 全局异常捕获
        err_msg = f"程序执行异常：{str(e)}\n{traceback.format_exc()}"
        print_log(err_msg, "error")

        # 生成错误报告
        err_path = os.path.join(OUTPUT_PATH, f"程序错误报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(err_path, "w", encoding="utf-8") as f:
            f.write(f"错误时间：{datetime.now()}\n")
            f.write(f"错误信息：{err_msg}\n")

        print_log(f"错误报告已保存：{err_path}", "error")

    # Windows防闪退
    if platform.system() == "Windows":
        input("\n按Enter键退出...")


if __name__ == "__main__":
    sys.setrecursionlimit(10000)
    main()