# multi_platform_query_system_v2.py
import re
import json
import time
import os
import pandas as pd
import numpy as np
import configparser
import concurrent.futures
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
import requests
from requests import Session
from bs4 import BeautifulSoup
from colorama import Fore, init, Style, Back
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import warnings
import logging
from fake_useragent import UserAgent

warnings.filterwarnings('ignore')

# 初始化colorama
init(autoreset=True)


# ======================== 配置文件读取 ========================
class ConfigManager:
    """配置文件管理器"""

    def __init__(self, config_path: str = "multi_platform_config_v2.ini"):
        self.config_path = config_path
        self.config = configparser.ConfigParser()

    # 在ConfigManager类的load_config方法中修改
    def load_config(self):
        """加载配置文件"""
        if not os.path.exists(self.config_path):
            self.create_default_config()

        self.config.read(self.config_path, encoding='utf-8')

        # 读取通用配置
        common_config = {
            'output_path': self.config.get('SETTINGS', 'output_path', fallback='桌面'),
            'bill_page_size': self.config.getint('SETTINGS', 'bill_page_size', fallback=50),
            'query_all_bills': self.config.getboolean('SETTINGS', 'query_all_bills', fallback=False),
            'max_workers': self.config.getint('SETTINGS', 'max_workers', fallback=3),
            'enable_threading': self.config.getboolean('SETTINGS', 'enable_threading', fallback=True),
            'request_timeout': self.config.getint('SETTINGS', 'request_timeout', fallback=30),
            'platform_delay': self.config.getfloat('SETTINGS', 'platform_delay', fallback=2.0),
            'days_for_recent': self.config.getint('SETTINGS', 'days_for_recent', fallback=30),
            'max_retries': self.config.getint('SETTINGS', 'max_retries', fallback=3),
            'retry_delay': self.config.getfloat('SETTINGS', 'retry_delay', fallback=1.0),
            'max_pages': self.config.getint('SETTINGS', 'max_pages', fallback=100),
            'enable_resume': self.config.getboolean('SETTINGS', 'enable_resume', fallback=True)
        }

        # 读取天机平台账号（过滤注释行）
        tianji_accounts = {}
        if self.config.has_section('TIANJI_ACCOUNTS'):
            for key, value in self.config.items('TIANJI_ACCOUNTS'):
                # 过滤注释行（以';'开头的键）
                if not key.strip().startswith(';'):
                    tianji_accounts[key] = value

        # 读取小台风平台账号（过滤注释行）
        xiaotaifeng_accounts = []
        if self.config.has_section('XIAOTAIENG_ACCOUNTS'):
            for key, value in self.config.items('XIAOTAIENG_ACCOUNTS'):
                # 过滤注释行
                if not key.strip().startswith(';'):
                    xiaotaifeng_accounts.append({"username": key, "password": value})

        # 读取妙月平台账号（过滤注释行）
        miaoyue_accounts = {}
        if self.config.has_section('MIAOYUE_ACCOUNTS'):
            for key, value in self.config.items('MIAOYUE_ACCOUNTS'):
                # 过滤注释行
                if not key.strip().startswith(';'):
                    miaoyue_accounts[key] = value

        return {
            'common': common_config,
            'tianji': {
                'base_url': 'https://sys.szlaina.com',
                'accounts': tianji_accounts,
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            },
            'xiaotaifeng': {
                'base_url': 'http://123.56.58.202:8085',
                'accounts': xiaotaifeng_accounts,
                'login_url': '/user/login',
                'balance_url': '/profit/profitcanwithdraw',
                'bill_list_url': '/profit/list'
            },
            'miaoyue': {
                'base_url': 'https://sapi.musmoon.com',
                'accounts': miaoyue_accounts,
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
        }

    def create_default_config(self):
        """创建默认配置文件（简化注释格式）"""
        self.config['SETTINGS'] = {
            'output_path': '桌面',
            'bill_page_size': '50',
            'query_all_bills': 'False',
            'max_workers': '3',
            'enable_threading': 'True',
            'request_timeout': '30',
            'platform_delay': '2.0',
            'days_for_recent': '30',
            'max_retries': '3',
            'retry_delay': '1.0',
            'max_pages': '100',
            'enable_resume': 'True'
        }

        # 直接在配置项前添加注释
        self.config['TIANJI_ACCOUNTS'] = {
            'Wdy': '90535de091e878a11a3e1724ab22bc10',
            'CFWS': 'a71a5ba407b3e4333d1a89689779446b',
            '晨阳科技': 'a71a5ba407b3e4333d1a89689779446b'
        }

        self.config['XIAOTAIENG_ACCOUNTS'] = {
            '超凡威视': '525231314.',
            '塘厦益雅贸易': '112233',
            '小姜安防': 'Wu5626480',
            '南鲁集镇': '525231314.'
        }

        self.config['MIAOYUE_ACCOUNTS'] = {
            'jiweishidai': '6D218509562ED94DB2808E28AE3DB3BB',
            'huangfangyi': '6F0A6BC78A79D8E922410BB0971FDE0A',
            '蓝硕商贸科技': '6F0A6BC78A79D8E922410BB0971FDE0A'
        }

        with open(self.config_path, 'w', encoding='utf-8') as f:
            self.config.write(f)

        # 打印使用说明
        print(f"✅ 已创建默认配置文件：{self.config_path}")
        print(f"\n📋 配置文件说明：")
        print(f"1. 天机平台账号格式：账号名 = 加密密码")
        print(f"2. 小台风平台账号格式：账号名 = 密码")
        print(f"3. 妙月平台账号格式：账号名 = 加密密码")
        print(f"\n⚠️  请修改配置文件中的API地址和账号信息")

# ======================== 数据清洗和标准化 ========================
class DataProcessorV2:
    """V2.0数据处理和标准化类（严格按照规范文档）"""

    @staticmethod
    def safe_str(value: Any, default: str = "") -> str:
        """安全转换为字符串，空值返回空字符串"""
        if pd.isna(value) or value is None:
            return default

        str_value = str(value).strip()
        if str_value in ['', 'null', 'NULL', 'Null', 'N/A', 'n/a', 'NaN', 'nan', 'None', 'none', '未采集', '未知', '-',
                         '--']:
            return default

        return str_value

    @staticmethod
    def safe_float(value: Any, default: str = "") -> Union[float, str]:
        """安全转换浮点数：非数字/空值返回空字符串"""
        if pd.isna(value) or value is None:
            return default

        str_value = str(value).strip()
        if str_value in ['', 'null', 'NULL', 'Null', 'N/A', 'n/a', 'NaN', 'nan', 'None', 'none', '未采集', '未知', '-',
                         '--']:
            return default

        try:
            # 处理可能包含逗号的千分位数字
            str_value = str_value.replace(',', '')
            num = float(str_value)
            # 四舍五入保留2位小数
            return round(num, 2)
        except (ValueError, TypeError):
            return default

    @staticmethod
    def standardize_datetime(dt_str: Any) -> str:
        """标准化时间格式：YYYY-MM-DD HH:MM:SS，空值留空"""
        if pd.isna(dt_str) or dt_str is None:
            return ""

        dt_str = str(dt_str).strip()
        if dt_str in ['', 'null', 'NULL', 'Null', 'N/A', 'n/a', 'NaN', 'nan', 'None', 'none', '未采集', '未知', '-',
                      '--']:
            return ""

        # 尝试多种时间格式
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y/%m/%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
            '%Y%m%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y/%m/%d %H:%M',
            '%Y年%m月%d日 %H:%M:%S',
            '%Y-%m-%d',
            '%Y/%m/%d'
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(dt_str, fmt)
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                continue

        # 如果是时间戳（秒或毫秒）
        if dt_str.isdigit():
            if len(dt_str) == 10:  # 秒级时间戳
                try:
                    dt = datetime.fromtimestamp(int(dt_str))
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            elif len(dt_str) == 13:  # 毫秒级时间戳
                try:
                    dt = datetime.fromtimestamp(int(dt_str) / 1000)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass

        # 无法转换则返回空字符串
        return ""

    @staticmethod
    def standardize_operator(operator_str: Any, platform: str = "") -> str:
        """运营商标准化（空值留空）"""
        if pd.isna(operator_str) or operator_str is None:
            return ""

        operator_str = str(operator_str).strip()
        if operator_str == "":
            return ""

        # 统一大写处理
        operator_upper = operator_str.upper()

        # 判断运营商类型
        if 'CM' in operator_upper or '移动' in operator_str:
            return "中国移动"
        elif 'CT' in operator_upper or '电信' in operator_str or 'TELECOM' in operator_upper:
            return "中国电信"
        elif 'CU' in operator_upper or '联通' in operator_str or 'UNICOM' in operator_upper:
            return "中国联通"
        else:
            # 非标准运营商名称，返回空字符串
            return ""

    @staticmethod
    def classify_income_type(platform: str, **kwargs) -> str:
        """收入类型归类（空值留空）"""
        if platform == "tianji":
            remarks = kwargs.get('remarks', '')
            remarks_str = DataProcessorV2.safe_str(remarks)
            if remarks_str == "":
                return ""

            if "续费" in remarks_str:
                return "续费"
            elif "套餐" in remarks_str or "充值" in remarks_str:
                return "出售套餐"
            else:
                return "未分类"

        elif platform == "xiaotaifeng":
            incometype = kwargs.get('incometype', '')
            incometype_str = DataProcessorV2.safe_str(incometype)
            if incometype_str == "":
                return ""

            if "出售套餐" in incometype_str:
                return "出售套餐"
            elif "续费" in incometype_str:
                return "续费"
            else:
                return "未分类"

        elif platform == "miaoyue":
            order_content = kwargs.get('order_content', '')
            order_content_str = DataProcessorV2.safe_str(order_content)
            if order_content_str == "":
                return ""

            if "续费" in order_content_str:
                return "续费"
            elif "月包" in order_content_str or "半年包" in order_content_str or "年包" in order_content_str:
                return "出售套餐"
            else:
                return "未分类"

        return ""

    @staticmethod
    def extract_card_number(card_number: Any) -> str:
        """提取卡号，只保留数字"""
        if pd.isna(card_number) or card_number is None:
            return ""

        card_str = str(card_number).strip()
        if card_str == "":
            return ""

        # 只保留数字
        digits = re.findall(r'\d+', card_str)
        if digits:
            return ''.join(digits)

        return ""

    @staticmethod
    def clean_remarks(remarks: Any) -> str:
        """清理备注字段"""
        if pd.isna(remarks) or remarks is None:
            return ""

        remarks_str = str(remarks).strip()
        if remarks_str == "":
            return ""

        # 去除特殊字符
        cleaned = re.sub(r'[/\\*#@$%^&|]', '', remarks_str)
        return cleaned.strip()


# ======================== 天机平台客户端 V2 ========================
class TianjiClientV2:
    """天机平台客户端 V2.0（严格按照规范文档）"""

    def __init__(self, config: dict):
        self.cfg = config
        self.common_cfg = config['common']
        self.base_url = config['tianji']['base_url']
        self.data_processor = DataProcessorV2()
        self.account_info = {}
        self.query_cache = {}

        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "keep-alive",
            "Referer": f"{self.base_url}/Index/index",
        }

    def login(self, username: str, password: str) -> tuple[Session, bool, str]:
        """登录天机平台"""
        session = Session()

        try:
            # 先访问首页获取初始cookie
            session.get(f"{self.base_url}/Index/index", headers=self.headers,
                        verify=False, timeout=self.common_cfg['request_timeout'])

            login_url = f"{self.base_url}/Login/doLogin"
            login_data = {"u_name": username, "pwd": password, "encry": "1"}

            resp = session.post(login_url, data=login_data, headers=self.headers,
                                verify=False, timeout=self.common_cfg['request_timeout'])
            resp.encoding = "utf-8"

            if session.cookies.get("PHPSESSID"):
                logger.info(f"天机平台账号 {username} 登录成功")
                return session, True, ""
            else:
                error_msg = "登录失败，无PHPSESSID"
                logger.error(f"天机平台账号 {username} {error_msg}")
                return session, False, error_msg

        except Exception as e:
            error_msg = f"登录异常: {str(e)}"
            logger.error(f"天机平台账号 {username} {error_msg}")
            return session, False, error_msg

    def get_balance(self, session: Session, username: str) -> tuple[float, str]:
        """获取余额"""
        try:
            profit_url = f"{self.base_url}/Profit/companyProfit"

            # 先访问一次利润页面
            session.get(f"{self.base_url}/Profit/listProfit", headers=self.headers,
                        verify=False, timeout=self.common_cfg['request_timeout'])

            resp = session.get(profit_url, headers=self.headers, verify=False,
                               timeout=self.common_cfg['request_timeout'])
            resp.encoding = "utf-8"

            if resp.status_code != 200:
                error_msg = f"余额接口请求失败，状态码：{resp.status_code}"
                logger.warning(f"天机平台账号 {username} {error_msg}")
                return 0.0, error_msg

            html = resp.text

            # 尝试多种方式提取余额
            patterns = [
                r'余额[:：]\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'(?:可用)?余额\s*[:：]?\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'当前余额\s*[:：]?\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'账户余额\s*[:：]?\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'余额\s*[：:]\s*<[^>]+>([-+]?\d+(?:\.\d+)?)</[^>]+>',
                r'余额\s*</span>\s*<span[^>]*>\s*([-+]?\d+(?:\.\d+)?)',
                r'<span[^>]*>\s*余额\s*</span>\s*<span[^>]*>\s*([-+]?\d+(?:\.\d+)?)',
                r'<td[^>]*>\s*余额\s*</td>\s*<td[^>]*>\s*([-+]?\d+(?:\.\d+)?)',
                r'余额\s*<em[^>]*>\s*([-+]?\d+(?:\.\d+)?)\s*</em>',
            ]

            for pattern in patterns:
                matches = re.findall(pattern, html, re.IGNORECASE)
                for match in matches:
                    try:
                        balance = float(match)
                        logger.info(f"天机平台账号 {username} 余额: {balance:.2f} 元")

                        self.account_info[username] = {
                            'balance': balance,
                            'recent_income': 0.0,
                            'recent_withdraw': 0.0,
                            'recent_refund': 0.0,
                            'total_bills': 0,
                            'last_query_page': 0
                        }
                        return balance, ""
                    except ValueError:
                        continue

            # 使用BeautifulSoup进行更精确的提取
            try:
                soup = BeautifulSoup(html, 'html.parser')

                # 查找包含"余额"的所有元素
                for text in soup.find_all(text=re.compile(r'余额')):
                    # 获取父元素
                    parent = text.parent
                    parent_text = parent.get_text()

                    # 在父文本中查找数字
                    matches = re.findall(r'[-+]?\d+(?:\.\d+)?', parent_text)
                    for match in matches:
                        try:
                            num = float(match)
                            # 检查数字是否在余额关键词附近
                            if re.search(r'余额[:：]\s*' + match, parent_text) or re.search(match + r'\s*元',
                                                                                           parent_text):
                                logger.info(f"天机平台账号 {username} BeautifulSoup提取余额: {num:.2f} 元")

                                self.account_info[username] = {
                                    'balance': num,
                                    'recent_income': 0.0,
                                    'recent_withdraw': 0.0,
                                    'recent_refund': 0.0,
                                    'total_bills': 0,
                                    'last_query_page': 0
                                }
                                return num, ""
                        except ValueError:
                            continue
            except Exception as e:
                logger.warning(f"天机平台账号 {username} BeautifulSoup解析异常: {str(e)}")

            # 最后尝试直接在整个HTML中查找数字模式
            all_numbers = re.findall(r'余额[:：]\s*[¥￥]?\s*(\d+(?:\.\d+)?)', html, re.IGNORECASE)
            for num_str in all_numbers:
                try:
                    balance = float(num_str)
                    logger.info(f"天机平台账号 {username} 直接提取余额: {balance:.2f} 元")

                    self.account_info[username] = {
                        'balance': balance,
                        'recent_income': 0.0,
                        'recent_withdraw': 0.0,
                        'recent_refund': 0.0,
                        'total_bills': 0,
                        'last_query_page': 0
                    }
                    return balance, ""
                except ValueError:
                    continue

            error_msg = "未提取到余额数据"
            logger.warning(f"天机平台账号 {username} {error_msg}")
            return 0.0, error_msg

        except Exception as e:
            error_msg = f"余额查询异常: {str(e)}"
            logger.error(f"天机平台账号 {username} {error_msg}")
            return 0.0, error_msg

    def get_bills(self, session: Session, username: str) -> tuple[List[Dict], str]:
        """获取账单（严格按照V2.0规范）"""
        all_bills = []
        errors = []

        # 获取上次查询的页码
        last_page = self.account_info.get(username, {}).get('last_query_page', 0)
        start_page = last_page + 1 if self.common_cfg['enable_resume'] else 1

        page = start_page
        page_size = self.common_cfg['bill_page_size']
        max_pages = self.common_cfg.get('max_pages', 100)

        logger.info(f"天机平台账号 {username} 开始查询账单，从第{page}页开始")

        while True:
            if page > max_pages:
                logger.info(f"天机平台账号 {username} 已达到最大查询页数 {max_pages}")
                break

            cache_key = f"{username}_page_{page}"
            if cache_key in self.query_cache:
                logger.debug(f"天机平台账号 {username} 第{page}页已缓存，跳过")
                page += 1
                continue

            bills, error = self._get_single_page_bills(session, username, page, page_size)

            if error:
                errors.append(error)
                if len(errors) > 3:
                    logger.warning(f"天机平台账号 {username} 连续多页查询失败，停止查询")
                    break
                time.sleep(2)
                continue

            if not bills:
                logger.info(f"天机平台账号 {username} 第{page}页无数据，查询完成")
                break

            all_bills.extend(bills)
            self.query_cache[cache_key] = True

            # 更新最后查询页码
            if username in self.account_info:
                self.account_info[username]['last_query_page'] = page
                self.account_info[username]['total_bills'] += len(bills)

            logger.info(f"天机平台账号 {username} 第{page}页获取到 {len(bills)} 条账单，累计 {len(all_bills)} 条")

            if not self.common_cfg['query_all_bills']:
                break

            page += 1

            # 添加延迟，避免请求过于频繁
            delay = self.common_cfg.get('page_delay', 0.5)
            if delay > 0:
                time.sleep(delay)

        if all_bills:
            logger.info(f"天机平台账号 {username} 共获取到 {len(all_bills)} 条账单")
        else:
            logger.warning(f"天机平台账号 {username} 未获取到账单数据")

        error_info = "; ".join(errors) if errors else ""
        return all_bills, error_info

    def _get_single_page_bills(self, session: Session, username: str, page: int, page_size: int) -> tuple[
        List[Dict], str]:
        """获取单页账单"""
        try:
            bill_url = f"{self.base_url}/Profit/billDetail"
            bill_data = {
                "page": page,
                "limit": page_size,
                "start_time": "",
                "end_time": "",
                "type": ""
            }

            bill_headers = self.headers.copy()
            bill_headers.update({
                "Accept": "*/*",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": f"{self.base_url}/Profit/listBillDetail"
            })

            resp = session.post(bill_url, data=bill_data, headers=bill_headers,
                                verify=False, timeout=self.common_cfg['request_timeout'])
            resp.encoding = "utf-8"

            if resp.status_code == 200:
                try:
                    bill_json = resp.json()

                    if isinstance(bill_json, dict):
                        if bill_json.get("status") in [1, "1", 200] or bill_json.get("message") in ["成功", "success"]:
                            bill_list = bill_json.get("list", [])
                            if bill_list:
                                processed_bills = self._process_bills(bill_list, username)
                                return processed_bills, ""
                            else:
                                return [], f"第{page}页无账单数据"
                        else:
                            error_msg = bill_json.get("message", "未知错误")
                            return [], f"第{page}页接口异常: {error_msg}"
                    else:
                        return [], f"第{page}页响应格式异常"
                except json.JSONDecodeError:
                    return [], f"第{page}页JSON解析失败"
            else:
                return [], f"第{page}页请求失败，状态码：{resp.status_code}"

        except Exception as e:
            return [], f"第{page}页查询异常: {str(e)}"

    def _process_bills(self, raw_bills: List[Dict], username: str) -> List[Dict]:
        """处理天机平台账单数据（严格按照V2.0规范）"""
        processed = []
        recent_income = 0.0
        recent_withdraw = 0.0
        recent_refund = 0.0
        days_for_recent = self.common_cfg.get('days_for_recent', 30)
        cutoff_date = datetime.now() - timedelta(days=days_for_recent)

        for bill in raw_bills:
            # 提取基础字段
            order_no = self.data_processor.safe_str(bill.get('order_no', ''))
            iccid = self.data_processor.safe_str(bill.get('iccid', ''))
            trans_time = self.data_processor.standardize_datetime(bill.get('trans_time_format', ''))

            # 交易金额字段
            sale_price = self.data_processor.safe_float(bill.get('income_money', ''))
            cost_price = self.data_processor.safe_float(bill.get('cost_money', ''))
            commission = self.data_processor.safe_float(bill.get('profit', ''))

            # 业务属性字段
            customer_name = self.data_processor.safe_str(bill.get('company_name', ''))
            product_name = self.data_processor.safe_str(bill.get('order_name', ''))
            operator = self.data_processor.standardize_operator(bill.get('second_operator_name', ''), 'tianji')

            # 收入类型归类
            income_type = self.data_processor.classify_income_type(
                'tianji',
                remarks=bill.get('remarks', '')
            )

            # 备注字段
            remark = self.data_processor.clean_remarks(bill.get('remarks', ''))

            # 卡号字段（天机平台无此字段）
            card_number = ""

            # 统计最近收益
            if trans_time:
                try:
                    trans_dt = datetime.strptime(trans_time, '%Y-%m-%d %H:%M:%S')
                    if trans_dt >= cutoff_date:
                        if isinstance(commission, float) and commission > 0:
                            recent_income += commission
                        elif isinstance(commission, float) and commission < 0:
                            # 判断是提现还是退款
                            if income_type == "提现支出":
                                recent_withdraw += abs(commission)
                            elif income_type == "退款":
                                recent_refund += abs(commission)
                except:
                    pass

            # 构建标准化账单记录（严格按照V2.0字段编码）
            processed_bill = {
                'order_no': order_no,
                'iccid': iccid.upper() if iccid != "" else "",  # ICCID统一大写
                'card_number': card_number,
                'trans_time': trans_time,
                'sale_price': sale_price if isinstance(sale_price, (int, float)) else "",
                'cost_price': cost_price if isinstance(cost_price, (int, float)) else "",
                'commission': commission if isinstance(commission, (int, float)) else "",
                'customer_name': customer_name,
                'product_name': product_name,
                'operator': operator,
                'income_type': income_type,
                'remark': remark,
                'platform': '天机',
                'account': username
            }
            processed.append(processed_bill)

        # 更新账号信息
        if username in self.account_info:
            self.account_info[username]['recent_income'] += recent_income
            self.account_info[username]['recent_withdraw'] += recent_withdraw
            self.account_info[username]['recent_refund'] += recent_refund

        return processed

    def get_account_info(self, username: str) -> Dict:
        """获取账号信息"""
        return self.account_info.get(username, {
            'balance': 0.0,
            'recent_income': 0.0,
            'recent_withdraw': 0.0,
            'recent_refund': 0.0,
            'total_bills': 0,
            'last_query_page': 0
        })


# ======================== 小台风平台客户端 V2 ========================
class XiaoTaiFengClientV2:
    """小台风平台客户端 V2.0（严格按照规范文档）"""

    def __init__(self, config: dict):
        self.cfg = config
        self.common_cfg = config['common']
        self.base_url = config['xiaotaifeng']['base_url']
        self.login_url = self.base_url + config['xiaotaifeng']['login_url']
        self.balance_url = self.base_url + config['xiaotaifeng']['balance_url']
        self.bill_list_url = self.base_url + config['xiaotaifeng']['bill_list_url']
        self.data_processor = DataProcessorV2()
        self.account_info = {}
        self.query_cache = {}

        self.request_headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "keep-alive",
            "Content-Type": "application/json",
            "Host": "123.56.58.202:8085",
            "Origin": "http://iot.xiaotaifeng.cn",
            "Referer": "http://iot.xiaotaifeng.cn/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36"
        }

    def login(self, username: str, password: str) -> tuple[Session, bool, str]:
        """登录小台风平台"""
        session = Session()
        session.headers.update(self.request_headers)

        try:
            login_data = {"username": username, "password": password}
            resp = session.post(
                self.login_url,
                json=login_data,
                timeout=self.common_cfg['request_timeout']
            )

            if resp.status_code == 200:
                result = resp.json()
                if result.get("code") == "0" and result.get("message") == "登录成功":
                    token = result.get("data", {}).get("token")
                    if token:
                        session.headers["X-Token"] = token
                        logger.info(f"小台风平台账号 {username} 登录成功")
                        return session, True, ""
                    else:
                        error_msg = "登录成功但未获取到token"
                        logger.error(f"小台风平台账号 {username} {error_msg}")
                        return session, False, error_msg
                else:
                    error_msg = f"登录失败: {result.get('message', '未知错误')}"
                    logger.error(f"小台风平台账号 {username} {error_msg}")
                    return session, False, error_msg
            else:
                error_msg = f"登录请求失败，状态码：{resp.status_code}"
                logger.error(f"小台风平台账号 {username} {error_msg}")
                return session, False, error_msg

        except Exception as e:
            error_msg = f"登录异常: {str(e)}"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return session, False, error_msg

    def get_balance(self, session: Session, username: str) -> tuple[float, str]:
        """获取余额"""
        try:
            resp = session.get(
                self.balance_url,
                timeout=self.common_cfg['request_timeout']
            )

            if resp.status_code == 200:
                result = resp.json()
                if result.get("code") == "0" and isinstance(result.get("data"), (int, float)):
                    balance = float(result["data"])
                    logger.info(f"小台风平台账号 {username} 余额: {balance:.2f} 元")

                    self.account_info[username] = {
                        'balance': balance,
                        'recent_income': 0.0,
                        'recent_withdraw': 0.0,
                        'recent_refund': 0.0,
                        'total_bills': 0,
                        'last_query_page': 0
                    }
                    return balance, ""
                else:
                    return 0.0, f"余额格式异常: {result}"
            else:
                return 0.0, f"余额请求失败，状态码：{resp.status_code}"

        except Exception as e:
            error_msg = f"余额查询异常: {str(e)}"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return 0.0, error_msg

    def get_bills(self, session: Session, username: str) -> tuple[List[Dict], str]:
        """获取账单（严格按照V2.0规范）"""
        all_bills = []
        errors = []

        last_page = self.account_info.get(username, {}).get('last_query_page', 0)
        start_page = last_page + 1 if self.common_cfg['enable_resume'] else 1

        page = start_page
        page_size = self.common_cfg['bill_page_size']
        max_pages = self.common_cfg.get('max_pages', 100)

        logger.info(f"小台风平台账号 {username} 开始查询账单，从第{page}页开始")

        while True:
            if page > max_pages:
                logger.info(f"小台风平台账号 {username} 已达到最大查询页数 {max_pages}")
                break

            cache_key = f"{username}_page_{page}"
            if cache_key in self.query_cache:
                logger.debug(f"小台风平台账号 {username} 第{page}页已缓存，跳过")
                page += 1
                continue

            bills, error = self._get_single_page_bills(session, username, page, page_size)

            if error:
                errors.append(error)
                if len(errors) > 3:
                    logger.warning(f"小台风平台账号 {username} 连续多页查询失败，停止查询")
                    break
                time.sleep(2)
                continue

            if not bills:
                logger.info(f"小台风平台账号 {username} 第{page}页无数据，查询完成")
                break

            all_bills.extend(bills)
            self.query_cache[cache_key] = True

            if username in self.account_info:
                self.account_info[username]['last_query_page'] = page
                self.account_info[username]['total_bills'] += len(bills)

            logger.info(f"小台风平台账号 {username} 第{page}页获取到 {len(bills)} 条账单，累计 {len(all_bills)} 条")

            if not self.common_cfg['query_all_bills']:
                break

            page += 1

            delay = self.common_cfg.get('page_delay', 0.5)
            if delay > 0:
                time.sleep(delay)

        if all_bills:
            logger.info(f"小台风平台账号 {username} 共获取到 {len(all_bills)} 条账单")
        else:
            logger.warning(f"小台风平台账号 {username} 未获取到账单数据")

        error_info = "; ".join(errors) if errors else ""
        return all_bills, error_info

    def _get_single_page_bills(self, session: Session, username: str, page: int, page_size: int) -> tuple[
        List[Dict], str]:
        """获取单页账单"""
        try:
            params = {
                "paytype": "",
                "account": "",
                "productid": "",
                "name": "",
                "page": page,
                "limit": page_size,
                "sort": "-d.ID"
            }

            resp = session.get(
                self.bill_list_url,
                params=params,
                timeout=self.common_cfg['request_timeout']
            )

            if resp.status_code == 200:
                result = resp.json()
                if result.get("code") == "0" and "data" in result and "items" in result["data"]:
                    bill_list = result["data"]["items"]
                    if bill_list:
                        processed_bills = self._process_bills(bill_list, username)
                        return processed_bills, ""
                    else:
                        return [], f"第{page}页无账单数据"
                else:
                    return [], f"第{page}页格式异常: {result.get('message', '未知错误')}"
            else:
                return [], f"第{page}页请求失败，状态码：{resp.status_code}"

        except Exception as e:
            return [], f"第{page}页查询异常: {str(e)}"

    def _process_bills(self, raw_bills: List[Dict], username: str) -> List[Dict]:
        """处理小台风平台账单数据（严格按照V2.0规范）"""
        processed = []
        recent_income = 0.0
        recent_withdraw = 0.0
        recent_refund = 0.0
        days_for_recent = self.common_cfg.get('days_for_recent', 30)
        cutoff_date = datetime.now() - timedelta(days=days_for_recent)

        for bill in raw_bills:
            # 提取基础字段
            order_no = self.data_processor.safe_str(bill.get('orderid', ''))
            iccid = self.data_processor.safe_str(bill.get('iccid', ''))
            card_number = self.data_processor.extract_card_number(bill.get('msisdn', ''))
            trans_time = self.data_processor.standardize_datetime(bill.get('purchasetime', ''))

            # 交易金额字段（可计算字段：成本 = 售价 - 佣金）
            sale_price = self.data_processor.safe_float(bill.get('amount', ''))
            commission = self.data_processor.safe_float(bill.get('profit', ''))

            # 计算成本（可计算字段）
            cost_price = ""
            if isinstance(sale_price, (int, float)) and isinstance(commission, (int, float)):
                cost_price = round(sale_price - commission, 2)

            # 客户名称：优先custom，其次account
            custom = self.data_processor.safe_str(bill.get('custom', ''))
            account = self.data_processor.safe_str(bill.get('account', ''))
            customer_name = custom if custom != "" else account

            # 业务属性字段
            product_name = self.data_processor.safe_str(bill.get('mpname', ''))
            operator = self.data_processor.standardize_operator(bill.get('yunyingshang', ''), 'xiaotaifeng')

            # 收入类型归类
            income_type = self.data_processor.classify_income_type(
                'xiaotaifeng',
                incometype=bill.get('incometype', '')
            )

            # 备注字段
            remark = self.data_processor.clean_remarks(bill.get('remark', ''))

            # 统计最近收益
            if trans_time:
                try:
                    trans_dt = datetime.strptime(trans_time, '%Y-%m-%d %H:%M:%S')
                    if trans_dt >= cutoff_date:
                        if isinstance(commission, float) and commission > 0:
                            recent_income += commission
                        elif isinstance(commission, float) and commission < 0:
                            if income_type == "提现支出":
                                recent_withdraw += abs(commission)
                            elif income_type == "退款":
                                recent_refund += abs(commission)
                except:
                    pass

            # 构建标准化账单记录（严格按照V2.0字段编码）
            processed_bill = {
                'order_no': order_no,
                'iccid': iccid.upper() if iccid != "" else "",  # ICCID统一大写
                'card_number': card_number,
                'trans_time': trans_time,
                'sale_price': sale_price if isinstance(sale_price, (int, float)) else "",
                'cost_price': cost_price if isinstance(cost_price, (int, float)) else "",
                'commission': commission if isinstance(commission, (int, float)) else "",
                'customer_name': customer_name,
                'product_name': product_name,
                'operator': operator,
                'income_type': income_type,
                'remark': remark,
                'platform': '小台风',
                'account': username
            }
            processed.append(processed_bill)

        # 更新账号信息
        if username in self.account_info:
            self.account_info[username]['recent_income'] += recent_income
            self.account_info[username]['recent_withdraw'] += recent_withdraw
            self.account_info[username]['recent_refund'] += recent_refund

        return processed

    def get_account_info(self, username: str) -> Dict:
        """获取账号信息"""
        return self.account_info.get(username, {
            'balance': 0.0,
            'recent_income': 0.0,
            'recent_withdraw': 0.0,
            'recent_refund': 0.0,
            'total_bills': 0,
            'last_query_page': 0
        })


# ======================== 妙月平台客户端 V2 ========================
class MiaoYueClientV2:
    """妙月平台客户端 V2.0（严格按照规范文档）"""

    def __init__(self, config: dict):
        self.cfg = config
        self.common_cfg = config['common']
        self.base_url = config['miaoyue']['base_url']
        self.ua = UserAgent()
        self.data_processor = DataProcessorV2()
        self.account_info = {}
        self.query_cache = {}

    def login(self, username: str, password: str) -> tuple[Optional[str], str]:
        """登录妙月平台"""
        try:
            login_url = f"{self.base_url}/card/user/password/login"
            login_params = {"username": username, "password": password}

            headers = {"User-Agent": self.ua.random}
            response = requests.post(login_url, params=login_params, headers=headers,
                                     timeout=self.common_cfg['request_timeout'])
            response.raise_for_status()

            result = response.json()
            if result.get("success") and result.get("statusCode") == 0:
                token = result["object"]["token"]
                logger.info(f"妙月平台账号 {username} 登录成功")
                return token, ""
            else:
                error_msg = f"登录失败: {result.get('content', '未知错误')}"
                logger.error(f"妙月平台账号 {username} {error_msg}")
                return None, error_msg

        except Exception as e:
            error_msg = f"登录异常: {str(e)}"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return None, error_msg

    def get_balance(self, token: str, username: str) -> tuple[float, float, float, str]:
        """获取余额"""
        try:
            balance_url = f"{self.base_url}/card/proxy/company/capital/account/info?currencyType=CNY"
            headers = {"x-token": f'{{"token":"{token}"}}', "User-Agent": self.ua.random}

            response = requests.get(balance_url, headers=headers,
                                    timeout=self.common_cfg['request_timeout'])
            response.raise_for_status()

            balance_raw = response.json()
            if balance_raw.get("success") and balance_raw.get("statusCode") == 0:
                balance_info = balance_raw.get("object", {})
                withdrawable = self.data_processor.safe_float(balance_info.get("withdrawAmount", 0))
                non_withdrawable = self.data_processor.safe_float(balance_info.get("nonWithdrawAmount", 0))
                total = withdrawable + non_withdrawable

                logger.info(f"妙月平台账号 {username} 余额: {total:.2f} 元 (可提现: {withdrawable:.2f} 元)")

                self.account_info[username] = {
                    'balance': total,
                    'withdrawable': withdrawable,
                    'non_withdrawable': non_withdrawable,
                    'recent_income': 0.0,
                    'recent_withdraw': 0.0,
                    'recent_refund': 0.0,
                    'total_bills': 0,
                    'last_query_page': 0
                }

                return total, withdrawable, non_withdrawable, ""
            else:
                return 0.0, 0.0, 0.0, f"余额查询失败: {balance_raw.get('content', '未知错误')}"

        except Exception as e:
            error_msg = f"余额查询异常: {str(e)}"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return 0.0, 0.0, 0.0, error_msg

    def get_bills(self, token: str, username: str) -> tuple[List[Dict], str]:
        """获取账单（严格按照V2.0规范）"""
        all_bills = []
        errors = []

        last_page = self.account_info.get(username, {}).get('last_query_page', 0)
        start_page = last_page + 1 if self.common_cfg['enable_resume'] else 1

        current = start_page
        max_pages = self.common_cfg.get('max_pages', 100)

        logger.info(f"妙月平台账号 {username} 开始查询账单，从第{current}页开始")

        while True:
            if current > max_pages:
                logger.info(f"妙月平台账号 {username} 已达到最大查询页数 {max_pages}")
                break

            cache_key = f"{username}_page_{current}"
            if cache_key in self.query_cache:
                logger.debug(f"妙月平台账号 {username} 第{current}页已缓存，跳过")
                current += 1
                continue

            bills, error = self._get_single_page_bills(token, username, current)

            if error:
                errors.append(error)
                if len(errors) > 3:
                    logger.warning(f"妙月平台账号 {username} 连续多页查询失败，停止查询")
                    break
                time.sleep(2)
                continue

            if not bills:
                logger.info(f"妙月平台账号 {username} 第{current}页无数据，查询完成")
                break

            all_bills.extend(bills)
            self.query_cache[cache_key] = True

            if username in self.account_info:
                self.account_info[username]['last_query_page'] = current
                self.account_info[username]['total_bills'] += len(bills)

            logger.info(f"妙月平台账号 {username} 第{current}页获取到 {len(bills)} 条账单，累计 {len(all_bills)} 条")

            if not self.common_cfg['query_all_bills']:
                break

            current += 1

            delay = self.common_cfg.get('page_delay', 0.5)
            if delay > 0:
                time.sleep(delay)

        if all_bills:
            logger.info(f"妙月平台账号 {username} 共获取到 {len(all_bills)} 条账单")
        else:
            logger.warning(f"妙月平台账号 {username} 未获取到账单数据")

        error_info = "; ".join(errors) if errors else ""
        return all_bills, error_info

    def _get_single_page_bills(self, token: str, username: str, current: int) -> tuple[List[Dict], str]:
        """获取单页账单"""
        try:
            bill_url = (f"{self.base_url}/card/proxy/user/bill/page?"
                        f"currency=CNY&billType=&orderNo=&cardValue=&"
                        f"orders[0].column=createTime&orders[0].asc=false&"
                        f"current={current}&size={self.common_cfg['bill_page_size']}")

            headers = {"x-token": f'{{"token":"{token}"}}', "User-Agent": self.ua.random}
            response = requests.get(bill_url, headers=headers,
                                    timeout=self.common_cfg['request_timeout'])
            response.raise_for_status()

            bill_raw = response.json()

            if bill_raw.get("success") and bill_raw.get("statusCode") == 0:
                records = bill_raw.get("object", {}).get("records", [])
                if records:
                    processed_bills = self._process_bills(records, username)
                    return processed_bills, ""
                else:
                    return [], f"第{current}页无账单数据"
            else:
                return [], f"第{current}页查询失败: {bill_raw.get('content', '未知错误')}"

        except Exception as e:
            return [], f"第{current}页查询异常: {str(e)}"

    def _process_bills(self, raw_bills: List[Dict], username: str) -> List[Dict]:
        """处理妙月平台账单数据（严格按照V2.0规范）"""
        processed = []
        recent_income = 0.0
        recent_withdraw = 0.0
        recent_refund = 0.0
        days_for_recent = self.common_cfg.get('days_for_recent', 30)
        cutoff_date = datetime.now() - timedelta(days=days_for_recent)

        for bill in raw_bills:
            # 提取基础字段
            order_no = self.data_processor.safe_str(bill.get('orderNo', ''))
            iccid = self.data_processor.safe_str(bill.get('cardIccid', ''))
            card_number = self.data_processor.extract_card_number(bill.get('cardNumber', ''))
            trans_time = self.data_processor.standardize_datetime(bill.get('settleTime', ''))

            # 交易金额字段（可计算字段：佣金 = commissionAmount - deCommissionAmount + extraAmount）
            commission_amount = self.data_processor.safe_float(bill.get('commissionAmount', ''))
            de_commission_amount = self.data_processor.safe_float(bill.get('deCommissionAmount', ''))
            extra_amount = self.data_processor.safe_float(bill.get('extraAmount', ''))

            # 计算佣金（可计算字段）
            commission = ""
            if (isinstance(commission_amount, (int, float)) and
                    isinstance(de_commission_amount, (int, float)) and
                    isinstance(extra_amount, (int, float))):
                commission = round(commission_amount - de_commission_amount + extra_amount, 2)

            # 售价和成本（妙月平台无此字段）
            sale_price = ""
            cost_price = ""

            # 客户名称（妙月平台无此字段）
            customer_name = ""

            # 业务属性字段
            product_name = self.data_processor.safe_str(bill.get('orderContent', ''))
            operator = ""  # 妙月平台无此字段

            # 收入类型归类
            income_type = self.data_processor.classify_income_type(
                'miaoyue',
                order_content=bill.get('orderContent', '')
            )

            # 备注字段
            remark = self.data_processor.clean_remarks(bill.get('mark', ''))

            # 统计最近收益
            if trans_time:
                try:
                    trans_dt = datetime.strptime(trans_time, '%Y-%m-%d %H:%M:%S')
                    if trans_dt >= cutoff_date:
                        if isinstance(commission, (int, float)) and commission > 0:
                            recent_income += commission
                        elif isinstance(commission, (int, float)) and commission < 0:
                            if income_type == "提现支出":
                                recent_withdraw += abs(commission)
                            elif income_type == "退款":
                                recent_refund += abs(commission)
                except:
                    pass

            # 构建标准化账单记录（严格按照V2.0字段编码）
            processed_bill = {
                'order_no': order_no,
                'iccid': iccid.upper() if iccid != "" else "",  # ICCID统一大写
                'card_number': card_number,
                'trans_time': trans_time,
                'sale_price': sale_price if isinstance(sale_price, (int, float)) else "",
                'cost_price': cost_price if isinstance(cost_price, (int, float)) else "",
                'commission': commission if isinstance(commission, (int, float)) else "",
                'customer_name': customer_name,
                'product_name': product_name,
                'operator': operator,
                'income_type': income_type,
                'remark': remark,
                'platform': '妙月',
                'account': username
            }
            processed.append(processed_bill)

        # 更新账号信息
        if username in self.account_info:
            self.account_info[username]['recent_income'] += recent_income
            self.account_info[username]['recent_withdraw'] += recent_withdraw
            self.account_info[username]['recent_refund'] += recent_refund

        return processed

    def get_account_info(self, username: str) -> Dict:
        """获取账号信息"""
        return self.account_info.get(username, {
            'balance': 0.0,
            'withdrawable': 0.0,
            'non_withdrawable': 0.0,
            'recent_income': 0.0,
            'recent_withdraw': 0.0,
            'recent_refund': 0.0,
            'total_bills': 0,
            'last_query_page': 0
        })


# ======================== 多平台管理器和Excel导出 V2 ========================
class MultiPlatformManagerV2:
    """多平台管理器 V2.0（严格按照规范文档）"""

    # 字段映射表（字段编码 -> 字段名称）
    FIELD_MAPPING = {
        'order_no': '订单号',
        'iccid': 'ICCID',
        'card_number': '卡号',
        'trans_time': '交易时间',
        'sale_price': '售价（元）',
        'cost_price': '成本（元）',
        'commission': '佣金（元）',
        'customer_name': '客户名称',
        'product_name': '套餐/产品名称',
        'operator': '运营商',
        'income_type': '收入类型',
        'remark': '备注',
        'platform': '平台',
        'account': '账号'
    }

    # 列宽设置（字符数）
    COLUMN_WIDTHS = {
        '订单号': 18,
        'ICCID': 20,
        '卡号': 15,
        '交易时间': 20,
        '售价（元）': 10,
        '成本（元）': 10,
        '佣金（元）': 10,
        '客户名称': 20,
        '套餐/产品名称': 25,
        '运营商': 10,
        '收入类型': 10,
        '备注': 15,
        '平台': 10,
        '账号': 15
    }

    def __init__(self, config_path: str = "multi_platform_config_v2.ini"):
        self.config_manager = ConfigManager(config_path)
        self.config = self.config_manager.load_config()
        self.all_bills = []
        self.platform_bills = {
            '天机': [],
            '小台风': [],
            '妙月': []
        }
        self.account_summary = []
        self.error_logs = []
        self.summary_data = {
            'total_accounts': 0,
            'success_accounts': 0,
            'failed_accounts': 0,
            'total_balance': 0.0,
            'total_recent_income': 0.0,
            'total_recent_withdraw': 0.0,
            'total_recent_refund': 0.0,
            'total_bills': 0
        }

    def get_output_path(self) -> Path:
        """获取输出路径"""
        output_path = self.config['common']['output_path']
        if output_path.lower() == '桌面':
            return Path.home() / 'Desktop'
        else:
            return Path(output_path)

    def log_error(self, platform: str, username: str, error_type: str, error_msg: str):
        """记录错误日志"""
        error_entry = {
            '时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '平台': platform,
            '账号': username,
            '错误类型': error_type,
            '错误信息': error_msg
        }
        self.error_logs.append(error_entry)

        print(f"{Fore.RED}❌ 【{platform}-{username}】{error_type}: {error_msg}")

    def query_tianji_accounts(self) -> Tuple[List[Dict], List[Dict]]:
        """查询天机平台所有账号"""
        print(f"\n{Fore.BLUE}【天机平台】查询开始")

        accounts = self.config['tianji']['accounts']
        all_bills = []
        account_info_list = []

        for username, password in accounts.items():
            try:
                self.summary_data['total_accounts'] += 1

                client = TianjiClientV2(self.config)
                session, login_ok, login_error = client.login(username, password)
                if not login_ok:
                    self.summary_data['failed_accounts'] += 1
                    self.log_error('天机', username, '登录失败', login_error)
                    continue

                balance, balance_error = client.get_balance(session, username)
                if balance_error:
                    self.log_error('天机', username, '余额查询失败', balance_error)
                else:
                    self.summary_data['total_balance'] += balance
                    print(f"{Fore.GREEN}【天机-{username}】余额: {balance:.2f}元")

                bills, bill_error = client.get_bills(session, username)
                if bill_error:
                    self.log_error('天机', username, '账单查询错误', bill_error)
                all_bills.extend(bills)

                account_info = client.get_account_info(username)
                account_info['平台'] = '天机'
                account_info['账号'] = username
                account_info_list.append(account_info)

                self.summary_data['total_recent_income'] += account_info['recent_income']
                self.summary_data['total_recent_withdraw'] += account_info['recent_withdraw']
                self.summary_data['total_recent_refund'] += account_info['recent_refund']

                self.summary_data['success_accounts'] += 1

                time.sleep(self.config['common']['platform_delay'])

            except Exception as e:
                error_msg = f"查询异常: {str(e)}"
                self.log_error('天机', username, '系统异常', error_msg)
                self.summary_data['failed_accounts'] += 1

        print(f"{Fore.BLUE}【天机平台】查询完成，成功{len(account_info_list)}个账号")
        return all_bills, account_info_list

    def query_xiaotaifeng_accounts(self) -> Tuple[List[Dict], List[Dict]]:
        """查询小台风平台所有账号"""
        print(f"\n{Fore.BLUE}【小台风平台】查询开始")

        accounts = self.config['xiaotaifeng']['accounts']
        all_bills = []
        account_info_list = []

        for account in accounts:
            username = account['username']
            password = account['password']

            try:
                self.summary_data['total_accounts'] += 1

                client = XiaoTaiFengClientV2(self.config)
                session, login_ok, login_error = client.login(username, password)
                if not login_ok:
                    self.summary_data['failed_accounts'] += 1
                    self.log_error('小台风', username, '登录失败', login_error)
                    continue

                balance, balance_error = client.get_balance(session, username)
                if balance_error:
                    self.log_error('小台风', username, '余额查询失败', balance_error)
                else:
                    self.summary_data['total_balance'] += balance
                    print(f"{Fore.GREEN}【小台风-{username}】余额: {balance:.2f}元")

                bills, bill_error = client.get_bills(session, username)
                if bill_error:
                    self.log_error('小台风', username, '账单查询错误', bill_error)
                all_bills.extend(bills)

                account_info = client.get_account_info(username)
                account_info['平台'] = '小台风'
                account_info['账号'] = username
                account_info_list.append(account_info)

                self.summary_data['total_recent_income'] += account_info['recent_income']
                self.summary_data['total_recent_withdraw'] += account_info['recent_withdraw']
                self.summary_data['total_recent_refund'] += account_info['recent_refund']

                self.summary_data['success_accounts'] += 1

                time.sleep(self.config['common']['platform_delay'])

            except Exception as e:
                error_msg = f"查询异常: {str(e)}"
                self.log_error('小台风', username, '系统异常', error_msg)
                self.summary_data['failed_accounts'] += 1

        print(f"{Fore.BLUE}【小台风平台】查询完成")
        return all_bills, account_info_list

    def query_miaoyue_accounts(self) -> Tuple[List[Dict], List[Dict]]:
        """查询妙月平台所有账号"""
        print(f"\n{Fore.BLUE}【妙月平台】查询开始")

        accounts = self.config['miaoyue']['accounts']
        all_bills = []
        account_info_list = []

        for username, password in accounts.items():
            try:
                self.summary_data['total_accounts'] += 1

                client = MiaoYueClientV2(self.config)
                token, login_error = client.login(username, password)
                if not token:
                    self.summary_data['failed_accounts'] += 1
                    self.log_error('妙月', username, '登录失败', login_error)
                    continue

                total, withdrawable, non_withdrawable, balance_error = client.get_balance(token, username)
                if balance_error:
                    self.log_error('妙月', username, '余额查询失败', balance_error)
                else:
                    self.summary_data['total_balance'] += total
                    print(f"{Fore.GREEN}【妙月-{username}】余额: {total:.2f}元（可提现 {withdrawable:.2f}元）")

                bills, bill_error = client.get_bills(token, username)
                if bill_error:
                    self.log_error('妙月', username, '账单查询错误', bill_error)
                all_bills.extend(bills)

                account_info = client.get_account_info(username)
                account_info['平台'] = '妙月'
                account_info['账号'] = username
                account_info['可提现余额'] = account_info.get('withdrawable', 0)
                account_info['不可提现余额'] = account_info.get('non_withdrawable', 0)
                account_info_list.append(account_info)

                self.summary_data['total_recent_income'] += account_info['recent_income']
                self.summary_data['total_recent_withdraw'] += account_info['recent_withdraw']
                self.summary_data['total_recent_refund'] += account_info['recent_refund']

                self.summary_data['success_accounts'] += 1

                time.sleep(self.config['common']['platform_delay'])

            except Exception as e:
                error_msg = f"查询异常: {str(e)}"
                self.log_error('妙月', username, '系统异常', error_msg)
                self.summary_data['failed_accounts'] += 1

        print(f"{Fore.BLUE}【妙月平台】查询完成")
        return all_bills, account_info_list

    def query_all_platforms(self):
        """查询所有平台"""
        start_time = time.time()

        print(f"\n{Fore.CYAN}⚡ 开始多平台查询，启用多线程: {self.config['common']['enable_threading']}")

        # 重置汇总数据
        self.summary_data = {
            'total_accounts': 0,
            'success_accounts': 0,
            'failed_accounts': 0,
            'total_balance': 0.0,
            'total_recent_income': 0.0,
            'total_recent_withdraw': 0.0,
            'total_recent_refund': 0.0,
            'total_bills': 0
        }

        if self.config['common']['enable_threading']:
            with concurrent.futures.ThreadPoolExecutor(
                    max_workers=min(3, self.config['common']['max_workers'])
            ) as executor:
                future_tianji = executor.submit(self.query_tianji_accounts)
                future_xiaotaifeng = executor.submit(self.query_xiaotaifeng_accounts)
                future_miaoyue = executor.submit(self.query_miaoyue_accounts)

                tianji_bills, tianji_accounts = future_tianji.result()
                xiaotaifeng_bills, xiaotaifeng_accounts = future_xiaotaifeng.result()
                miaoyue_bills, miaoyue_accounts = future_miaoyue.result()
        else:
            tianji_bills, tianji_accounts = self.query_tianji_accounts()
            xiaotaifeng_bills, xiaotaifeng_accounts = self.query_xiaotaifeng_accounts()
            miaoyue_bills, miaoyue_accounts = self.query_miaoyue_accounts()

        self.all_bills = tianji_bills + xiaotaifeng_bills + miaoyue_bills
        self.platform_bills['天机'] = tianji_bills
        self.platform_bills['小台风'] = xiaotaifeng_bills
        self.platform_bills['妙月'] = miaoyue_bills

        self.account_summary = tianji_accounts + xiaotaifeng_accounts + miaoyue_accounts
        self.summary_data['total_bills'] = len(self.all_bills)

        elapsed_time = time.time() - start_time

        # 计算净收益
        net_income = self.summary_data['total_recent_income'] - self.summary_data['total_recent_refund']

        print(f"\n{Fore.GREEN}{'=' * 60}")
        print(f"{Fore.GREEN}查询完成！用时：{elapsed_time:.1f}秒")
        print(f"{Fore.GREEN}{'=' * 60}")
        print(f"📊 汇总信息：")
        print(f"   总账号数：{self.summary_data['total_accounts']}")
        print(f"   成功账号：{self.summary_data['success_accounts']}")
        print(f"   失败账号：{self.summary_data['failed_accounts']}")
        print(f"   总余额：{self.summary_data['total_balance']:.2f} 元")
        print(
            f"   最近{self.config['common'].get('days_for_recent', 30)}天总收益：{self.summary_data['total_recent_income']:.2f} 元")
        print(
            f"   最近{self.config['common'].get('days_for_recent', 30)}天总提现：{self.summary_data['total_recent_withdraw']:.2f} 元")
        print(
            f"   最近{self.config['common'].get('days_for_recent', 30)}天总退款：{self.summary_data['total_recent_refund']:.2f} 元")
        print(f"   净收益（总收益 - 总退款）：{net_income:.2f} 元")
        print(f"   总账单数：{self.summary_data['total_bills']}")
        print(f"{Fore.GREEN}{'=' * 60}")

        # 显示查询完成率
        if self.summary_data['total_accounts'] > 0:
            success_rate = (self.summary_data['success_accounts'] / self.summary_data['total_accounts']) * 100
            print(f"✅ 查询完成率：{success_rate:.1f}%")

            # 显示各平台账单数
            print(f"\n📊 各平台账单统计：")
            for platform in ['天机', '小台风', '妙月']:
                bills_count = len(self.platform_bills[platform])
                if self.summary_data['total_bills'] > 0:
                    percentage = (bills_count / self.summary_data['total_bills']) * 100
                    print(f"   {platform}: {bills_count} 条 ({percentage:.1f}%)")

    def export_to_excel(self):
        """导出数据到Excel（严格按照V2.0规范）"""
        if not self.all_bills and not self.account_summary:
            print(f"{Fore.YELLOW}⚠️  无任何数据，跳过导出")
            return

        output_path = self.get_output_path()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_path / f"多平台账单汇总_V2_{timestamp}.xlsx"

        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 1. 多平台账单汇总表（主表）
        if self.all_bills:
            ws_summary = wb.create_sheet(title="多平台账单汇总")
            self._write_bills_to_sheet(ws_summary, self.all_bills, "多平台账单汇总", sort_by_time=True)

        # 2. 各平台单独工作表
        for platform in ['天机', '小台风', '妙月']:
            platform_bills = self.platform_bills[platform]
            if platform_bills:
                ws_platform = wb.create_sheet(title=f"{platform}账单")
                sort_needed = (platform == '天机')
                self._write_bills_to_sheet(ws_platform, platform_bills, f"{platform}平台账单", sort_by_time=sort_needed)

        # 3. 平台账号汇总表
        if self.account_summary:
            ws_accounts = wb.create_sheet(title="平台账号汇总")
            self._write_account_summary_to_sheet(ws_accounts)

        # 4. 数据统计表
        ws_stats = wb.create_sheet(title="数据统计")
        self._write_statistics_to_sheet(ws_stats)

        # 5. 错误日志表
        if self.error_logs:
            ws_errors = wb.create_sheet(title="错误日志")
            self._write_error_logs_to_sheet(ws_errors)

        # 6. 字段映射说明表（新增）
        ws_mapping = wb.create_sheet(title="字段映射说明")
        self._write_field_mapping_to_sheet(ws_mapping)

        wb.save(excel_file)
        print(f"\n{Fore.GREEN}✅ Excel文件已保存：{excel_file}")

        return excel_file

    def _write_bills_to_sheet(self, ws, bills_data, sheet_title, sort_by_time=False):
        """将账单数据写入工作表（严格按照V2.0规范）"""
        if not bills_data:
            ws.append(["无数据"])
            return

        # 转换为DataFrame
        df = pd.DataFrame(bills_data)

        # 按交易时间降序排序
        if 'trans_time' in df.columns and sort_by_time:
            df['trans_time_temp'] = pd.to_datetime(df['trans_time'], errors='coerce', format='%Y-%m-%d %H:%M:%S')
            df = df.dropna(subset=['trans_time_temp'])
            df = df.sort_values('trans_time_temp', ascending=False)
            df = df.drop('trans_time_temp', axis=1)

        # 重新排列列顺序（按照规范文档顺序）
        column_order = ['order_no', 'iccid', 'card_number', 'trans_time', 'sale_price', 'cost_price',
                        'commission', 'customer_name', 'product_name', 'operator', 'income_type',
                        'remark', 'platform', 'account']

        # 只保留存在的列
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment_num = Alignment(horizontal='right', vertical='center')
        data_alignment_str = Alignment(horizontal='left', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头（使用字段映射）
        for col_idx, field_code in enumerate(existing_columns, 1):
            field_name = self.FIELD_MAPPING.get(field_code, field_code)
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # 设置列宽
            width = self.COLUMN_WIDTHS.get(field_name, 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 写入数据
        for row_idx, row in df.iterrows():
            for col_idx, field_code in enumerate(existing_columns, 1):
                cell_value = row[field_code]

                # 处理空值：空字符串或None都设置为None（Excel显示为空）
                if cell_value == "" or pd.isna(cell_value) or cell_value is None:
                    cell_value = None

                cell = ws.cell(row=row_idx + 2, column=col_idx, value=cell_value)
                cell.font = data_font
                cell.border = thin_border

                # 设置对齐方式
                field_name = self.FIELD_MAPPING.get(field_code, field_code)
                if field_name in ['售价（元）', '成本（元）', '佣金（元）']:
                    cell.alignment = data_alignment_num
                    # 格式化为两位小数
                    if isinstance(cell_value, (int, float)):
                        cell.number_format = '0.00'
                elif field_name in ['平台', '账号', '收入类型', '运营商']:
                    cell.alignment = data_alignment_center
                else:
                    cell.alignment = data_alignment_str

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        # 添加标题
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(existing_columns))
        title_cell = ws.cell(row=1, column=1, value=sheet_title)
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

        # 冻结表头
        ws.freeze_panes = ws['A3']

    def _write_account_summary_to_sheet(self, ws):
        """写入平台账号汇总表"""
        # 表头
        headers = ['平台', '账号', '总余额（元）', '可提现余额（元）', '不可提现余额（元）',
                   f'最近{self.config["common"].get("days_for_recent", 30)}天收益（元）',
                   f'最近{self.config["common"].get("days_for_recent", 30)}天提现（元）',
                   f'最近{self.config["common"].get("days_for_recent", 30)}天退款（元）',
                   '总账单数', '最后查询页码']

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment_num = Alignment(horizontal='right', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        column_widths = [10, 20, 15, 15, 15, 18, 18, 18, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 写入数据
        row_idx = 3
        for account_info in self.account_summary:
            platform = account_info.get('平台', '')
            username = account_info.get('账号', '')
            balance = account_info.get('balance', 0)
            withdrawable = account_info.get('withdrawable', account_info.get('balance', 0))
            non_withdrawable = account_info.get('non_withdrawable', 0)
            recent_income = account_info.get('recent_income', 0)
            recent_withdraw = account_info.get('recent_withdraw', 0)
            recent_refund = account_info.get('recent_refund', 0)
            total_bills = account_info.get('total_bills', 0)
            last_query_page = account_info.get('last_query_page', 0)

            data_row = [platform, username, balance, withdrawable, non_withdrawable,
                        recent_income, recent_withdraw, recent_refund, total_bills, last_query_page]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border

                # 设置对齐方式
                if col_idx in [1, 2, 9, 10]:  # 平台、账号、总账单数、最后查询页码居中
                    cell.alignment = data_alignment_center
                elif col_idx in [3, 4, 5, 6, 7, 8]:  # 金额右对齐
                    cell.alignment = data_alignment_num
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                else:
                    cell.alignment = data_alignment_center

            row_idx += 1

        # 添加汇总行
        if self.account_summary:
            row_idx += 1

            totals = [
                "",
                "总计",
                sum(acc.get('balance', 0) for acc in self.account_summary),
                sum(acc.get('withdrawable', acc.get('balance', 0)) for acc in self.account_summary),
                sum(acc.get('non_withdrawable', 0) for acc in self.account_summary),
                sum(acc.get('recent_income', 0) for acc in self.account_summary),
                sum(acc.get('recent_withdraw', 0) for acc in self.account_summary),
                sum(acc.get('recent_refund', 0) for acc in self.account_summary),
                sum(acc.get('total_bills', 0) for acc in self.account_summary),
                ""
            ]

            for col_idx, total in enumerate(totals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=total)
                if col_idx > 2:
                    cell.font = Font(bold=True)
                if col_idx in [3, 4, 5, 6, 7, 8]:
                    cell.alignment = data_alignment_num
                    if isinstance(total, (int, float)):
                        cell.number_format = '0.00'

            # 设置汇总行样式
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                if col_idx > 2:
                    cell.font = Font(bold=True)

        # 自动调整行高
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        # 添加标题
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="平台账号汇总")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')

        # 冻结表头
        ws.freeze_panes = ws['A3']

    def _write_statistics_to_sheet(self, ws):
        """写入数据统计表"""
        net_income = self.summary_data['total_recent_income'] - self.summary_data['total_recent_refund']

        stats = [
            ["统计项目", "数值", "说明"],
            ["总账号数", self.summary_data['total_accounts'], "配置文件中所有平台账号总数"],
            ["成功查询账号", self.summary_data['success_accounts'], "成功登录并获取数据的账号数"],
            ["查询失败账号", self.summary_data['failed_accounts'], "登录失败或查询异常的账号数"],
            ["查询完成率",
             f"{(self.summary_data['success_accounts'] / self.summary_data['total_accounts'] * 100):.1f}%",
             "成功账号占总账号比例"],
            ["总余额（元）", self.summary_data['total_balance'], "所有账号余额总和"],
            [f"最近{self.config['common'].get('days_for_recent', 30)}天总收益（元）",
             self.summary_data['total_recent_income'],
             f"最近{self.config['common'].get('days_for_recent', 30)}天的正数收益总和"],
            [f"最近{self.config['common'].get('days_for_recent', 30)}天总提现（元）",
             self.summary_data['total_recent_withdraw'],
             f"最近{self.config['common'].get('days_for_recent', 30)}天的提现支出总和"],
            [f"最近{self.config['common'].get('days_for_recent', 30)}天总退款（元）",
             self.summary_data['total_recent_refund'],
             f"最近{self.config['common'].get('days_for_recent', 30)}天的退款总和"],
            ["净收益（总收益 - 总退款）（元）", net_income, "实际净收益（总收益减去退款）"],
            ["总账单数", self.summary_data['total_bills'], "所有账单记录总数"],
            ["", "", ""],
            ["平台", "账号数", "总余额（元）", "最近收益（元）", "最近提现（元）", "最近退款（元）", "净收益（元）", "账单数",
             "占比"]
        ]

        # 添加各平台统计
        for platform in ['天机', '小台风', '妙月']:
            platform_accounts = [acc for acc in self.account_summary if acc['平台'] == platform]
            platform_balance = sum(acc.get('balance', 0) for acc in platform_accounts)
            platform_recent_income = sum(acc.get('recent_income', 0) for acc in platform_accounts)
            platform_recent_withdraw = sum(acc.get('recent_withdraw', 0) for acc in platform_accounts)
            platform_recent_refund = sum(acc.get('recent_refund', 0) for acc in platform_accounts)
            platform_net_income = platform_recent_income - platform_recent_refund
            platform_bills_count = len(self.platform_bills[platform])
            percentage = (platform_bills_count / self.summary_data['total_bills'] * 100) if self.summary_data[
                                                                                                'total_bills'] > 0 else 0

            stats.append([
                platform,
                len(platform_accounts),
                platform_balance,
                platform_recent_income,
                platform_recent_withdraw,
                platform_recent_refund,
                platform_net_income,
                platform_bills_count,
                f"{percentage:.1f}%"
            ])

        for row_idx, row in enumerate(stats, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if row_idx == 1 or row_idx == 13:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx <= 12:
                    if col_idx == 2 and isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    if col_idx in [3, 4, 5, 6, 7, 8] and isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        column_widths = [20, 15, 30, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 添加标题
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        title_cell = ws.cell(row=1, column=1, value="数据统计")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

    def _write_error_logs_to_sheet(self, ws):
        """写入错误日志表"""
        headers = ['时间', '平台', '账号', '错误类型', '错误信息']

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        column_widths = [18, 10, 15, 15, 50]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 写入数据
        for row_idx, error in enumerate(self.error_logs, 2):
            data_row = [
                error.get('时间', ''),
                error.get('平台', ''),
                error.get('账号', ''),
                error.get('错误类型', ''),
                error.get('错误信息', '')
            ]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        # 添加标题
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="错误日志")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

        # 冻结表头
        ws.freeze_panes = ws['A3']

    def _write_field_mapping_to_sheet(self, ws):
        """写入字段映射说明表"""
        headers = ['字段编码', '字段名称', '数据类型', '业务定义', '空值处理规则', '数据校验规则']

        # 字段映射数据（按照规范文档）
        field_data = [
            ['order_no', '订单号', '字符串', '各平台唯一订单标识，跨平台对账主键', '源字段为空则留空',
             '非空时长度≤64字符'],
            ['iccid', 'ICCID', '字符串', '流量卡唯一标识，卡级交易溯源核心', '源字段为空则留空',
             '非空时符合ICCID编码规则（19-20位）'],
            ['card_number', '卡号', '字符串', '流量卡关联的手机号/物理卡号，辅助卡信息溯源', '源字段为空则留空',
             '非空时长度≤20字符，仅含数字'],
            ['trans_time', '交易时间', '字符串', '订单交易/结算的时间节点，财务对账核心维度',
             '源字段为空/格式异常则留空', '非空时符合"YYYY-MM-DD HH:MM:SS"格式'],
            ['sale_price', '售价（元）', '浮点数', '订单实际销售金额，收入核算基础', '源字段为空则留空',
             '非空时≥0，精度≤2位小数'],
            ['cost_price', '成本（元）', '浮点数', '订单对应卡/套餐的采购成本，佣金核算基础', '源字段为空则留空',
             '非空时≥0、≤售价，精度≤2位小数'],
            ['commission', '佣金（元）', '浮点数', '订单实际收益（原"利润"）', '源字段为空则留空',
             '非空时精度≤2位小数，=售价-成本（误差±0.01）'],
            ['customer_name', '客户名称', '字符串', '交易对应的客户/企业名称，业务溯源核心', '源字段为空则留空',
             '非空时长度≤64字符，去除特殊字符'],
            ['product_name', '套餐/产品名称', '字符串', '订单对应的套餐/产品类型，区分业务场景', '源字段为空则留空',
             '非空时长度≤128字符'],
            ['operator', '运营商', '字符串', '流量卡所属基础运营商，区分业务维度', '源字段为空则留空',
             '非空时仅允许"中国移动/中国电信/中国联通/其他"'],
            ['income_type', '收入类型', '字符串', '区分交易场景，支撑业务类型分析', '推导依据为空则留空',
             '非空时仅允许"出售套餐/续费/未分类"'],
            ['remark', '备注', '字符串', '交易补充说明，异常账单排查依据', '源字段为空则留空', '非空时长度≤256字符']
        ]

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        column_widths = [15, 15, 10, 40, 30, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 写入字段数据
        for row_idx, field_row in enumerate(field_data, 2):
            for col_idx, value in enumerate(field_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 25

        # 添加标题
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="字段映射说明（V2.0规范）")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')

        # 冻结表头
        ws.freeze_panes = ws['A3']

    def run(self):
        """运行主程序"""
        print(f"{Fore.CYAN}{'=' * 60}")
        print(f"{Fore.CYAN}多平台账单查询系统 V2.0（严格按照规范文档开发）")
        print(f"{Fore.CYAN}启动时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{Fore.CYAN}{'=' * 60}")

        try:
            self.query_all_platforms()

            if self.all_bills or self.account_summary:
                excel_file = self.export_to_excel()

                print(f"\n{Fore.GREEN}{'=' * 60}")
                print(f"{Fore.GREEN}导出文件包含以下工作表：")
                print(f"{Fore.GREEN}1. 多平台账单汇总 - 所有平台的标准化账单数据（V2.0规范）")
                print(f"{Fore.GREEN}2. 天机账单 - 天机平台的标准化账单数据")
                print(f"{Fore.GREEN}3. 小台风账单 - 小台风平台的标准化账单数据")
                print(f"{Fore.GREEN}4. 妙月账单 - 妙月平台的标准化账单数据")
                print(f"{Fore.GREEN}5. 平台账号汇总 - 各账号余额、收益、提现、退款统计")
                print(f"{Fore.GREEN}6. 数据统计 - 整体统计信息（包含净收益计算）")
                print(f"{Fore.GREEN}7. 字段映射说明 - V2.0规范字段映射说明")
                if self.error_logs:
                    print(f"{Fore.GREEN}8. 错误日志 - 所有错误记录")
                print(f"{Fore.GREEN}{'=' * 60}")

                print(f"\n{Fore.YELLOW}💡 V2.0规范特点：")
                print(f"{Fore.YELLOW}   1. 空值统一留空（不再填充'未采集'等默认值）")
                print(f"{Fore.YELLOW}   2. 严格按照字段编码映射表处理数据")
                print(f"{Fore.YELLOW}   3. 可计算字段按规则计算，依赖字段空则结果空")
                print(f"{Fore.YELLOW}   4. 标准化格式化（时间、运营商、收入类型等）")
            else:
                print(f"{Fore.YELLOW}⚠️  未获取到任何数据")

        except Exception as e:
            logger.error(f"程序运行异常: {str(e)}", exc_info=True)
            print(f"{Fore.RED}❌ 程序运行异常：{str(e)}")
            print(f"{Fore.RED}详细错误信息请查看日志文件：{log_file}")

        finally:
            print(f"\n{Fore.CYAN}{'=' * 60}")
            print(f"{Fore.CYAN}程序执行完成")
            print(f"{Fore.CYAN}详细日志请查看: {log_file}")
            print(f"{Fore.CYAN}{'=' * 60}")


# ======================== 配置日志 ========================
def setup_logging():
    """配置日志系统"""
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"platform_query_v2_{timestamp}.log"

    logger = logging.getLogger('platform_query_v2')
    logger.setLevel(logging.INFO)

    if logger.hasHandlers():
        logger.handlers.clear()

    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.WARNING)

    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger, log_file


# ======================== 主程序入口 ========================
def main():
    """主函数"""
    # 配置日志
    global logger, log_file
    logger, log_file = setup_logging()

    try:
        from fake_useragent import UserAgent
    except ImportError:
        print(f"{Fore.YELLOW}⚠️  缺少依赖库: fake_useragent")
        print(f"{Fore.YELLOW}正在安装依赖库...def load_config(self):")
        import subprocess
        import sys

        packages = ['fake_useragent', 'pandas', 'openpyxl', 'colorama', 'requests', 'beautifulsoup4']
        for package in packages:
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            except:
                pass

        print(f"{Fore.GREEN}✅ 依赖安装完成，请重新运行程序")
        return

    manager = MultiPlatformManagerV2()
    manager.run()


if __name__ == "__main__":
    main()