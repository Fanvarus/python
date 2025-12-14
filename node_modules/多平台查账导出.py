# multi_platform_query_system_optimized.py
import re
import json
import time
import os
import pandas as pd
import numpy as np
import configparser
import concurrent.futures
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import requests
from requests import Session
from bs4 import BeautifulSoup
from colorama import Fore, init, Style, Back
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import warnings
import logging
from fake_useragent import UserAgent

warnings.filterwarnings('ignore')

# 初始化colorama
init(autoreset=True)


# 配置日志
def setup_logging():
    """配置日志系统"""
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"platform_query_{timestamp}.log"

    # 创建logger
    logger = logging.getLogger('platform_query')
    logger.setLevel(logging.INFO)

    # 防止重复添加handler
    if logger.hasHandlers():
        logger.handlers.clear()

    # 文件处理器
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)

    # 控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.WARNING)

    # 格式化器
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger, log_file


# 创建logger
logger, log_file = setup_logging()


# ======================== 配置文件读取 ========================
class ConfigManager:
    """配置文件管理器"""

    def __init__(self, config_path: str = "multi_platform_config.ini"):
        self.config_path = config_path
        self.config = configparser.ConfigParser()

    def load_config(self):
        """加载配置文件"""
        if not os.path.exists(self.config_path):
            self.create_default_config()

        self.config.read(self.config_path, encoding='utf-8')

        # 读取通用配置
        common_config = {
            'output_path': self.config.get('SETTINGS', 'output_path', fallback='桌面'),
            'bill_page_size': self.config.getint('SETTINGS', 'bill_page_size', fallback=10),
            'query_all_bills': self.config.getboolean('SETTINGS', 'query_all_bills', fallback=False),
            'max_workers': self.config.getint('SETTINGS', 'max_workers', fallback=5),
            'enable_threading': self.config.getboolean('SETTINGS', 'enable_threading', fallback=True),
            'request_timeout': self.config.getint('SETTINGS', 'request_timeout', fallback=15),
            'platform_delay': self.config.getfloat('SETTINGS', 'platform_delay', fallback=0.5),
            'days_for_recent': self.config.getint('SETTINGS', 'days_for_recent', fallback=30)
        }

        # 读取天机平台账号
        tianji_accounts = {}
        if self.config.has_section('TIANJI_ACCOUNTS'):
            tianji_accounts = dict(self.config.items('TIANJI_ACCOUNTS'))

        # 读取小台风平台账号
        xiaotaifeng_accounts = []
        if self.config.has_section('XIAOTAIENG_ACCOUNTS'):
            xiaotaifeng_accounts = [
                {"username": k, "password": v}
                for k, v in dict(self.config.items('XIAOTAIENG_ACCOUNTS')).items()
            ]

        # 读取妙月平台账号
        miaoyue_accounts = {}
        if self.config.has_section('MIAOYUE_ACCOUNTS'):
            miaoyue_accounts = dict(self.config.items('MIAOYUE_ACCOUNTS'))

        return {
            'common': common_config,
            'tianji': {
                'base_url': 'https://sys.szlaina.com',
                'accounts': tianji_accounts,
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            },
            'xiaotaifeng': {
                'base_url': 'http://123.56.58.202:8085',
                'accounts': xiaotaifeng_accounts,
                'login_url': '/user/login',
                'balance_url': '/profit/profitcanwithdraw',
                'bill_list_url': '/profit/list'
            },
            'miaoyue': {
                'base_url': 'https://sapi.musmoon.com',
                'accounts': miaoyue_accounts,
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
        }

    def create_default_config(self):
        """创建默认配置文件"""
        self.config['SETTINGS'] = {
            '; 输出路径（支持"桌面"或具体路径）': '',
            'output_path': '桌面',
            '; 每次查询账单条数': '',
            'bill_page_size': '10',
            '; 是否查询全部账单（True=全部，False=仅查询bill_page_size条）': '',
            'query_all_bills': 'False',
            '; 最大线程数（用于多平台并发查询）': '',
            'max_workers': '5',
            '; 是否启用多线程': '',
            'enable_threading': 'True',
            '; 请求超时时间（秒）': '',
            'request_timeout': '15',
            '; 平台间查询延迟（秒）': '',
            'platform_delay': '0.5',
            '; 统计最近多少天的收益': '',
            'days_for_recent': '30'
        }

        self.config['TIANJI_ACCOUNTS'] = {
            '; 格式：账号名 = 加密密码': '',
            'Wdy': '90535de091e878a11a3e1724ab22bc10',
            'CFWS': 'a71a5ba407b3e4333d1a89689779446b',
            '晨阳科技': 'a71a5ba407b3e4333d1a89689779446b'
        }

        self.config['XIAOTAIENG_ACCOUNTS'] = {
            '; 格式：账号名 = 密码': '',
            '超凡威视': '525231314.',
            '塘厦益雅贸易': '112233',
            '小姜安防': 'Wu5626480',
            '南鲁集镇': '525231314.'
        }

        self.config['MIAOYUE_ACCOUNTS'] = {
            '; 格式：账号名 = 加密密码': '',
            'jiweishidai': '6D218509562ED94DB2808E28AE3DB3BB',
            'huangfangyi': '6F0A6BC78A79D8E922410BB0971FDE0A',
            '蓝硕商贸科技': '6F0A6BC78A79D8E922410BB0971FDE0A'
        }

        with open(self.config_path, 'w', encoding='utf-8') as f:
            self.config.write(f)

        print(f"✅ 已创建默认配置文件：{self.config_path}")
        print(f"⚠️  请修改配置文件中的API地址和账号信息")


# ======================== 数据清洗和标准化 ========================
class DataProcessor:
    """数据处理和标准化类"""

    @staticmethod
    def standardize_operator(operator: str) -> str:
        """运营商标准化"""
        if not operator or pd.isna(operator) or operator in ['', '未采集', '未采集 ']:
            return ""

        operator = str(operator).strip().upper()
        if 'CM' in operator or '移动' in operator:
            return "中国移动"
        elif '电信' in operator or 'TELECOM' in operator:
            return "中国电信"
        elif '联通' in operator or 'UNICOM' in operator:
            return "中国联通"
        else:
            return operator

    @staticmethod
    def classify_income_type(remarks: str, order_content: str = "", income_type: str = "", bill_type: str = "",
                             amount: float = 0.0) -> str:
        """收入类型归类，增加对负数金额的判断"""
        # 统一转换为字符串处理
        remarks_str = str(remarks or "").strip().lower()
        order_content_str = str(order_content or "").strip().lower()
        income_type_str = str(income_type or "").strip().lower()
        bill_type_str = str(bill_type or "").strip()

        # 先判断是否为负数（表示支出）
        if amount < 0:
            # 判断是提现还是退款
            if "提现" in remarks_str or "withdraw" in remarks_str or bill_type_str == "userWithdraw":
                return "提现支出"
            elif "退款" in remarks_str or "refund" in remarks_str or bill_type_str == "orderRefundBill":
                return "退款"
            else:
                return "其他支出"

        # 正数金额的处理逻辑
        if bill_type_str:
            if bill_type_str == "orderCommissionBill":
                return "佣金收入"
            elif bill_type_str == "userWithdraw":
                return "提现支出"
            elif bill_type_str == "orderRefundBill":
                return "退款"
            else:
                return f"其他-{bill_type_str}"

        if "套餐" in order_content_str:
            return "出售套餐"

        if "续费" in remarks_str:
            return "续费"
        elif "出售" in remarks_str or "套餐" in remarks_str:
            return "出售套餐"

        if "续费" in income_type_str:
            return "续费"
        elif "出售" in income_type_str or "套餐" in income_type_str:
            return "出售套餐"

        return "未分类"

    @staticmethod
    def safe_float(value, default=0.0) -> float:
        """安全转换为浮点数"""
        if pd.isna(value) or value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    @staticmethod
    def safe_str(value, default="") -> str:
        """安全转换为字符串，如果是'未采集'返回空字符串"""
        if pd.isna(value) or value is None:
            return default

        str_value = str(value).strip()
        if str_value in ['未采集', '未采集 ', '无', '无 ', '-']:
            return ""
        return str_value

    @staticmethod
    def standardize_datetime(dt_str: str) -> str:
        """标准化时间格式"""
        if not dt_str or pd.isna(dt_str) or str(dt_str).strip() in ['无交易时间', '未知时间', '-']:
            return ""

        try:
            dt_str = str(dt_str).strip()

            formats = [
                '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                '%Y-%m-%dT%H:%M:%S', '%Y%m%d %H:%M:%S',
                '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M',
                '%Y年%m月%d日 %H:%M:%S', '%Y-%m-%d'
            ]

            for fmt in formats:
                try:
                    dt = datetime.strptime(dt_str, fmt)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    continue

            if dt_str.isdigit():
                if len(dt_str) == 10:
                    dt = datetime.fromtimestamp(int(dt_str))
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                elif len(dt_str) == 13:
                    dt = datetime.fromtimestamp(int(dt_str) / 1000)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')

        except Exception as e:
            logger.warning(f"时间格式转换失败: {dt_str}, 错误: {e}")

        return dt_str


# ======================== 天机平台客户端 ========================
class TianjiClient:
    def __init__(self, config: dict):
        self.cfg = config
        self.common_cfg = config['common']
        self.base_url = config['tianji']['base_url']

        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "keep-alive",
            "Referer": f"{self.base_url}/Index/index",
            "Sec-Ch-Ua": '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "iframe",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
            "Priority": "u=0, i"
        }

        self.data_processor = DataProcessor()
        self.account_info = {}

    def login(self, username: str, password: str) -> tuple[Session, bool, str]:
        """登录天机平台"""
        session = Session()
        login_url = f"{self.base_url}/Login/doLogin"
        login_data = {"u_name": username, "pwd": password, "encry": "1"}

        try:
            session.get(f"{self.base_url}/Index/index", headers=self.headers,
                        verify=False, timeout=self.common_cfg['request_timeout'])
        except Exception as e:
            logger.warning(f"访问首页异常: {str(e)}")

        error_msgs = []

        for retry in range(3):
            try:
                resp = session.post(login_url, data=login_data, headers=self.headers,
                                    verify=False, timeout=self.common_cfg['request_timeout'])
                resp.encoding = "utf-8"

                if session.cookies.get("PHPSESSID"):
                    logger.info(f"天机平台账号 {username} 登录成功")
                    return session, True, ""
                else:
                    error_msg = f"第{retry + 1}次尝试登录失败，无PHPSESSID"
                    error_msgs.append(error_msg)
                    logger.warning(f"天机平台账号 {username} {error_msg}")

                    if retry < 2:
                        time.sleep(0.5)
                        continue
                    else:
                        final_error = f"登录失败: {'; '.join(error_msgs)}"
                        logger.error(f"天机平台账号 {username} {final_error}")
                        return session, False, final_error

            except requests.exceptions.Timeout:
                error_msg = f"第{retry + 1}次尝试登录超时"
                error_msgs.append(error_msg)
                logger.warning(f"天机平台账号 {username} {error_msg}")
                if retry < 2:
                    time.sleep(0.5)
                    continue
            except requests.exceptions.ConnectionError:
                error_msg = f"第{retry + 1}次尝试连接失败"
                error_msgs.append(error_msg)
                logger.warning(f"天机平台账号 {username} {error_msg}")
                if retry < 2:
                    time.sleep(0.5)
                    continue
            except Exception as e:
                error_msg = f"第{retry + 1}次尝试异常: {str(e)}"
                error_msgs.append(error_msg)
                logger.error(f"天机平台账号 {username} {error_msg}")
                if retry < 2:
                    time.sleep(0.5)
                    continue

        final_error = f"多次尝试登录失败: {'; '.join(error_msgs)}"
        logger.error(f"天机平台账号 {username} {final_error}")
        return session, False, final_error

    def get_balance(self, session: Session, username: str) -> tuple[float, str]:
        """获取余额"""
        profit_url = f"{self.base_url}/Profit/companyProfit"
        try:
            session.get(f"{self.base_url}/Profit/listProfit", headers=self.headers,
                        verify=False, timeout=self.common_cfg['request_timeout'])

            resp = session.get(profit_url, headers=self.headers, verify=False,
                               timeout=self.common_cfg['request_timeout'])
            resp.encoding = "utf-8"

            if resp.status_code != 200:
                error_msg = f"余额接口请求失败，状态码：{resp.status_code}"
                logger.warning(f"天机平台账号 {username} {error_msg}")
                return 0.0, error_msg

            html = resp.text

            try:
                soup = BeautifulSoup(html, 'html.parser')

                balance_elements = soup.find_all(text=re.compile(r'余额[:：]?\s*'))
                for element in balance_elements:
                    parent = element.parent
                    if parent:
                        parent_text = parent.get_text().strip()
                        matches = re.findall(r'[-+]?\d+\.?\d*', parent_text)
                        for match in matches:
                            try:
                                num = float(match)
                                pattern1 = r'余额[:：]\s*' + re.escape(match)
                                pattern2 = match + r'\s*元'

                                if (re.search(pattern1, parent_text) or
                                        re.search(pattern2, parent_text)):
                                    logger.info(f"天机平台账号 {username} 提取余额: {num:.2f} 元")

                                    self.account_info[username] = {
                                        'balance': num,
                                        'recent_income': 0.0,
                                        'recent_withdraw': 0.0,
                                        'recent_refund': 0.0,
                                        'total_bills': 0
                                    }
                                    return num, ""
                            except ValueError:
                                continue

                for class_name in ['info-box-number', 'balance', 'amount', 'money']:
                    elements = soup.find_all(class_=re.compile(class_name, re.I))
                    for element in elements:
                        element_text = element.get_text().strip()
                        matches = re.findall(r'[-+]?\d+\.?\d*', element_text)
                        for match in matches:
                            try:
                                num = float(match)
                                parent = element.parent
                                for _ in range(3):
                                    if parent:
                                        parent_text = parent.get_text()
                                        if '余额' in parent_text:
                                            logger.info(f"天机平台账号 {username} class提取余额: {num:.2f} 元")

                                            self.account_info[username] = {
                                                'balance': num,
                                                'recent_income': 0.0,
                                                'recent_withdraw': 0.0,
                                                'recent_refund': 0.0,
                                                'total_bills': 0
                                            }
                                            return num, ""
                                    parent = parent.parent if parent else None
                            except ValueError:
                                continue

            except Exception as e:
                logger.error(f"天机平台账号 {username} BeautifulSoup解析异常: {str(e)}")

            precise_patterns = [
                r'余额[:：]\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'(?:可用)?余额\s*[:：]?\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'([-+]?\d+(?:\.\d+)?)\s*元\s*(?:的)?余额',
                r'当前余额\s*[:：]?\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
                r'账户余额\s*[:：]?\s*([-+]?\d+(?:\.\d+)?)\s*(?:元|￥|¥)?',
            ]

            for pattern in precise_patterns:
                matches = re.findall(pattern, html, re.IGNORECASE)
                for match in matches:
                    try:
                        num = float(match)
                        logger.info(f"天机平台账号 {username} 正则提取余额: {num:.2f} 元")

                        self.account_info[username] = {
                            'balance': num,
                            'recent_income': 0.0,
                            'recent_withdraw': 0.0,
                            'recent_refund': 0.0,
                            'total_bills': 0
                        }
                        return num, ""
                    except ValueError:
                        continue

            error_msg = "未提取到余额数据"
            logger.warning(f"天机平台账号 {username} {error_msg}")
            return 0.0, error_msg

        except requests.exceptions.Timeout:
            error_msg = "余额查询超时"
            logger.error(f"天机平台账号 {username} {error_msg}")
            return 0.0, error_msg
        except requests.exceptions.ConnectionError:
            error_msg = "余额查询连接失败"
            logger.error(f"天机平台账号 {username} {error_msg}")
            return 0.0, error_msg
        except Exception as e:
            error_msg = f"余额查询异常: {str(e)}"
            logger.error(f"天机平台账号 {username} {error_msg}")
            return 0.0, error_msg

    def get_bills(self, session: Session, username: str) -> tuple[List[Dict], str]:
        """获取账单"""
        all_bills = []
        page = 1
        page_size = self.common_cfg['bill_page_size']
        errors = []

        while True:
            bill_url = f"{self.base_url}/Profit/billDetail"
            bill_data = {
                "page": page,
                "limit": page_size,
                "start_time": "",
                "end_time": "",
                "type": ""
            }

            bill_headers = self.headers.copy()
            bill_headers.update({
                "Accept": "*/*",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": f"{self.base_url}/Profit/listBillDetail"
            })

            try:
                resp = session.post(bill_url, data=bill_data, headers=bill_headers,
                                    verify=False, timeout=self.common_cfg['request_timeout'])
                resp.encoding = "utf-8"

                if resp.status_code == 200:
                    bill_json = resp.json()
                    if bill_json.get("status") in [1, "1"] and bill_json.get("message") == "成功":
                        bill_list = bill_json.get("list", [])
                        if bill_list:
                            processed_bills = self._process_bills(bill_list, username)
                            all_bills.extend(processed_bills)
                        else:
                            break
                    else:
                        error_msg = f"账单接口状态异常: {bill_json.get('status')}, {bill_json.get('message')}"
                        errors.append(error_msg)
                        break
                else:
                    error_msg = f"账单接口请求失败，状态码：{resp.status_code}"
                    errors.append(error_msg)
                    break

            except requests.exceptions.Timeout:
                error_msg = f"第{page}页账单查询超时"
                errors.append(error_msg)
                break
            except requests.exceptions.ConnectionError:
                error_msg = f"第{page}页账单查询连接失败"
                errors.append(error_msg)
                break
            except Exception as e:
                error_msg = f"第{page}页账单查询异常: {str(e)}"
                errors.append(error_msg)
                break

            if not self.common_cfg['query_all_bills'] or len(all_bills) >= 100:
                break
            page += 1
            time.sleep(0.3)

        if all_bills:
            logger.info(f"天机平台账号 {username} 获取到 {len(all_bills)} 条账单")
        else:
            logger.warning(f"天机平台账号 {username} 未获取到账单数据")

        error_info = "; ".join(errors) if errors else ""
        return all_bills, error_info

    def _process_bills(self, raw_bills: List[Dict], username: str) -> List[Dict]:
        """处理天机平台账单数据"""
        processed = []
        recent_income = 0.0
        recent_withdraw = 0.0
        recent_refund = 0.0
        days_for_recent = self.common_cfg.get('days_for_recent', 30)
        cutoff_date = datetime.now() - timedelta(days=days_for_recent)

        for bill in raw_bills:
            trans_time_str = bill.get('trans_time_format') or bill.get('create_time_format')
            trans_time = None
            if trans_time_str:
                try:
                    trans_time = datetime.strptime(str(trans_time_str), '%Y-%m-%d %H:%M:%S')
                except:
                    pass

            # 获取金额信息
            profit = self.data_processor.safe_float(bill.get('profit', 0))
            amount = self.data_processor.safe_float(bill.get('income_money', 0))
            cost = amount - profit if amount > 0 else self.data_processor.safe_float(bill.get('cost_money', 0))

            # 分类收入类型
            income_type = self.data_processor.classify_income_type(
                bill.get('remarks', ''),
                bill.get('order_name', ''),
                '',
                '',
                profit  # 传入金额用于判断是否为负数
            )

            # 统计最近收益
            if trans_time and trans_time >= cutoff_date:
                if profit > 0:
                    # 正数：收入
                    recent_income += profit
                elif profit < 0:
                    # 负数：判断是提现还是退款
                    if income_type == "提现支出":
                        recent_withdraw += abs(profit)  # 取绝对值累加
                    elif income_type == "退款":
                        recent_refund += abs(profit)  # 取绝对值累加

            processed_bill = {
                '订单号': self.data_processor.safe_str(bill.get('order_no', '')),
                'ICCID': self.data_processor.safe_str(bill.get('iccid', '')),
                '卡号': '',
                '交易时间': self.data_processor.standardize_datetime(trans_time_str),
                '售价（元）': amount,
                '成本（元）': cost,
                '佣金（元）': profit,
                '客户名称': self.data_processor.safe_str(bill.get('company_name', '')),
                '套餐/产品名称': self.data_processor.safe_str(bill.get('order_name', '')),
                '运营商': self.data_processor.standardize_operator(bill.get('second_operator_name', '')),
                '收入类型': income_type,
                '备注': self.data_processor.safe_str(bill.get('remarks', '')),
                '平台': '天机',
                '账号': username
            }
            processed.append(processed_bill)

        # 更新账号信息
        if username in self.account_info:
            self.account_info[username]['recent_income'] = recent_income
            self.account_info[username]['recent_withdraw'] = recent_withdraw
            self.account_info[username]['recent_refund'] = recent_refund
            self.account_info[username]['total_bills'] += len(processed)

        return processed

    def get_account_info(self, username: str) -> Dict:
        """获取账号信息"""
        return self.account_info.get(username, {
            'balance': 0.0,
            'recent_income': 0.0,
            'recent_withdraw': 0.0,
            'recent_refund': 0.0,
            'total_bills': 0
        })


# ======================== 小台风平台客户端 ========================
class XiaoTaiFengClient:
    def __init__(self, config: dict):
        self.cfg = config
        self.common_cfg = config['common']
        self.base_url = config['xiaotaifeng']['base_url']
        self.login_url = self.base_url + config['xiaotaifeng']['login_url']
        self.balance_url = self.base_url + config['xiaotaifeng']['balance_url']
        self.bill_list_url = self.base_url + config['xiaotaifeng']['bill_list_url']

        self.request_headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "keep-alive",
            "Content-Type": "application/json",
            "Host": "123.56.58.202:8085",
            "Origin": "http://iot.xiaotaifeng.cn",
            "Referer": "http://iot.xiaotaifeng.cn/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36"
        }

        self.data_processor = DataProcessor()
        self.account_info = {}

    def login(self, username: str, password: str) -> tuple[Session, bool, str]:
        """登录小台风平台"""
        session = Session()
        session.headers.update(self.request_headers)

        try:
            login_data = {"username": username, "password": password}
            resp = session.post(
                self.login_url,
                json=login_data,
                timeout=self.common_cfg['request_timeout']
            )

            if resp.status_code == 200:
                result = resp.json()
                if result.get("code") == "0" and result.get("message") == "登录成功":
                    token = result.get("data", {}).get("token")
                    if token:
                        session.headers["X-Token"] = token
                        logger.info(f"小台风平台账号 {username} 登录成功")
                        return session, True, ""
                    else:
                        error_msg = "登录成功但未获取到token"
                        logger.error(f"小台风平台账号 {username} {error_msg}")
                        return session, False, error_msg
                else:
                    error_msg = f"登录失败: {result.get('message', '未知错误')}"
                    logger.error(f"小台风平台账号 {username} {error_msg}")
                    return session, False, error_msg
            else:
                error_msg = f"登录请求失败，状态码：{resp.status_code}"
                logger.error(f"小台风平台账号 {username} {error_msg}")
                return session, False, error_msg

        except requests.exceptions.Timeout:
            error_msg = "登录请求超时"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return session, False, error_msg
        except requests.exceptions.ConnectionError:
            error_msg = "无法连接到小台风平台"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return session, False, error_msg
        except Exception as e:
            error_msg = f"登录异常: {str(e)}"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return session, False, error_msg

    def get_balance(self, session: Session, username: str) -> tuple[float, str]:
        """获取余额"""
        try:
            resp = session.get(
                self.balance_url,
                timeout=self.common_cfg['request_timeout']
            )

            if resp.status_code == 200:
                result = resp.json()
                if result.get("code") == "0" and isinstance(result.get("data"), (int, float)):
                    balance = float(result["data"])
                    logger.info(f"小台风平台账号 {username} 余额: {balance:.2f} 元")

                    self.account_info[username] = {
                        'balance': balance,
                        'recent_income': 0.0,
                        'recent_withdraw': 0.0,
                        'recent_refund': 0.0,
                        'total_bills': 0
                    }
                    return balance, ""
                else:
                    error_msg = f"余额格式异常: {result}"
                    logger.error(f"小台风平台账号 {username} {error_msg}")
                    return 0.0, error_msg
            else:
                error_msg = f"余额请求失败，状态码：{resp.status_code}"
                logger.error(f"小台风平台账号 {username} {error_msg}")
                return 0.0, error_msg

        except requests.exceptions.Timeout:
            error_msg = "余额查询超时"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return 0.0, error_msg
        except requests.exceptions.ConnectionError:
            error_msg = "余额查询连接失败"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return 0.0, error_msg
        except Exception as e:
            error_msg = f"余额查询异常: {str(e)}"
            logger.error(f"小台风平台账号 {username} {error_msg}")
            return 0.0, error_msg

    def get_bills(self, session: Session, username: str) -> tuple[List[Dict], str]:
        """获取账单"""
        all_bills = []
        page = 1
        page_size = self.common_cfg['bill_page_size']
        errors = []

        while True:
            try:
                params = {
                    "paytype": "",
                    "account": "",
                    "productid": "",
                    "name": "",
                    "page": page,
                    "limit": page_size,
                    "sort": "-d.ID"
                }

                resp = session.get(
                    self.bill_list_url,
                    params=params,
                    timeout=self.common_cfg['request_timeout']
                )

                if resp.status_code == 200:
                    result = resp.json()
                    if result.get("code") == "0" and "data" in result and "items" in result["data"]:
                        bill_list = result["data"]["items"]
                        if bill_list:
                            processed_bills = self._process_bills(bill_list, username)
                            all_bills.extend(processed_bills)
                        else:
                            break
                    else:
                        error_msg = f"账单格式异常: {result.get('message', '未知错误')}"
                        errors.append(error_msg)
                        break
                else:
                    error_msg = f"账单请求失败，状态码：{resp.status_code}"
                    errors.append(error_msg)
                    break

            except requests.exceptions.Timeout:
                error_msg = f"第{page}页账单查询超时"
                errors.append(error_msg)
                break
            except requests.exceptions.ConnectionError:
                error_msg = f"第{page}页账单查询连接失败"
                errors.append(error_msg)
                break
            except Exception as e:
                error_msg = f"第{page}页账单查询异常: {str(e)}"
                errors.append(error_msg)
                break

            if not self.common_cfg['query_all_bills'] or len(all_bills) >= 100:
                break
            page += 1
            time.sleep(0.3)

        if all_bills:
            logger.info(f"小台风平台账号 {username} 获取到 {len(all_bills)} 条账单")

        error_info = "; ".join(errors) if errors else ""
        return all_bills, error_info

    def _process_bills(self, raw_bills: List[Dict], username: str) -> List[Dict]:
        """处理小台风平台账单数据"""
        processed = []
        recent_income = 0.0
        recent_withdraw = 0.0
        recent_refund = 0.0
        days_for_recent = self.common_cfg.get('days_for_recent', 30)
        cutoff_date = datetime.now() - timedelta(days=days_for_recent)

        for bill in raw_bills:
            trans_time_str = bill.get('createtime')
            trans_time = None
            if trans_time_str and trans_time_str != '-':
                try:
                    trans_time = datetime.strptime(str(trans_time_str), '%Y-%m-%d %H:%M:%S')
                except:
                    pass

            amount = self.data_processor.safe_float(bill.get('amount', 0))
            profit = self.data_processor.safe_float(bill.get('profit', 0))
            cost = amount - profit if amount > 0 else 0

            # 分类收入类型
            income_type = self.data_processor.classify_income_type(
                '',
                '',
                bill.get('incometype', ''),
                '',
                profit
            )

            # 统计最近收益
            if trans_time and trans_time >= cutoff_date:
                if profit > 0:
                    recent_income += profit
                elif profit < 0:
                    if income_type == "提现支出":
                        recent_withdraw += abs(profit)
                    elif income_type == "退款":
                        recent_refund += abs(profit)

            processed_bill = {
                '订单号': self.data_processor.safe_str(bill.get('orderid', bill.get('order_no', ''))),
                'ICCID': self.data_processor.safe_str(bill.get('iccid', '')),
                '卡号': '',
                '交易时间': self.data_processor.standardize_datetime(trans_time_str),
                '售价（元）': amount,
                '成本（元）': cost,
                '佣金（元）': profit,
                '客户名称': self.data_processor.safe_str(bill.get('account', '')),
                '套餐/产品名称': self.data_processor.safe_str(bill.get('productname', '')),
                '运营商': self.data_processor.standardize_operator(bill.get('yunyingshang', '')),
                '收入类型': income_type,
                '备注': '',
                '平台': '小台风',
                '账号': username
            }
            processed.append(processed_bill)

        # 更新账号信息
        if username in self.account_info:
            self.account_info[username]['recent_income'] = recent_income
            self.account_info[username]['recent_withdraw'] = recent_withdraw
            self.account_info[username]['recent_refund'] = recent_refund
            self.account_info[username]['total_bills'] += len(processed)

        return processed

    def get_account_info(self, username: str) -> Dict:
        """获取账号信息"""
        return self.account_info.get(username, {
            'balance': 0.0,
            'recent_income': 0.0,
            'recent_withdraw': 0.0,
            'recent_refund': 0.0,
            'total_bills': 0
        })


# ======================== 妙月平台客户端 ========================
class MiaoYueClient:
    def __init__(self, config: dict):
        self.cfg = config
        self.common_cfg = config['common']
        self.base_url = config['miaoyue']['base_url']
        self.ua = UserAgent()
        self.data_processor = DataProcessor()
        self.account_info = {}

    def login(self, username: str, password: str) -> tuple[Optional[str], str]:
        """登录妙月平台"""
        login_url = f"{self.base_url}/card/user/password/login"
        try:
            login_params = {"username": username, "password": password}
            response = requests.post(login_url, params=login_params,
                                     timeout=self.common_cfg['request_timeout'])
            response.raise_for_status()

            result = response.json()
            if result.get("success") and result.get("statusCode") == 0:
                token = result["object"]["token"]
                logger.info(f"妙月平台账号 {username} 登录成功")
                return token, ""
            else:
                error_msg = f"登录失败: {result.get('content', '未知错误')}"
                logger.error(f"妙月平台账号 {username} {error_msg}")
                return None, error_msg

        except requests.exceptions.Timeout:
            error_msg = "登录请求超时"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return None, error_msg
        except requests.exceptions.ConnectionError:
            error_msg = "无法连接到妙月平台"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return None, error_msg
        except Exception as e:
            error_msg = f"登录异常: {str(e)}"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return None, error_msg

    def get_balance(self, token: str, username: str) -> tuple[float, float, float, str]:
        """获取余额"""
        balance_url = f"{self.base_url}/card/proxy/company/capital/account/info?currencyType=CNY"
        try:
            headers = {"x-token": f'{{"token":"{token}"}}', "User-Agent": self.ua.random}
            response = requests.get(balance_url, headers=headers,
                                    timeout=self.common_cfg['request_timeout'])
            response.raise_for_status()

            balance_raw = response.json()
            if balance_raw.get("success") and balance_raw.get("statusCode") == 0:
                balance_info = balance_raw.get("object", {})
                withdrawable = self.data_processor.safe_float(balance_info.get("withdrawAmount", 0))
                non_withdrawable = self.data_processor.safe_float(balance_info.get("nonWithdrawAmount", 0))
                total = withdrawable + non_withdrawable

                logger.info(f"妙月平台账号 {username} 余额: {total:.2f} 元 (可提现: {withdrawable:.2f} 元)")

                self.account_info[username] = {
                    'balance': total,
                    'withdrawable': withdrawable,
                    'non_withdrawable': non_withdrawable,
                    'recent_income': 0.0,
                    'recent_withdraw': 0.0,
                    'recent_refund': 0.0,
                    'total_bills': 0
                }

                return total, withdrawable, non_withdrawable, ""
            else:
                error_msg = f"余额查询失败: {balance_raw.get('content', '未知错误')}"
                logger.error(f"妙月平台账号 {username} {error_msg}")
                return 0.0, 0.0, 0.0, error_msg

        except requests.exceptions.Timeout:
            error_msg = "余额查询超时"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return 0.0, 0.0, 0.0, error_msg
        except requests.exceptions.ConnectionError:
            error_msg = "余额查询连接失败"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return 0.0, 0.0, 0.0, error_msg
        except Exception as e:
            error_msg = f"余额查询异常: {str(e)}"
            logger.error(f"妙月平台账号 {username} {error_msg}")
            return 0.0, 0.0, 0.0, error_msg

    def get_bills(self, token: str, username: str) -> tuple[List[Dict], str]:
        """获取账单"""
        all_bills = []
        current = 1
        errors = []

        while True:
            try:
                bill_url = (f"{self.base_url}/card/proxy/user/bill/page?"
                            f"currency=CNY&billType=&orderNo=&cardValue=&"
                            f"orders[0].column=createTime&orders[0].asc=false&"
                            f"current={current}&size={self.common_cfg['bill_page_size']}")

                headers = {"x-token": f'{{"token":"{token}"}}', "User-Agent": self.ua.random}
                response = requests.get(bill_url, headers=headers,
                                        timeout=self.common_cfg['request_timeout'])
                response.raise_for_status()

                bill_raw = response.json()

                if bill_raw.get("success") and bill_raw.get("statusCode") == 0:
                    records = bill_raw.get("object", {}).get("records", [])
                    if records:
                        processed_bills = self._process_bills(records, username)
                        all_bills.extend(processed_bills)
                    else:
                        break
                else:
                    error_msg = f"账单查询失败: {bill_raw.get('content', '未知错误')}"
                    errors.append(error_msg)
                    break

            except requests.exceptions.Timeout:
                error_msg = f"第{current}页账单查询超时"
                errors.append(error_msg)
                break
            except requests.exceptions.ConnectionError:
                error_msg = f"第{current}页账单查询连接失败"
                errors.append(error_msg)
                break
            except Exception as e:
                error_msg = f"第{current}页账单查询异常: {str(e)}"
                errors.append(error_msg)
                break

            if not self.common_cfg['query_all_bills'] or current >= 10:
                break

            current += 1
            time.sleep(0.3)

        if all_bills:
            logger.info(f"妙月平台账号 {username} 获取到 {len(all_bills)} 条账单")

        error_info = "; ".join(errors) if errors else ""
        return all_bills, error_info

    def _process_bills(self, raw_bills: List[Dict], username: str) -> List[Dict]:
        """处理妙月平台账单数据"""
        processed = []
        recent_income = 0.0
        recent_withdraw = 0.0
        recent_refund = 0.0
        days_for_recent = self.common_cfg.get('days_for_recent', 30)
        cutoff_date = datetime.now() - timedelta(days=days_for_recent)

        for bill in raw_bills:
            trans_time_str = bill.get('createTime')
            trans_time = None
            if trans_time_str:
                try:
                    trans_time = datetime.strptime(str(trans_time_str), '%Y-%m-%d %H:%M:%S')
                except:
                    pass

            bill_amount = self.data_processor.safe_float(bill.get('billAmount', 0))
            bill_type = bill.get('billType', '')

            # 分类收入类型
            income_type = self.data_processor.classify_income_type(
                bill.get('remarks', ''),
                '',
                '',
                bill_type,
                bill_amount  # 传入金额用于判断是否为负数
            )

            # 统计最近收益
            if trans_time and trans_time >= cutoff_date:
                if bill_amount > 0 and income_type not in ["提现支出", "退款"]:
                    # 正数且不是提现或退款，计入收入
                    recent_income += bill_amount
                elif bill_amount < 0:
                    # 负数：判断是提现还是退款
                    if income_type == "提现支出":
                        recent_withdraw += abs(bill_amount)
                    elif income_type == "退款":
                        recent_refund += abs(bill_amount)

            order_content = ""
            order_no = bill.get('orderNo', '')
            if order_no:
                order_content = f"订单号: {order_no}"

            processed_bill = {
                '订单号': self.data_processor.safe_str(bill.get('orderNo', '')),
                'ICCID': self.data_processor.safe_str(bill.get('cardIccid', '')),
                '卡号': self.data_processor.safe_str(bill.get('cardNumber', '')),
                '交易时间': self.data_processor.standardize_datetime(trans_time_str),
                '售价（元）': 0.0,
                '成本（元）': 0.0,
                '佣金（元）': bill_amount,
                '客户名称': '',
                '套餐/产品名称': order_content,
                '运营商': '',
                '收入类型': income_type,
                '备注': self.data_processor.safe_str(bill.get('remarks', '')),
                '平台': '妙月',
                '账号': username,
                '账单类型': bill_type
            }
            processed.append(processed_bill)

        # 更新账号信息
        if username in self.account_info:
            self.account_info[username]['recent_income'] = recent_income
            self.account_info[username]['recent_withdraw'] = recent_withdraw
            self.account_info[username]['recent_refund'] = recent_refund
            self.account_info[username]['total_bills'] += len(processed)

        return processed

    def get_account_info(self, username: str) -> Dict:
        """获取账号信息"""
        return self.account_info.get(username, {
            'balance': 0.0,
            'withdrawable': 0.0,
            'non_withdrawable': 0.0,
            'recent_income': 0.0,
            'recent_withdraw': 0.0,
            'recent_refund': 0.0,
            'total_bills': 0
        })


# ======================== 多平台管理器和Excel导出 ========================
class MultiPlatformManager:
    """多平台管理器"""

    def __init__(self, config_path: str = "multi_platform_config.ini"):
        self.config_manager = ConfigManager(config_path)
        self.config = self.config_manager.load_config()
        self.all_bills = []
        self.platform_bills = {
            '天机': [],
            '小台风': [],
            '妙月': []
        }
        self.account_summary = []
        self.error_logs = []
        self.summary_data = {
            'total_accounts': 0,
            'success_accounts': 0,
            'failed_accounts': 0,
            'total_balance': 0.0,
            'total_recent_income': 0.0,
            'total_recent_withdraw': 0.0,
            'total_recent_refund': 0.0,
            'total_bills': 0
        }

    def get_output_path(self) -> Path:
        """获取输出路径"""
        output_path = self.config['common']['output_path']
        if output_path.lower() == '桌面':
            return Path.home() / 'Desktop'
        else:
            return Path(output_path)

    def log_error(self, platform: str, username: str, error_type: str, error_msg: str):
        """记录错误日志"""
        error_entry = {
            '时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '平台': platform,
            '账号': username,
            '错误类型': error_type,
            '错误信息': error_msg
        }
        self.error_logs.append(error_entry)

        print(f"{Fore.RED}❌ 【{platform}-{username}】{error_type}: {error_msg}")

    def query_tianji_accounts(self) -> Tuple[List[Dict], List[Dict]]:
        """查询天机平台所有账号"""
        print(f"\n{Fore.BLUE}【天机平台】查询开始")

        accounts = self.config['tianji']['accounts']
        all_bills = []
        account_info_list = []

        for username, password in accounts.items():
            try:
                self.summary_data['total_accounts'] += 1

                client = TianjiClient(self.config)
                session, login_ok, login_error = client.login(username, password)
                if not login_ok:
                    self.summary_data['failed_accounts'] += 1
                    self.log_error('天机', username, '登录失败', login_error)
                    continue

                balance, balance_error = client.get_balance(session, username)
                if balance_error:
                    self.log_error('天机', username, '余额查询失败', balance_error)
                else:
                    self.summary_data['total_balance'] += balance
                    print(f"{Fore.GREEN}【天机-{username}】余额: {balance:.2f}元")

                bills, bill_error = client.get_bills(session, username)
                if bill_error:
                    self.log_error('天机', username, '账单查询错误', bill_error)
                all_bills.extend(bills)

                account_info = client.get_account_info(username)
                account_info['平台'] = '天机'
                account_info['账号'] = username
                account_info_list.append(account_info)

                # 累加最近收益和提现
                self.summary_data['total_recent_income'] += account_info['recent_income']
                self.summary_data['total_recent_withdraw'] += account_info['recent_withdraw']
                self.summary_data['total_recent_refund'] += account_info['recent_refund']

                self.summary_data['success_accounts'] += 1

                time.sleep(self.config['common']['platform_delay'])

            except Exception as e:
                error_msg = f"查询异常: {str(e)}"
                self.log_error('天机', username, '系统异常', error_msg)
                self.summary_data['failed_accounts'] += 1

        print(f"{Fore.BLUE}【天机平台】查询完成，成功{self.summary_data['success_accounts']}个账号")
        return all_bills, account_info_list

    def query_xiaotaifeng_accounts(self) -> Tuple[List[Dict], List[Dict]]:
        """查询小台风平台所有账号"""
        print(f"\n{Fore.BLUE}【小台风平台】查询开始")

        accounts = self.config['xiaotaifeng']['accounts']
        all_bills = []
        account_info_list = []

        for account in accounts:
            username = account['username']
            password = account['password']

            try:
                self.summary_data['total_accounts'] += 1

                client = XiaoTaiFengClient(self.config)
                session, login_ok, login_error = client.login(username, password)
                if not login_ok:
                    self.summary_data['failed_accounts'] += 1
                    self.log_error('小台风', username, '登录失败', login_error)
                    continue

                balance, balance_error = client.get_balance(session, username)
                if balance_error:
                    self.log_error('小台风', username, '余额查询失败', balance_error)
                else:
                    self.summary_data['total_balance'] += balance
                    print(f"{Fore.GREEN}【小台风-{username}】余额: {balance:.2f}元")

                bills, bill_error = client.get_bills(session, username)
                if bill_error:
                    self.log_error('小台风', username, '账单查询错误', bill_error)
                all_bills.extend(bills)

                account_info = client.get_account_info(username)
                account_info['平台'] = '小台风'
                account_info['账号'] = username
                account_info_list.append(account_info)

                self.summary_data['total_recent_income'] += account_info['recent_income']
                self.summary_data['total_recent_withdraw'] += account_info['recent_withdraw']
                self.summary_data['total_recent_refund'] += account_info['recent_refund']

                self.summary_data['success_accounts'] += 1

                time.sleep(self.config['common']['platform_delay'])

            except Exception as e:
                error_msg = f"查询异常: {str(e)}"
                self.log_error('小台风', username, '系统异常', error_msg)
                self.summary_data['failed_accounts'] += 1

        print(f"{Fore.BLUE}【小台风平台】查询完成")
        return all_bills, account_info_list

    def query_miaoyue_accounts(self) -> Tuple[List[Dict], List[Dict]]:
        """查询妙月平台所有账号"""
        print(f"\n{Fore.BLUE}【妙月平台】查询开始")

        accounts = self.config['miaoyue']['accounts']
        all_bills = []
        account_info_list = []

        for username, password in accounts.items():
            try:
                self.summary_data['total_accounts'] += 1

                client = MiaoYueClient(self.config)
                token, login_error = client.login(username, password)
                if not token:
                    self.summary_data['failed_accounts'] += 1
                    self.log_error('妙月', username, '登录失败', login_error)
                    continue

                total, withdrawable, non_withdrawable, balance_error = client.get_balance(token, username)
                if balance_error:
                    self.log_error('妙月', username, '余额查询失败', balance_error)
                else:
                    self.summary_data['total_balance'] += total
                    print(f"{Fore.GREEN}【妙月-{username}】余额: {total:.2f}元（可提现 {withdrawable:.2f}元）")

                bills, bill_error = client.get_bills(token, username)
                if bill_error:
                    self.log_error('妙月', username, '账单查询错误', bill_error)
                all_bills.extend(bills)

                account_info = client.get_account_info(username)
                account_info['平台'] = '妙月'
                account_info['账号'] = username
                account_info['可提现余额'] = account_info.get('withdrawable', 0)
                account_info['不可提现余额'] = account_info.get('non_withdrawable', 0)
                account_info_list.append(account_info)

                self.summary_data['total_recent_income'] += account_info['recent_income']
                self.summary_data['total_recent_withdraw'] += account_info['recent_withdraw']
                self.summary_data['total_recent_refund'] += account_info['recent_refund']

                self.summary_data['success_accounts'] += 1

                time.sleep(self.config['common']['platform_delay'])

            except Exception as e:
                error_msg = f"查询异常: {str(e)}"
                self.log_error('妙月', username, '系统异常', error_msg)
                self.summary_data['failed_accounts'] += 1

        print(f"{Fore.BLUE}【妙月平台】查询完成")
        return all_bills, account_info_list

    def query_all_platforms(self):
        """查询所有平台"""
        start_time = time.time()

        print(f"\n{Fore.CYAN}⚡ 开始多平台查询，启用多线程: {self.config['common']['enable_threading']}")

        # 重置汇总数据
        self.summary_data = {
            'total_accounts': 0,
            'success_accounts': 0,
            'failed_accounts': 0,
            'total_balance': 0.0,
            'total_recent_income': 0.0,
            'total_recent_withdraw': 0.0,
            'total_recent_refund': 0.0,
            'total_bills': 0
        }

        if self.config['common']['enable_threading']:
            with concurrent.futures.ThreadPoolExecutor(
                    max_workers=min(3, self.config['common']['max_workers'])
            ) as executor:
                future_tianji = executor.submit(self.query_tianji_accounts)
                future_xiaotaifeng = executor.submit(self.query_xiaotaifeng_accounts)
                future_miaoyue = executor.submit(self.query_miaoyue_accounts)

                tianji_bills, tianji_accounts = future_tianji.result()
                xiaotaifeng_bills, xiaotaifeng_accounts = future_xiaotaifeng.result()
                miaoyue_bills, miaoyue_accounts = future_miaoyue.result()
        else:
            tianji_bills, tianji_accounts = self.query_tianji_accounts()
            xiaotaifeng_bills, xiaotaifeng_accounts = self.query_xiaotaifeng_accounts()
            miaoyue_bills, miaoyue_accounts = self.query_miaoyue_accounts()

        self.all_bills = tianji_bills + xiaotaifeng_bills + miaoyue_bills
        self.platform_bills['天机'] = tianji_bills
        self.platform_bills['小台风'] = xiaotaifeng_bills
        self.platform_bills['妙月'] = miaoyue_bills

        self.account_summary = tianji_accounts + xiaotaifeng_accounts + miaoyue_accounts
        self.summary_data['total_bills'] = len(self.all_bills)

        elapsed_time = time.time() - start_time

        # 计算净收益（总收入 - 总退款，不包括提现）
        net_income = self.summary_data['total_recent_income'] - self.summary_data['total_recent_refund']

        print(f"\n{Fore.GREEN}{'=' * 60}")
        print(f"{Fore.GREEN}查询完成！用时：{elapsed_time:.1f}秒")
        print(f"{Fore.GREEN}{'=' * 60}")
        print(f"📊 汇总信息：")
        print(f"   总账号数：{self.summary_data['total_accounts']}")
        print(f"   成功账号：{self.summary_data['success_accounts']}")
        print(f"   失败账号：{self.summary_data['failed_accounts']}")
        print(f"   总余额：{self.summary_data['total_balance']:.2f} 元")
        print(
            f"   最近{self.config['common'].get('days_for_recent', 30)}天总收益：{self.summary_data['total_recent_income']:.2f} 元")
        print(
            f"   最近{self.config['common'].get('days_for_recent', 30)}天总提现：{self.summary_data['total_recent_withdraw']:.2f} 元")
        print(
            f"   最近{self.config['common'].get('days_for_recent', 30)}天总退款：{self.summary_data['total_recent_refund']:.2f} 元")
        print(f"   净收益（总收益 - 总退款）：{net_income:.2f} 元")
        print(f"   总账单数：{self.summary_data['total_bills']}")
        print(f"{Fore.GREEN}{'=' * 60}")

    def export_to_excel(self):
        """导出数据到Excel"""
        if not self.all_bills and not self.account_summary:
            print(f"{Fore.YELLOW}⚠️  无任何数据，跳过导出")
            return

        output_path = self.get_output_path()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_path / f"多平台账单汇总_{timestamp}.xlsx"

        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 1. 多平台账单汇总表
        if self.all_bills:
            ws_summary = wb.create_sheet(title="多平台账单汇总")
            self._write_bills_to_sheet(ws_summary, self.all_bills, "多平台账单汇总", sort_by_time=True)

        # 2. 各平台单独工作表
        for platform in ['天机', '小台风', '妙月']:
            platform_bills = self.platform_bills[platform]
            if platform_bills:
                ws_platform = wb.create_sheet(title=f"{platform}账单")
                sort_needed = (platform == '天机')
                self._write_bills_to_sheet(ws_platform, platform_bills, f"{platform}平台账单", sort_by_time=sort_needed)

        # 3. 平台账号汇总表
        if self.account_summary:
            ws_accounts = wb.create_sheet(title="平台账号汇总")
            self._write_account_summary_to_sheet(ws_accounts)

        # 4. 数据统计表
        ws_stats = wb.create_sheet(title="数据统计")
        self._write_statistics_to_sheet(ws_stats)

        # 5. 错误日志表
        if self.error_logs:
            ws_errors = wb.create_sheet(title="错误日志")
            self._write_error_logs_to_sheet(ws_errors)

        wb.save(excel_file)
        print(f"\n{Fore.GREEN}✅ Excel文件已保存：{excel_file}")

        return excel_file

    def _write_bills_to_sheet(self, ws, bills_data, sheet_title, sort_by_time=False):
        """将账单数据写入工作表"""
        if not bills_data:
            ws.append(["无数据"])
            return

        df = pd.DataFrame(bills_data)

        if '交易时间' in df.columns and sort_by_time:
            df['交易时间_temp'] = pd.to_datetime(df['交易时间'], errors='coerce', format='%Y-%m-%d %H:%M:%S')
            df = df.dropna(subset=['交易时间_temp'])
            df = df.sort_values('交易时间_temp', ascending=False)
            df = df.drop('交易时间_temp', axis=1)

        column_order = ['订单号', 'ICCID', '卡号', '交易时间', '售价（元）', '成本（元）',
                        '佣金（元）', '客户名称', '套餐/产品名称', '运营商', '收入类型',
                        '备注', '平台', '账号']

        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment_num = Alignment(horizontal='right', vertical='center')
        data_alignment_str = Alignment(horizontal='left', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        column_widths = {
            '订单号': 18,
            'ICCID': 20,
            '卡号': 15,
            '交易时间': 20,
            '售价（元）': 10,
            '成本（元）': 10,
            '佣金（元）': 10,
            '客户名称': 20,
            '套餐/产品名称': 25,
            '运营商': 10,
            '收入类型': 10,
            '备注': 15,
            '平台': 10,
            '账号': 15
        }

        for col_idx, column in enumerate(existing_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            width = column_widths.get(column, 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, row in df.iterrows():
            for col_idx, column in enumerate(existing_columns, 1):
                cell_value = row[column]

                if cell_value == "" or pd.isna(cell_value):
                    cell_value = None

                cell = ws.cell(row=row_idx + 2, column=col_idx, value=cell_value)
                cell.font = data_font
                cell.border = thin_border

                if '（元）' in column or column in ['售价（元）', '成本（元）', '佣金（元）']:
                    cell.alignment = data_alignment_num
                    if isinstance(cell_value, (int, float)):
                        cell.number_format = '0.00'
                elif column in ['平台', '账号', '收入类型']:
                    cell.alignment = data_alignment_center
                else:
                    cell.alignment = data_alignment_str

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(existing_columns))
        title_cell = ws.cell(row=1, column=1, value=sheet_title)
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

        ws.freeze_panes = ws['A3']

    def _write_account_summary_to_sheet(self, ws):
        """写入平台账号汇总表"""
        # 更新表头，增加提现和退款统计
        headers = ['平台', '账号', '总余额（元）', '可提现余额（元）', '不可提现余额（元）',
                   f'最近{self.config["common"].get("days_for_recent", 30)}天收益（元）',
                   f'最近{self.config["common"].get("days_for_recent", 30)}天提现（元）',
                   f'最近{self.config["common"].get("days_for_recent", 30)}天退款（元）',
                   '总账单数']

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment_num = Alignment(horizontal='right', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        column_widths = [10, 20, 15, 15, 15, 18, 18, 18, 12]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row_idx = 3
        for account_info in self.account_summary:
            platform = account_info.get('平台', '')
            username = account_info.get('账号', '')
            balance = account_info.get('balance', 0)
            withdrawable = account_info.get('withdrawable', account_info.get('balance', 0))
            non_withdrawable = account_info.get('non_withdrawable', 0)
            recent_income = account_info.get('recent_income', 0)
            recent_withdraw = account_info.get('recent_withdraw', 0)
            recent_refund = account_info.get('recent_refund', 0)
            total_bills = account_info.get('total_bills', 0)

            data_row = [platform, username, balance, withdrawable, non_withdrawable,
                        recent_income, recent_withdraw, recent_refund, total_bills]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border

                if col_idx in [1, 2, 9]:  # 平台、账号、总账单数居中
                    cell.alignment = data_alignment_center
                elif col_idx in [3, 4, 5, 6, 7, 8]:  # 金额右对齐
                    cell.alignment = data_alignment_num
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                else:
                    cell.alignment = data_alignment_center

            row_idx += 1

        if self.account_summary:
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="总计").font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value="").font = Font(bold=True)

            # 计算各列总计
            totals = [
                sum(acc.get('balance', 0) for acc in self.account_summary),
                sum(acc.get('withdrawable', acc.get('balance', 0)) for acc in self.account_summary),
                sum(acc.get('non_withdrawable', 0) for acc in self.account_summary),
                sum(acc.get('recent_income', 0) for acc in self.account_summary),
                sum(acc.get('recent_withdraw', 0) for acc in self.account_summary),
                sum(acc.get('recent_refund', 0) for acc in self.account_summary),
                sum(acc.get('total_bills', 0) for acc in self.account_summary)
            ]

            for col_idx, total in enumerate(totals, 3):
                cell = ws.cell(row=row_idx, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.alignment = data_alignment_num
                cell.number_format = '0.00'

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                cell.font = Font(bold=True)
                if col_idx in [3, 4, 5, 6, 7, 8]:
                    cell.alignment = data_alignment_num
                    if col_idx in [3, 4, 5, 6, 7, 8]:
                        cell.number_format = '0.00'

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="平台账号汇总")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')

        ws.freeze_panes = ws['A3']

    def _write_statistics_to_sheet(self, ws):
        """写入数据统计表"""
        # 计算净收益
        net_income = self.summary_data['total_recent_income'] - self.summary_data['total_recent_refund']

        stats = [
            ["统计项目", "数值", "说明"],
            ["总账号数", self.summary_data['total_accounts'], "配置文件中所有平台账号总数"],
            ["成功查询账号", self.summary_data['success_accounts'], "成功登录并获取数据的账号数"],
            ["查询失败账号", self.summary_data['failed_accounts'], "登录失败或查询异常的账号数"],
            ["总余额（元）", self.summary_data['total_balance'], "所有账号余额总和"],
            [f"最近{self.config['common'].get('days_for_recent', 30)}天总收益（元）",
             self.summary_data['total_recent_income'],
             f"最近{self.config['common'].get('days_for_recent', 30)}天的正数收益总和"],
            [f"最近{self.config['common'].get('days_for_recent', 30)}天总提现（元）",
             self.summary_data['total_recent_withdraw'],
             f"最近{self.config['common'].get('days_for_recent', 30)}天的提现支出总和"],
            [f"最近{self.config['common'].get('days_for_recent', 30)}天总退款（元）",
             self.summary_data['total_recent_refund'],
             f"最近{self.config['common'].get('days_for_recent', 30)}天的退款总和"],
            ["净收益（总收益 - 总退款）（元）", net_income, "实际净收益（总收益减去退款）"],
            ["总账单数", self.summary_data['total_bills'], "所有账单记录总数"],
            ["", "", ""],
            ["平台", "账号数", "总余额（元）", "最近收益（元）", "最近提现（元）", "最近退款（元）", "净收益（元）", "账单数"]
        ]

        # 添加各平台统计
        for platform in ['天机', '小台风', '妙月']:
            platform_accounts = [acc for acc in self.account_summary if acc['平台'] == platform]
            platform_balance = sum(acc.get('balance', 0) for acc in platform_accounts)
            platform_recent_income = sum(acc.get('recent_income', 0) for acc in platform_accounts)
            platform_recent_withdraw = sum(acc.get('recent_withdraw', 0) for acc in platform_accounts)
            platform_recent_refund = sum(acc.get('recent_refund', 0) for acc in platform_accounts)
            platform_net_income = platform_recent_income - platform_recent_refund
            platform_bills_count = len(self.platform_bills[platform])

            stats.append(
                [platform, len(platform_accounts), platform_balance,
                 platform_recent_income, platform_recent_withdraw, platform_recent_refund,
                 platform_net_income, platform_bills_count])

        for row_idx, row in enumerate(stats, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if row_idx == 1 or row_idx == 12:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx <= 11:
                    if col_idx == 2 and isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    if col_idx in [3, 4, 5, 6, 7] and isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        column_widths = [20, 15, 30, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        title_cell = ws.cell(row=1, column=1, value="数据统计")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

    def _write_error_logs_to_sheet(self, ws):
        """写入错误日志表"""
        headers = ['时间', '平台', '账号', '错误类型', '错误信息']

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        column_widths = [18, 10, 15, 15, 50]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, error in enumerate(self.error_logs, 2):
            data_row = [
                error.get('时间', ''),
                error.get('平台', ''),
                error.get('账号', ''),
                error.get('错误类型', ''),
                error.get('错误信息', '')
            ]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="错误日志")
        title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

        ws.freeze_panes = ws['A3']

    def run(self):
        """运行主程序"""
        print(f"{Fore.CYAN}{'=' * 60}")
        print(f"{Fore.CYAN}多平台账单查询系统")
        print(f"{Fore.CYAN}启动时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{Fore.CYAN}{'=' * 60}")

        try:
            self.query_all_platforms()

            if self.all_bills or self.account_summary:
                excel_file = self.export_to_excel()

                print(f"\n{Fore.GREEN}{'=' * 60}")
                print(f"{Fore.GREEN}导出文件包含以下工作表：")
                print(f"{Fore.GREEN}1. 多平台账单汇总 - 所有平台的账单数据（按时间排序）")
                print(f"{Fore.GREEN}2. 天机账单 - 天机平台的账单数据（按时间排序）")
                print(f"{Fore.GREEN}3. 小台风账单 - 小台风平台的账单数据")
                print(f"{Fore.GREEN}4. 妙月账单 - 妙月平台的账单数据")
                print(f"{Fore.GREEN}5. 平台账号汇总 - 各账号余额、收益、提现、退款统计")
                print(f"{Fore.GREEN}6. 数据统计 - 整体统计信息（包含净收益计算）")
                if self.error_logs:
                    print(f"{Fore.GREEN}7. 错误日志 - 所有错误记录")
                print(f"{Fore.GREEN}{'=' * 60}")
            else:
                print(f"{Fore.YELLOW}⚠️  未获取到任何数据")

        except Exception as e:
            logger.error(f"程序运行异常: {str(e)}", exc_info=True)
            print(f"{Fore.RED}❌ 程序运行异常：{str(e)}")
            print(f"{Fore.RED}详细错误信息请查看日志文件：{log_file}")

        finally:
            print(f"\n{Fore.CYAN}{'=' * 60}")
            print(f"{Fore.CYAN}程序执行完成")
            print(f"{Fore.CYAN}详细日志请查看: {log_file}")
            print(f"{Fore.CYAN}{'=' * 60}")


# ======================== 主程序入口 ========================
def main():
    """主函数"""
    try:
        from fake_useragent import UserAgent
    except ImportError:
        print(f"{Fore.YELLOW}⚠️  缺少依赖库: fake_useragent")
        print(f"{Fore.YELLOW}正在安装依赖库...")
        import subprocess
        import sys

        packages = ['fake_useragent', 'pandas', 'openpyxl', 'colorama', 'requests', 'beautifulsoup4']
        for package in packages:
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            except:
                pass

        print(f"{Fore.GREEN}✅ 依赖安装完成，请重新运行程序")
        return

    manager = MultiPlatformManager()
    manager.run()


if __name__ == "__main__":
    main()